from __future__ import annotations

import os
import json
from pathlib import Path

import pandas as pd
import polars as pl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from lib.catalog_report_filter import filter_reportable_catalog
except ModuleNotFoundError:
    from scripts.lib.catalog_report_filter import filter_reportable_catalog


BASE = Path(r"C:\royalties_pipeline")
MARTS_DIR = BASE / "warehouse" / "marts"
REPORTS_DIR = BASE / "reports"
STANDARDIZED_ALL_PATH = MARTS_DIR / "standardized_raw_all_sources.parquet"
STATEMENT_SUMMARY_PATH = MARTS_DIR / "statement_summary_all_sources.parquet"
CATALOG_RELEASE_METADATA_PATH = MARTS_DIR / "catalog_release_metadata.parquet"
REGISTRY_DIR = BASE / "warehouse" / "registry"
STATEMENT_POLICY_PATH = REGISTRY_DIR / "distributor_account_policies.json"
CONTRACT_CUTOFFS_PATH = REGISTRY_DIR / "contract_cutoffs.json"

FUGA_STATEMENT_FACTOR = 0.977832
GUSTY_VARIANT_SOURCE = "onerpm"
GUSTY_VARIANT_BASE_ACCOUNT = "gusty_dj"
GUSTY_VARIANT_ACCOUNT = "gusty_dj_post_motorcito"
GUSTY_VARIANT_TITLE = "ONErpm Gusty DJ - post Motorcito"
NEW_REPORT_ONERPM_ACCOUNTS = {"henry_remix"}
LA_NUEVA_SANGRE_REFERENCE_TERMS = [
    "ni ahi",
    "ni ah",
    "ni aca",
    "ni ac",
    "ni alla",
    "ni all",
    "ni aya",
]
LA_NUEVA_SANGRE_CONTRACT_CUTOFF_DATE = "2023-06-15"
def use_release_metadata_in_statement() -> bool:
    return (
        os.getenv("VPO_USE_RELEASE_METADATA_IN_STATEMENT", "").strip().lower()
        in {"1", "true", "yes", "on"}
    )
NEW_REPORT_VARIANT_TITLES = {
    ("onerpm", "gusty_dj"): "ONErpm Gusty DJ",
    ("onerpm", "la_nueva_sangre"): "ONErpm La Nueva Sangre",
}

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ARTIST_FILL = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
TITLE_FILL = PatternFill(start_color="EAF2F8", end_color="EAF2F8", fill_type="solid")
SHARE_ALERT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
CONFIG_FILL = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")

TITLE_FONT = Font(size=16, bold=True, color="1F4E78")
TOTAL_FONT = Font(bold=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")


def has_col(schema: dict[str, pl.DataType], col: str) -> bool:
    return col in schema


def col_or_null(schema: dict[str, pl.DataType], col: str, dtype=pl.Utf8) -> pl.Expr:
    if has_col(schema, col):
        return pl.col(col).cast(dtype, strict=False)
    return pl.lit(None).cast(dtype)


def amount_expr(schema: dict[str, pl.DataType]) -> pl.Expr:
    candidates = []
    for col in ["amount_usd", "net_amount_usd", "net_amount"]:
        if has_col(schema, col):
            candidates.append(pl.col(col).cast(pl.Float64, strict=False))

    if not candidates:
        return pl.lit(0.0)

    return pl.coalesce(candidates)


def format_sheet(ws, pivot, company_name, share_flags=None):
    max_col = len(pivot.columns) + 1
    max_row = len(pivot) + 2

    for col in range(2, max_col + 1):
        ws.cell(row=1, column=col).value = None

    ws["A1"] = company_name
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = LEFT_ALIGN
    ws.row_dimensions[1].height = 24

    ws["A2"] = "ARTISTA"
    ws["A2"].fill = HEADER_FILL
    ws["A2"].font = HEADER_FONT
    ws["A2"].alignment = CENTER_ALIGN

    for col_idx, col_name in enumerate(pivot.columns, start=2):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = col_name
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN

    for row in range(3, max_row + 1):
        label = str(ws.cell(row=row, column=1).value or "")
        artist_cell = ws.cell(row=row, column=1)
        artist_cell.fill = ARTIST_FILL
        artist_cell.alignment = LEFT_ALIGN
        is_total_row = label.upper().startswith("TOTAL")

        if is_total_row:
            artist_cell.fill = TOTAL_FILL
            artist_cell.font = TOTAL_FONT

        for col in range(2, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.number_format = '#,##0.00'
            cell.alignment = CENTER_ALIGN

            if is_total_row:
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT

    if share_flags is not None and not share_flags.empty:
        for artist_idx, artist_name in enumerate(pivot.index, start=3):
            if str(artist_name).upper().startswith("TOTAL"):
                continue

            if artist_name not in share_flags.index:
                continue

            for col_idx, period in enumerate(pivot.columns, start=2):
                if period == "TOTAL" or period not in share_flags.columns:
                    continue

                try:
                    has_alert = int(share_flags.loc[artist_name, period]) == 1
                except Exception:
                    has_alert = False

                if has_alert:
                    ws.cell(row=artist_idx, column=col_idx).fill = SHARE_ALERT_FILL

    total_col = max_col
    for row in range(2, max_row + 1):
        cell = ws.cell(row=row, column=total_col)
        if row == 2:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        else:
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT

    ws.column_dimensions["A"].width = 30
    for col in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A2:{get_column_letter(max_col)}{max_row}"


def add_totals_and_clean(pivot, include_zero_total_artists: bool = False):
    pivot = pivot.reindex(sorted(pivot.columns), axis=1)
    pivot["TOTAL"] = pivot.sum(axis=1)
    if not include_zero_total_artists:
        pivot = pivot[pivot["TOTAL"].round(2) != 0]
    return pivot


def scope_key(source, account) -> str:
    return f"{source}_{account}"


def add_config_sheet(
    writer,
    scopes: list[tuple[str, str]],
    min_artist_total_usd: float = 0.0,
    include_zero_total_artists: bool = False,
    default_excluded_scopes: set[str] | None = None,
) -> dict[str, int]:
    ws = writer.book.create_sheet("Config", 0)
    headers = ["Distribuidora/Cuenta", "Incluir en TOTAL", "Umbral aplicado USD", "Incluir artistas en cero"]
    ws.append(headers)
    default_excluded_scopes = default_excluded_scopes or set()

    scope_rows = {}
    for row_idx, (source, account) in enumerate(scopes, start=2):
        key = scope_key(source, account)
        ws.cell(row=row_idx, column=1).value = key
        ws.cell(row=row_idx, column=2).value = "NO" if key in default_excluded_scopes else "SI"
        ws.cell(row=row_idx, column=3).value = float(min_artist_total_usd)
        ws.cell(row=row_idx, column=4).value = "SI" if include_zero_total_artists else "NO"
        scope_rows[key] = row_idx

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN

    for row in range(2, len(scopes) + 2):
        ws.cell(row=row, column=1).fill = CONFIG_FILL
        ws.cell(row=row, column=2).alignment = CENTER_ALIGN
        ws.cell(row=row, column=3).alignment = CENTER_ALIGN
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=4).alignment = CENTER_ALIGN

    validation = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False)
    ws.add_data_validation(validation)
    validation.add(f"B2:B{len(scopes) + 1}")
    zero_validation = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False)
    ws.add_data_validation(zero_validation)
    zero_validation.add(f"D2:D{len(scopes) + 1}")

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.freeze_panes = "A2"
    return scope_rows


def enable_formula_recalculation(workbook) -> None:
    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
    except Exception:
        pass


def add_total_data_sheet(writer, total_source: pd.DataFrame, config_rows: int) -> int:
    detail = (
        total_source
        .groupby(["artist", "_scope_key", "statement_period"], dropna=False, as_index=False)["total"]
        .sum()
    )
    detail = detail.sort_values(["artist", "_scope_key", "statement_period"])
    detail["include_flag"] = 0

    sheet_name = "_TOTAL_DATA"
    detail.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]

    last_row = len(detail) + 1
    for row in range(2, last_row + 1):
        ws.cell(row=row, column=5).value = (
            f'=IFERROR(--(VLOOKUP(B{row},Config!$A$2:$B${config_rows},2,FALSE)="SI"),0)'
        )

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.sheet_state = "hidden"
    return last_row


def apply_total_config_formulas(ws, pivot_total, total_data_last_row: int) -> None:
    first_amount_col = 2
    last_col = len(pivot_total.columns) + 1
    total_col = last_col
    total_row = 2 + len(pivot_total)

    data_sheet = "'_TOTAL_DATA'"
    amount_range = f"{data_sheet}!$D$2:$D${total_data_last_row}"
    artist_range = f"{data_sheet}!$A$2:$A${total_data_last_row}"
    period_range = f"{data_sheet}!$C$2:$C${total_data_last_row}"
    include_range = f"{data_sheet}!$E$2:$E${total_data_last_row}"

    for row in range(3, total_row):
        for col in range(first_amount_col, total_col):
            letter = get_column_letter(col)
            ws.cell(row=row, column=col).value = (
                f"=SUMIFS({amount_range},{artist_range},$A{row},"
                f"{period_range},{letter}$2,{include_range},1)"
            )

        first_letter = get_column_letter(first_amount_col)
        previous_letter = get_column_letter(total_col - 1)
        ws.cell(row=row, column=total_col).value = f"=SUM({first_letter}{row}:{previous_letter}{row})"

    for col in range(first_amount_col, total_col + 1):
        letter = get_column_letter(col)
        ws.cell(row=total_row, column=col).value = f"=SUM({letter}3:{letter}{total_row - 1})"


def hide_initial_zero_total_rows(
    ws,
    pivot_total: pd.DataFrame,
    total_source: pd.DataFrame,
    scopes: list[tuple[str, str]],
    include_zero_total_artists: bool = False,
    default_excluded_scopes: set[str] | None = None,
) -> None:
    if include_zero_total_artists:
        return

    default_excluded_scopes = default_excluded_scopes or set()
    default_included_scopes = {
        scope_key(source, account)
        for source, account in scopes
        if scope_key(source, account) not in default_excluded_scopes
    }

    included_totals = (
        total_source.loc[total_source["_scope_key"].isin(default_included_scopes)]
        .groupby("artist", dropna=False)["total"]
        .sum()
    )

    for row_idx, artist_name in enumerate(pivot_total.index, start=3):
        if str(artist_name).upper().startswith("TOTAL"):
            continue

        total = float(included_totals.get(artist_name, 0.0))
        if round(abs(total), 2) == 0:
            ws.row_dimensions[row_idx].hidden = True


def clean_text_expr(expr: pl.Expr) -> pl.Expr:
    return expr.cast(pl.Utf8, strict=False).str.strip_chars()


def first_available_text(schema: dict[str, pl.DataType], columns: list[str]) -> pl.Expr:
    candidates = [
        clean_text_expr(pl.col(col))
        for col in columns
        if has_col(schema, col)
    ]
    if not candidates:
        return pl.lit(None).cast(pl.Utf8)

    return pl.coalesce(candidates)


def non_empty(expr: pl.Expr) -> pl.Expr:
    return expr.is_not_null() & (expr != "")


def gusty_track_key_expr(schema: dict[str, pl.DataType]) -> pl.Expr:
    isrc = first_available_text(schema, ["asset_isrc", "ISRC"])
    track_id = first_available_text(
        schema,
        ["label_track_id", "Label Track ID", "Track ID", "Video ID", "ID", "Parent ID"],
    )
    title = first_available_text(
        schema,
        [
            "track_statement_style",
            "asset_title_statement",
            "Track Title",
            "TRACK",
            "Asset Title",
            "Album Title",
            "album_statement_style",
            "Title",
            "Video Title",
            "Product Title",
            "YouTube Video Title",
        ],
    )
    artist = first_available_text(
        schema,
        [
            "artist_statement_style",
            "asset_artist_statement",
            "track_artist_statement",
            "Track Artists",
            "Artist Name",
            "Channel Name",
            "Product Artist",
        ],
    )

    return (
        pl.when(non_empty(isrc))
        .then(pl.concat_str([pl.lit("isrc:"), isrc.str.to_uppercase()]))
        .when(non_empty(track_id))
        .then(pl.concat_str([pl.lit("track_id:"), track_id.str.to_lowercase()]))
        .when(non_empty(title))
        .then(pl.concat_str([
            pl.lit("title_artist:"),
            title.str.to_lowercase(),
            pl.lit("|"),
            artist.fill_null("").str.to_lowercase(),
        ]))
        .otherwise(pl.lit(None).cast(pl.Utf8))
    )


def legacy_gusty_track_key_expr(schema: dict[str, pl.DataType]) -> pl.Expr:
    isrc = first_available_text(schema, ["asset_isrc", "ISRC"])
    track_id = first_available_text(schema, ["label_track_id", "Label Track ID", "Track ID"])
    title = first_available_text(
        schema,
        ["track_statement_style", "Track Title", "TRACK", "Album Title", "album_statement_style", "Title"],
    )
    artist = first_available_text(
        schema,
        ["artist_statement_style", "track_artist_statement", "Track Artists", "Artist Name"],
    )

    return (
        pl.when(non_empty(isrc))
        .then(pl.concat_str([pl.lit("isrc:"), isrc.str.to_uppercase()]))
        .when(non_empty(track_id))
        .then(pl.concat_str([pl.lit("track_id:"), track_id.str.to_lowercase()]))
        .when(non_empty(title))
        .then(pl.concat_str([
            pl.lit("title_artist:"),
            title.str.to_lowercase(),
            pl.lit("|"),
            artist.fill_null("").str.to_lowercase(),
        ]))
        .otherwise(pl.lit(None).cast(pl.Utf8))
    )


def title_match_expr(schema: dict[str, pl.DataType], terms: list[str]) -> pl.Expr:
    columns = [
        "track_statement_style",
        "asset_title_statement",
        "Track Title",
        "TRACK",
        "Asset Title",
        "Album Title",
        "album_statement_style",
        "Title",
        "Video Title",
        "Product Title",
        "YouTube Video Title",
    ]
    candidates = [
        clean_text_expr(pl.col(col)).str.to_lowercase().str.contains(term, literal=True)
        for col in columns
        if has_col(schema, col)
        for term in terms
    ]
    if not candidates:
        return pl.lit(False)

    return pl.any_horizontal(candidates)


def legacy_title_match_expr(schema: dict[str, pl.DataType], terms: list[str]) -> pl.Expr:
    columns = [
        "track_statement_style",
        "Track Title",
        "TRACK",
        "Album Title",
        "album_statement_style",
        "Title",
    ]
    candidates = [
        clean_text_expr(pl.col(col)).str.to_lowercase().str.contains(term, literal=True)
        for col in columns
        if has_col(schema, col)
        for term in terms
    ]
    if not candidates:
        return pl.lit(False)

    return pl.any_horizontal(candidates)


def gusty_motorcito_match_expr(schema: dict[str, pl.DataType]) -> pl.Expr:
    return title_match_expr(schema, ["motorcito"])


def build_post_reference_totals(
    standardized_path: Path | None,
    source: str,
    base_account: str,
    output_account: str,
    reference_terms: list[str],
    fallback_cutoff_month: str | None = None,
    reference_source: str | None = None,
    reference_account: str | None = None,
    cutoff_basis: str = "transaction_month",
    track_key_builder=gusty_track_key_expr,
    title_match_builder=title_match_expr,
) -> pd.DataFrame:
    if standardized_path is None or not standardized_path.exists():
        return pd.DataFrame(columns=["source", "account", "artist", "statement_period", "total", "has_share_in_out"])

    lf = pl.scan_parquet(standardized_path)
    schema = lf.collect_schema()

    has_share_expr = (
        pl.col("has_share_in_out").cast(pl.Boolean, strict=False)
        if has_col(schema, "has_share_in_out")
        else pl.lit(False)
    )

    base = (
        lf
        .filter(
            (pl.col("source") == source)
            & (pl.col("account") == base_account)
        )
        .pipe(lambda frame: filter_reportable_catalog(frame, set(schema.names())))
        .with_columns([
            col_or_null(schema, "statement_period").alias("_statement_period"),
            col_or_null(schema, "transaction_month").alias("_transaction_month"),
            col_or_null(schema, "artist_statement_style").alias("_artist_raw"),
            amount_expr(schema).alias("_amount"),
            has_share_expr.alias("_has_share_in_out"),
            track_key_builder(schema).alias("_track_key"),
            title_match_builder(schema, reference_terms).alias("_is_reference_track"),
        ])
        .filter(
            pl.col("_statement_period").is_not_null()
            & (pl.col("_statement_period").str.strip_chars() != "")
        )
    )

    reference_source = reference_source or source
    reference_account = reference_account or base_account
    cutoff_col = "_transaction_month" if cutoff_basis == "transaction_month" else "_statement_period"
    reference_base = (
        lf
        .filter(
            (pl.col("source") == reference_source)
            & (pl.col("account") == reference_account)
        )
        .with_columns([
            col_or_null(schema, "statement_period").alias("_statement_period"),
            col_or_null(schema, "transaction_month").alias("_transaction_month"),
            title_match_builder(schema, reference_terms).alias("_is_reference_track"),
        ])
    )

    reference_periods = (
        reference_base
        .filter(pl.col("_is_reference_track"))
        .select([
            pl.min(cutoff_col).alias("first_cutoff_month"),
        ])
        .collect()
    )

    first_cutoff_month = None
    if not reference_periods.is_empty():
        first_cutoff_month = reference_periods["first_cutoff_month"][0]

    if not first_cutoff_month:
        first_cutoff_month = fallback_cutoff_month

    cutoff_month = first_cutoff_month
    if not cutoff_month:
        return pd.DataFrame(columns=["source", "account", "artist", "statement_period", "total", "has_share_in_out"])

    old_track_keys = (
        base
        .filter(
            pl.col("_track_key").is_not_null()
            & pl.col(cutoff_col).is_not_null()
        )
        .group_by("_track_key")
        .agg(pl.min(cutoff_col).alias("_first_cutoff_month"))
        .filter(pl.col("_first_cutoff_month") < cutoff_month)
        .select("_track_key")
        .collect()
        .get_column("_track_key")
        .to_list()
    )

    return (
        base
        .filter(
            pl.col("_track_key").is_null()
            | (~pl.col("_track_key").is_in(old_track_keys))
        )
        .with_columns(
            pl.when(
                pl.col("_artist_raw").is_null()
                | (pl.col("_artist_raw").str.strip_chars() == "")
            )
            .then(pl.lit("SIN ARTISTA"))
            .otherwise(pl.col("_artist_raw").str.strip_chars())
            .alias("artist")
        )
        .group_by(["artist", "_statement_period"])
        .agg([
            pl.sum("_amount").round(2).alias("total"),
            pl.max("_has_share_in_out").cast(pl.Int64).alias("has_share_in_out"),
        ])
        .with_columns([
            pl.lit(source).alias("source"),
            pl.lit(output_account).alias("account"),
        ])
        .rename({"_statement_period": "statement_period"})
        .select(["source", "account", "artist", "statement_period", "total", "has_share_in_out"])
        .collect()
        .to_pandas()
    )


def build_source_sheet_totals(
    standardized_path: Path | None,
    source: str,
    account: str,
    source_sheet: str,
) -> pd.DataFrame:
    if standardized_path is None or not standardized_path.exists():
        return pd.DataFrame(columns=["source", "account", "artist", "statement_period", "total", "has_share_in_out"])

    lf = pl.scan_parquet(standardized_path)
    schema = lf.collect_schema()

    has_share_expr = (
        pl.col("has_share_in_out").cast(pl.Boolean, strict=False)
        if has_col(schema, "has_share_in_out")
        else pl.lit(False)
    )

    return (
        lf
        .filter(
            (pl.col("source") == source)
            & (pl.col("account") == account)
            & (col_or_null(schema, "source_sheet") == source_sheet)
        )
        .pipe(lambda frame: filter_reportable_catalog(frame, set(schema.names())))
        .with_columns([
            col_or_null(schema, "statement_period").alias("_statement_period"),
            col_or_null(schema, "artist_statement_style").alias("_artist_raw"),
            amount_expr(schema).alias("_amount"),
            has_share_expr.alias("_has_share_in_out"),
        ])
        .filter(
            pl.col("_statement_period").is_not_null()
            & (pl.col("_statement_period").str.strip_chars() != "")
        )
        .with_columns(
            pl.when(
                pl.col("_artist_raw").is_null()
                | (pl.col("_artist_raw").str.strip_chars() == "")
            )
            .then(pl.lit("SIN ARTISTA"))
            .otherwise(pl.col("_artist_raw").str.strip_chars())
            .alias("artist")
        )
        .group_by(["artist", "_statement_period"])
        .agg([
            pl.sum("_amount").round(2).alias("total"),
            pl.max("_has_share_in_out").cast(pl.Int64).alias("has_share_in_out"),
        ])
        .with_columns([
            pl.lit(source).alias("source"),
            pl.lit(account).alias("account"),
        ])
        .rename({"_statement_period": "statement_period"})
        .select(["source", "account", "artist", "statement_period", "total", "has_share_in_out"])
        .collect()
        .to_pandas()
    )


def release_metadata_lookup_key_expr(schema: dict[str, pl.DataType]) -> pl.Expr:
    source_sheet = col_or_null(schema, "source_sheet")
    isrc = first_available_text(schema, ["ISRC", "asset_isrc"])
    video_id = first_available_text(schema, ["Video ID", "ID", "Parent ID"])

    return (
        pl.when((source_sheet == "Masters") & non_empty(isrc))
        .then(pl.concat_str([pl.lit("isrc:"), isrc.str.to_uppercase()]))
        .when((source_sheet == "Youtube Channels") & non_empty(video_id))
        .then(pl.concat_str([pl.lit("video:"), video_id]))
        .otherwise(pl.lit(None).cast(pl.Utf8))
    )


def release_metadata_provider_expr(schema: dict[str, pl.DataType]) -> pl.Expr:
    source_sheet = col_or_null(schema, "source_sheet")
    return (
        pl.when(source_sheet == "Masters")
        .then(pl.lit("spotify"))
        .when(source_sheet == "Youtube Channels")
        .then(pl.lit("youtube"))
        .otherwise(pl.lit(None).cast(pl.Utf8))
    )


def build_release_metadata_totals(
    standardized_path: Path | None,
    *,
    source: str,
    account: str,
    output_account: str,
    metadata_path: Path = CATALOG_RELEASE_METADATA_PATH,
) -> pd.DataFrame:
    empty = pd.DataFrame(columns=["source", "account", "artist", "statement_period", "total", "has_share_in_out"])
    if not use_release_metadata_in_statement():
        return empty
    if (
        standardized_path is None
        or not standardized_path.exists()
        or not metadata_path.exists()
    ):
        return empty

    metadata = pl.scan_parquet(metadata_path)
    metadata_schema = metadata.collect_schema()
    required_metadata_cols = {
        "source",
        "account",
        "preferred_provider",
        "preferred_lookup_key",
        "release_date",
        "include_after_release_cutoff",
    }
    if not required_metadata_cols.issubset(set(metadata_schema.names())):
        return empty

    metadata_ready = (
        metadata
        .filter(
            (pl.col("source") == source)
            & (pl.col("account") == account)
            & pl.col("release_date").is_not_null()
        )
        .select([
            pl.col("preferred_provider").alias("_metadata_provider"),
            pl.col("preferred_lookup_key").alias("_metadata_lookup_key"),
            pl.col("include_after_release_cutoff").cast(pl.Boolean).alias("_include_after_release_cutoff"),
        ])
        .collect()
    )

    if metadata_ready.is_empty():
        return empty

    lf = pl.scan_parquet(standardized_path)
    schema = lf.collect_schema()

    has_share_expr = (
        pl.col("has_share_in_out").cast(pl.Boolean, strict=False)
        if has_col(schema, "has_share_in_out")
        else pl.lit(False)
    )

    return (
        lf
        .filter(
            (pl.col("source") == source)
            & (pl.col("account") == account)
        )
        .pipe(lambda frame: filter_reportable_catalog(frame, set(schema.names())))
        .with_columns([
            col_or_null(schema, "statement_period").alias("_statement_period"),
            col_or_null(schema, "artist_statement_style").alias("_artist_raw"),
            amount_expr(schema).alias("_amount"),
            has_share_expr.alias("_has_share_in_out"),
            release_metadata_provider_expr(schema).alias("_metadata_provider"),
            release_metadata_lookup_key_expr(schema).alias("_metadata_lookup_key"),
        ])
        .join(
            metadata_ready.lazy(),
            on=["_metadata_provider", "_metadata_lookup_key"],
            how="left",
        )
        .filter(
            (pl.col("_include_after_release_cutoff") == True)
            & pl.col("_statement_period").is_not_null()
            & (pl.col("_statement_period").str.strip_chars() != "")
        )
        .with_columns(
            pl.when(
                pl.col("_artist_raw").is_null()
                | (pl.col("_artist_raw").str.strip_chars() == "")
            )
            .then(pl.lit("SIN ARTISTA"))
            .otherwise(pl.col("_artist_raw").str.strip_chars())
            .alias("artist")
        )
        .group_by(["artist", "_statement_period"])
        .agg([
            pl.sum("_amount").round(2).alias("total"),
            pl.max("_has_share_in_out").cast(pl.Int64).alias("has_share_in_out"),
        ])
        .with_columns([
            pl.lit(source).alias("source"),
            pl.lit(output_account).alias("account"),
        ])
        .rename({"_statement_period": "statement_period"})
        .select(["source", "account", "artist", "statement_period", "total", "has_share_in_out"])
        .collect()
        .to_pandas()
    )


def build_gusty_post_motorcito_totals(standardized_path: Path | None) -> pd.DataFrame:
    return build_post_reference_totals(
        standardized_path=standardized_path,
        source=GUSTY_VARIANT_SOURCE,
        base_account=GUSTY_VARIANT_BASE_ACCOUNT,
        output_account=GUSTY_VARIANT_ACCOUNT,
        reference_terms=["motorcito"],
        cutoff_basis="statement_period",
        track_key_builder=legacy_gusty_track_key_expr,
        title_match_builder=legacy_title_match_expr,
    )


def build_statement_report_new_variants(standardized_path: Path | None) -> pd.DataFrame:
    variants = []

    mawz_masters = build_source_sheet_totals(
        standardized_path=standardized_path,
        source="onerpm",
        account="mawzrecords",
        source_sheet="Masters",
    )
    if not mawz_masters.empty:
        variants.append(mawz_masters)

    gusty = build_post_reference_totals(
        standardized_path=standardized_path,
        source="onerpm",
        base_account="gusty_dj",
        output_account="gusty_dj",
        reference_terms=["motorcito"],
    )
    if not gusty.empty:
        variants.append(gusty)

    la_nueva_sangre = build_release_metadata_totals(
        standardized_path=standardized_path,
        source="onerpm",
        account="la_nueva_sangre",
        output_account="la_nueva_sangre",
    )
    if la_nueva_sangre.empty:
        la_nueva_sangre = build_post_reference_totals(
            standardized_path=standardized_path,
            source="onerpm",
            base_account="la_nueva_sangre",
            output_account="la_nueva_sangre",
            reference_terms=LA_NUEVA_SANGRE_REFERENCE_TERMS,
            reference_source="onerpm",
            reference_account="mawzrecords",
            fallback_cutoff_month=LA_NUEVA_SANGRE_CONTRACT_CUTOFF_DATE[:7],
        )
    if not la_nueva_sangre.empty:
        variants.append(la_nueva_sangre)

    if not variants:
        return pd.DataFrame(columns=["source", "account", "artist", "statement_period", "total", "has_share_in_out"])

    return pd.concat(variants, ignore_index=True)


def read_registry_entries(path: Path) -> list[dict]:
    if not path.exists():
        return []
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return []
    entries = payload.get("entries", [])
    return entries if isinstance(entries, list) else []


def normalize_statement_totals(df: pd.DataFrame) -> pd.DataFrame:
    columns = ["source", "account", "artist", "statement_period", "total", "has_share_in_out"]
    if df.empty:
        return pd.DataFrame(columns=columns)

    out = df.copy()
    for col in ["source", "account", "artist", "statement_period"]:
        out[col] = out[col].fillna("").astype(str)
    out["total"] = pd.to_numeric(out["total"], errors="coerce").fillna(0.0)
    if "has_share_in_out" not in out.columns:
        out["has_share_in_out"] = 0
    out["has_share_in_out"] = pd.to_numeric(
        out["has_share_in_out"],
        errors="coerce",
    ).fillna(0).astype(int)

    grouped = (
        out
        .groupby(["source", "account", "artist", "statement_period"], dropna=False, as_index=False)
        .agg(
            total=("total", "sum"),
            has_share_in_out=("has_share_in_out", "max"),
        )
    )
    grouped["total"] = grouped["total"].round(2)
    return grouped[columns]


def cutoff_reference_scope(cutoff: dict) -> tuple[str | None, str | None]:
    observed_source = str(cutoff.get("evidence_observed_source") or "")
    if "/" not in observed_source:
        return None, None
    source, account = observed_source.split("/", 1)
    return source or None, account or None


def build_policy_statement_view(
    base_df: pd.DataFrame,
    standardized_path: Path | None,
    policy_path: Path = STATEMENT_POLICY_PATH,
    cutoffs_path: Path = CONTRACT_CUTOFFS_PATH,
) -> pd.DataFrame:
    policies = read_registry_entries(policy_path)
    if not policies:
        return pd.DataFrame(columns=["source", "account", "artist", "statement_period", "total", "has_share_in_out"])

    cutoffs = {
        entry.get("cutoff_id"): entry
        for entry in read_registry_entries(cutoffs_path)
        if entry.get("cutoff_id")
    }
    pieces: list[pd.DataFrame] = []

    for policy in policies:
        source = str(policy.get("source") or "")
        account = str(policy.get("account") or "")
        if not source or not account:
            continue
        if policy.get("statement_view_enabled") is not True:
            continue

        cutoff_id = policy.get("contract_cutoff_id")
        if cutoff_id:
            cutoff = cutoffs.get(cutoff_id, {})
            evidence_terms = [str(term) for term in (cutoff.get("evidence_terms") or [])]
            cutoff_basis = str(cutoff.get("cutoff_basis") or "transaction_month")

            if account == "la_nueva_sangre":
                release_based = build_release_metadata_totals(
                    standardized_path=standardized_path,
                    source=source,
                    account=account,
                    output_account=account,
                )
                if not release_based.empty:
                    pieces.append(release_based)
                    continue

            reference_source, reference_account = cutoff_reference_scope(cutoff)
            fallback_month = cutoff.get("contract_start_month")
            pieces.append(
                build_post_reference_totals(
                    standardized_path=standardized_path,
                    source=source,
                    base_account=account,
                    output_account=account,
                    reference_terms=evidence_terms,
                    fallback_cutoff_month=str(fallback_month) if fallback_month else None,
                    reference_source=reference_source,
                    reference_account=reference_account,
                    cutoff_basis="transaction_month"
                    if "transaction_month" in cutoff_basis
                    else "statement_period",
                )
            )
            continue

        if source == "onerpm":
            sheet_rules = policy.get("sheet_rules") or {}
            for source_sheet, rule in sheet_rules.items():
                if isinstance(rule, dict) and rule.get("statement_view") is True:
                    pieces.append(
                        build_source_sheet_totals(
                            standardized_path=standardized_path,
                            source=source,
                            account=account,
                            source_sheet=str(source_sheet),
                        )
                    )
            continue

        pieces.append(
            base_df.loc[
                (base_df["source"] == source)
                & (base_df["account"] == account)
            ].copy()
        )

    if not pieces:
        return pd.DataFrame(columns=["source", "account", "artist", "statement_period", "total", "has_share_in_out"])

    return normalize_statement_totals(pd.concat(pieces, ignore_index=True))


def build_hardcoded_new_statement_view(
    base_df: pd.DataFrame,
    standardized_path: Path | None,
) -> pd.DataFrame:
    df = base_df.loc[
        (base_df["source"] != "onerpm")
        | (base_df["account"].isin(NEW_REPORT_ONERPM_ACCOUNTS))
    ].copy()
    variants = build_statement_report_new_variants(standardized_path)
    if not variants.empty:
        df = pd.concat([df, variants], ignore_index=True)
    return normalize_statement_totals(df)


def aggregate_statement_data(standardized_path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    if not standardized_path.exists():
        raise FileNotFoundError(f"No existe mart standardized: {standardized_path}")

    lf = pl.scan_parquet(standardized_path)
    schema = lf.collect_schema()

    source = col_or_null(schema, "source")
    amount = amount_expr(schema)
    adjusted_amount = (
        pl.when(source == "fuga")
        .then(amount * FUGA_STATEMENT_FACTOR)
        .otherwise(amount)
    )

    has_share_expr = (
        pl.col("has_share_in_out").cast(pl.Boolean, strict=False)
        if has_col(schema, "has_share_in_out")
        else pl.lit(False)
    )

    base = lf.with_columns([
        source.alias("_source"),
        col_or_null(schema, "account").alias("_account"),
        col_or_null(schema, "statement_period").alias("_statement_period"),
        col_or_null(schema, "artist_statement_style").alias("_artist_raw"),
        col_or_null(schema, "source_sheet").alias("_source_sheet"),
        adjusted_amount.alias("_amount"),
        has_share_expr.alias("_has_share_in_out"),
    ])

    base = base.filter(
        pl.col("_statement_period").is_not_null()
        & (pl.col("_statement_period").str.strip_chars() != "")
    )

    base = base.filter(
        (pl.col("_source") != "soundon")
        | (pl.col("_source_sheet") == "my_royalty")
    )
    base = filter_reportable_catalog(base, set(schema.names()))

    totals = (
        base
        .with_columns(
            pl.when(
                pl.col("_artist_raw").is_null()
                | (pl.col("_artist_raw").str.strip_chars() == "")
            )
            .then(pl.lit("SIN ARTISTA"))
            .otherwise(pl.col("_artist_raw").str.strip_chars())
            .alias("artist")
        )
        .group_by(["_source", "_account", "artist", "_statement_period"])
        .agg([
            pl.sum("_amount").round(2).alias("total"),
            pl.max("_has_share_in_out").cast(pl.Int64).alias("has_share_in_out"),
        ])
        .rename({
            "_source": "source",
            "_account": "account",
            "_statement_period": "statement_period",
        })
        .collect()
        .to_pandas()
    )

    fuga_eur = pd.DataFrame(columns=["source", "account", "statement_period", "total_eur"])
    if has_col(schema, "net_amount"):
        fuga_eur = (
            lf
            .filter(
                (pl.col("source") == "fuga")
                & pl.col("statement_period").is_not_null()
                & (pl.col("statement_period").cast(pl.Utf8).str.strip_chars() != "")
            )
            .group_by(["source", "account", "statement_period"])
            .agg(pl.sum("net_amount").round(2).alias("total_eur"))
            .collect()
            .to_pandas()
        )

    return totals, fuga_eur


def aggregate_statement_summary(summary_path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    if not summary_path.exists():
        raise FileNotFoundError(f"No existe mart statement summary: {summary_path}")

    df = pl.read_parquet(summary_path).to_pandas()
    if df.empty:
        return pd.DataFrame(), pd.DataFrame()

    artist_rows = df.loc[df["row_type"] == "artist_total"].copy()
    fuga_eur_rows = df.loc[df["row_type"] == "fuga_eur_total"].copy()

    totals = artist_rows[
        ["source", "account", "artist", "statement_period", "total", "has_share_in_out"]
    ].copy()
    totals["total"] = pd.to_numeric(totals["total"], errors="coerce").fillna(0.0)
    totals["has_share_in_out"] = pd.to_numeric(
        totals["has_share_in_out"],
        errors="coerce",
    ).fillna(0).astype(int)

    fuga_eur = pd.DataFrame(columns=["source", "account", "statement_period", "total_eur"])
    if not fuga_eur_rows.empty:
        fuga_eur = fuga_eur_rows[
            ["source", "account", "statement_period", "total_eur"]
        ].copy()
        fuga_eur["total_eur"] = pd.to_numeric(fuga_eur["total_eur"], errors="coerce").fillna(0.0)

    return totals, fuga_eur


def filter_by_scope_artist_threshold(
    df: pd.DataFrame,
    min_artist_total_usd: float,
    include_zero_total_artists: bool = False,
) -> pd.DataFrame:
    if df.empty:
        return df

    scope_totals = (
        df
        .groupby(["source", "account", "artist"], dropna=False, as_index=False)["total"]
        .sum()
    )
    if min_artist_total_usd <= 0 and include_zero_total_artists:
        return df

    if min_artist_total_usd <= 0:
        keep_mask = scope_totals["total"].abs().round(2) != 0
    else:
        keep_mask = scope_totals["total"].abs() >= min_artist_total_usd

    keep = scope_totals.loc[
        keep_mask,
        ["source", "account", "artist"],
    ]

    return df.merge(keep, on=["source", "account", "artist"], how="inner")


def write_statement_report(
    df: pd.DataFrame,
    fuga_eur_df: pd.DataFrame,
    output_path: Path,
    min_artist_total_usd: float = 0.0,
    include_zero_total_artists: bool = False,
    standardized_path: Path | None = STANDARDIZED_ALL_PATH,
    report_version: str = "legacy",
) -> Path:
    if df.empty:
        raise ValueError("No hay datos para generar el reporte por statement.")

    default_excluded_scopes: set[str] = {"onerpm_gusty_dj"}
    if report_version == "new":
        default_excluded_scopes = set()
        policy_df = build_policy_statement_view(df, standardized_path)
        if policy_df.empty:
            policy_df = build_hardcoded_new_statement_view(df, standardized_path)
        df = policy_df
    else:
        gusty_variant = build_gusty_post_motorcito_totals(standardized_path)
        if not gusty_variant.empty:
            df = pd.concat([df, gusty_variant], ignore_index=True)

    df = filter_by_scope_artist_threshold(df, min_artist_total_usd, include_zero_total_artists)
    if df.empty:
        raise ValueError("No hay datos para generar el reporte por statement con el umbral seleccionado.")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        scopes = sorted(
            {
                (str(source), str(account))
                for source, account in df[["source", "account"]].drop_duplicates().itertuples(index=False, name=None)
            }
        )
        add_config_sheet(writer, scopes, min_artist_total_usd, include_zero_total_artists, default_excluded_scopes)
        config_rows = len(scopes) + 1

        total_source = df.copy()
        total_source["_scope_key"] = total_source.apply(
            lambda row: scope_key(row["source"], row["account"]),
            axis=1,
        )
        total_data_last_row = add_total_data_sheet(writer, total_source, config_rows)

        pivot_total = total_source.pivot_table(
            index="artist",
            columns="statement_period",
            values="total",
            aggfunc="sum",
            fill_value=0,
        )

        pivot_total = add_totals_and_clean(pivot_total, include_zero_total_artists)

        if not pivot_total.empty:
            total_general = pivot_total.sum(axis=0)
            total_general.name = "TOTAL USD"
            pivot_total = pd.concat([pivot_total, total_general.to_frame().T])
            pivot_total.to_excel(writer, sheet_name="TOTAL", startrow=1)
            format_sheet(writer.sheets["TOTAL"], pivot_total, "SALDO TOTAL ARTISTAS")
            apply_total_config_formulas(writer.sheets["TOTAL"], pivot_total, total_data_last_row)
            hide_initial_zero_total_rows(
                writer.sheets["TOTAL"],
                pivot_total,
                total_source,
                scopes,
                include_zero_total_artists,
                default_excluded_scopes,
            )

        for (source, account), group in df.groupby(["source", "account"], dropna=False):
            pivot = group.pivot_table(
                index="artist",
                columns="statement_period",
                values="total",
                aggfunc="sum",
                fill_value=0,
            )

            pivot = add_totals_and_clean(pivot, include_zero_total_artists)

            if pivot.empty:
                continue

            total_usd = pivot.sum(axis=0)
            total_usd.name = "TOTAL USD"
            rows_to_add = [total_usd.to_frame().T]

            if source == "fuga":
                matches = fuga_eur_df[
                    (fuga_eur_df["source"] == source)
                    & (fuga_eur_df["account"] == account)
                ]

                total_eur = pd.Series(0.0, index=pivot.columns)
                total_eur.name = "TOTAL EUR"

                for col in pivot.columns:
                    if col == "TOTAL":
                        continue

                    period_matches = matches.loc[matches["statement_period"] == col, "total_eur"]
                    if not period_matches.empty:
                        total_eur[col] = float(period_matches.iloc[0])

                total_eur["TOTAL"] = total_eur.drop(labels=["TOTAL"], errors="ignore").sum()
                rows_to_add.append(total_eur.to_frame().T)

            share_flags = None
            if source == "onerpm" and account == "gusty_dj" and "has_share_in_out" in group.columns:
                share_flags = group.pivot_table(
                    index="artist",
                    columns="statement_period",
                    values="has_share_in_out",
                    aggfunc="max",
                    fill_value=0,
                )
                share_flags = share_flags.reindex(
                    index=pivot.index,
                    columns=[col for col in pivot.columns if col != "TOTAL"],
                    fill_value=0,
                )

            pivot = pd.concat([pivot] + rows_to_add)
            sheet_name = f"{source}_{account}"[:31]
            pivot.to_excel(writer, sheet_name=sheet_name, startrow=1)
            company_name = NEW_REPORT_VARIANT_TITLES.get(
                (str(source), str(account)),
                GUSTY_VARIANT_TITLE if account == GUSTY_VARIANT_ACCOUNT else f"{str(source).upper()} - {account}",
            )
            format_sheet(writer.sheets[sheet_name], pivot, company_name, share_flags)

        enable_formula_recalculation(writer.book)

    return output_path


def build_statement_report_from_mart(
    standardized_path: Path = STANDARDIZED_ALL_PATH,
    output_path: Path | None = None,
    min_artist_total_usd: float = 0.0,
    include_zero_total_artists: bool = False,
    report_version: str = "legacy",
) -> Path:
    if output_path is None:
        output_path = REPORTS_DIR / "reporte_ingresos_digitales_por_mes_de_statement_marts.xlsx"

    df, fuga_eur_df = aggregate_statement_data(standardized_path)
    return write_statement_report(
        df,
        fuga_eur_df,
        output_path,
        min_artist_total_usd,
        include_zero_total_artists,
        standardized_path,
        report_version,
    )


def build_statement_report_from_summary(
    summary_path: Path = STATEMENT_SUMMARY_PATH,
    output_path: Path | None = None,
    min_artist_total_usd: float = 0.0,
    include_zero_total_artists: bool = False,
    standardized_path: Path | None = STANDARDIZED_ALL_PATH,
    report_version: str = "legacy",
) -> Path:
    if output_path is None:
        output_path = REPORTS_DIR / "reporte_ingresos_digitales_por_mes_de_statement_marts.xlsx"

    df, fuga_eur_df = aggregate_statement_summary(summary_path)
    return write_statement_report(
        df,
        fuga_eur_df,
        output_path,
        min_artist_total_usd,
        include_zero_total_artists,
        standardized_path,
        report_version,
    )


def main():
    output = build_statement_report_from_mart()
    print("Reporte generado en:")
    print(output)


if __name__ == "__main__":
    main()
