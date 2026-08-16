from __future__ import annotations

import argparse
from datetime import date, datetime
from decimal import Decimal
import json
import os
from pathlib import Path
import sys
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.operational_db import operational_connect


DEFAULT_WORKBOOK = ROOT / "outputs" / "agenda_indyana_20260816" / "agenda_indyana_validacion_simple_20260816.xlsx"
DEFAULT_SECRET = ROOT / ".secrets" / "cloudsql_operational.env"
SOURCE_SYSTEM = "agenda_indyana_excel"
SOURCE_SHEET = "Agenda para validar"
ACTOR = "system_agenda_import"

EXISTING_EVENT_IDS: dict[int, list[int]] = {
    5: [1287],
    6: [1286],
    7: [1288],
    8: [1289],
    9: [1293],
    16: [1294],
    17: [1351],
    19: [1297],
    26: [1302],
    27: [1301],
    36: [1308],
    37: [1309],
    42: [1315],
    43: [1317],
    44: [1318],
    57: [1352],
    60: [1353],
    86: [1332],
    92: [1334, 1335, 1336],
    97: [1354],
    98: [1337],
}

AVAILABILITY_ROWS = {76, 139, 143, 229, 232}
LOGISTICS_ROWS = {197, 235, 237, 239, 242}
PROSPECT_ROWS = {245, 246}
GROUP_ROWS = {166}
SHARED_SHOW_ROWS = {119, 149, 168, 198, 209}

DISPLAY_NAMES: dict[int, str] = {
    33: "FA",
    65: "Gustavo",
    72: "Un poco de ruido",
    76: "No trabaja",
    106: "Fer Almiron - Chubut",
    107: "Pico Truncado",
    108: "Caleta Olivia",
    117: "San Vicente",
    118: "Hawaian",
    119: "Maza Bar",
    124: "Sergio x2 - Azul y otra",
    128: "J. C. Paz",
    129: "Boulevard Bailable",
    130: "Grimaldi",
    131: "Grand Rex",
    139: "No trabaja",
    143: "No trabaja",
    147: "La Peña del Morfi",
    149: "Vita",
    150: "El Quincho",
    151: "Parrilla Grill",
    154: "La Reyna",
    155: "La Retro",
    156: "Terraza",
    162: "Hemisferio",
    163: "Ama Burger",
    164: "Chavela",
    166: "Teodolina las 2",
    168: "Neuquén",
    171: "Salta",
    174: "Salta",
    177: "Salta",
    180: "Salta",
    181: "Auditorio Haedo",
    182: "Eros",
    186: "Campana",
    187: "Yankee",
    188: "Tornado",
    197: "Vuelo 08:15 - Aeroparque",
    198: "Cervantes",
    203: "San Clemente",
    209: "Chaitén",
    219: "Maria Club",
    221: "Madariaga",
    223: "Salto, Uruguay",
    229: "No trabaja",
    232: "No trabaja",
    235: "EEUU",
    237: "EEUU",
    239: "EEUU",
    242: "EEUU",
    243: "Catamarca",
    244: "Catamarca",
    245: "Posible teatro",
    246: "Posible teatro",
    248: "Club Araos",
}

CACHETS: dict[int, Decimal] = {
    106: Decimal("10000000"),
    117: Decimal("3500000"),
    118: Decimal("3200000"),
    119: Decimal("6000000"),
    151: Decimal("3500000"),
    162: Decimal("3500000"),
    164: Decimal("3200000"),
    166: Decimal("19000000"),
    181: Decimal("3500000"),
    187: Decimal("7000000"),
    198: Decimal("23000000"),
    203: Decimal("12000000"),
    209: Decimal("20000000"),
    221: Decimal("10000000"),
    248: Decimal("8000000"),
}


def load_secret_env(path: Path) -> None:
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))
    os.environ["VPO_OPERATIONAL_DB_DRIVER"] = "postgres"
    os.environ["VPO_POSTGRES_CONNECT_MODE"] = "local_proxy"
    os.environ["VPO_OPERATIONAL_DB_NAME"] = os.environ["CLOUDSQL_DATABASE"]
    os.environ["VPO_OPERATIONAL_DB_USER"] = os.environ["CLOUDSQL_USER"]
    os.environ["VPO_OPERATIONAL_DB_PASSWORD"] = os.environ["CLOUDSQL_PASSWORD"]
    os.environ["VPO_CLOUDSQL_CONNECTION_NAME"] = os.environ["CLOUDSQL_CONNECTION_NAME"]
    os.environ["VPO_POSTGRES_LOCAL_PROXY_HOST"] = "127.0.0.1"
    os.environ["VPO_POSTGRES_LOCAL_PROXY_PORT"] = "5432"


def json_default(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    raise TypeError(f"Unsupported JSON value: {type(value).__name__}")


def read_candu_rows(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[SOURCE_SHEET]
    rows: list[dict[str, Any]] = []
    for row_number in range(5, sheet.max_row + 1):
        values = [sheet.cell(row_number, column).value for column in range(1, 12)]
        artist = str(values[1] or "")
        if "candu" not in artist.casefold():
            continue
        event_date = values[0]
        if not isinstance(event_date, (date, datetime)):
            raise ValueError(f"Renglon {row_number}: fecha invalida")
        rows.append(
            {
                "row_number": row_number,
                "event_date": event_date.date() if isinstance(event_date, datetime) else event_date,
                "artist_source": artist,
                "source_text": str(values[2] or "").strip(),
                "match_result": str(values[3] or "").strip(),
                "workbook_system_id": values[4],
                "booking_artist": values[5],
                "booking_venue": values[6],
                "booking_city": values[7],
                "responsible": values[8],
                "validation": str(values[9] or "").strip(),
                "observations": str(values[10] or "").strip(),
            }
        )
    if len(rows) != 77:
        raise ValueError(f"Se esperaban 77 renglones de Candu y se encontraron {len(rows)}")
    return rows


def event_type_for(row_number: int) -> str:
    if row_number in AVAILABILITY_ROWS:
        return "availability_block"
    if row_number in LOGISTICS_ROWS:
        return "logistics"
    if row_number in PROSPECT_ROWS:
        return "prospect"
    if row_number in GROUP_ROWS:
        return "show_group"
    return "show"


def source_reference(workbook_path: Path, row_number: int) -> str:
    return f"{workbook_path.stem}:{SOURCE_SHEET}:{row_number}"


def event_statuses(event_type: str, event_date: date) -> tuple[str, str, str]:
    if event_type == "show":
        return "confirmado", "realizado" if event_date < date.today() else "programado", "no_iniciada"
    if event_type == "show_group":
        return "no_aplica", "realizado" if event_date < date.today() else "programado", "no_aplica"
    if event_type == "availability_block":
        return "no_aplica", "bloqueado", "no_aplica"
    if event_type == "logistics":
        return "no_aplica", "informativo", "no_aplica"
    return "prospecto", "programado", "no_aplica"


def source_payload(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "sheet": SOURCE_SHEET,
        "row": row["row_number"],
        "artist_source": row["artist_source"],
        "match_result": row["match_result"],
        "workbook_system_id": row["workbook_system_id"],
        "validation": row["validation"],
        "observations": row["observations"],
    }


def insert_source_link(
    conn: Any,
    *,
    event_id: int,
    reference: str,
    role: str,
    source_text: str,
    payload: dict[str, Any],
) -> None:
    conn.execute(
        """
        INSERT INTO booking_event_source_links (
            event_id, source_system, source_reference, source_role,
            source_text, source_payload_json, created_by
        )
        VALUES (%s, %s, %s, %s, %s, %s::jsonb, %s)
        ON CONFLICT (event_id, source_system, source_reference) DO NOTHING
        """,
        (event_id, SOURCE_SYSTEM, reference, role, source_text, json.dumps(payload, ensure_ascii=False), ACTOR),
    )


def insert_event(
    conn: Any,
    *,
    event_type: str,
    event_date: date,
    venue: str,
    artists: list[dict[str, Any]],
    cachet: Decimal = Decimal("0"),
    group_event_id: int | None = None,
    group_position: int | None = None,
    notes: str | None = None,
) -> int:
    commercial, operational, settlement = event_statuses(event_type, event_date)
    booking_mode = "shared" if len(artists) > 1 else "individual"
    row = conn.execute(
        """
        INSERT INTO booking_events (
            event_type, event_date, venue, booking_mode,
            commercial_status, operational_status, deposit_status, settlement_status,
            contracted_cachet_amount, currency, group_event_id, group_position,
            notes, created_by
        )
        VALUES (%s, %s, %s, %s, %s, %s, 'no_informada', %s, %s, 'ARS', %s, %s, %s, %s)
        RETURNING id
        """,
        (
            event_type,
            event_date,
            venue,
            booking_mode,
            commercial,
            operational,
            settlement,
            cachet,
            group_event_id,
            group_position,
            notes,
            ACTOR,
        ),
    ).fetchone()
    event_id = int(row["id"])
    for position, artist in enumerate(artists, start=1):
        conn.execute(
            """
            INSERT INTO booking_event_artists (event_id, artist_id, artist_name, position)
            VALUES (%s, %s, %s, %s)
            """,
            (event_id, artist["id"], artist["stage_name"], position),
        )
    return event_id


def build_plan(rows: list[dict[str, Any]]) -> dict[str, Any]:
    existing_sources = [row for row in rows if row["row_number"] in EXISTING_EVENT_IDS]
    new_sources = [row for row in rows if row["row_number"] not in EXISTING_EVENT_IDS]
    return {
        "source_rows": len(rows),
        "existing_source_rows": len(existing_sources),
        "existing_event_links": sum(len(EXISTING_EVENT_IDS[row["row_number"]]) for row in existing_sources),
        "new_show_sources": sum(event_type_for(row["row_number"]) in {"show", "show_group"} for row in new_sources),
        "availability_blocks": sum(event_type_for(row["row_number"]) == "availability_block" for row in new_sources),
        "logistics": sum(event_type_for(row["row_number"]) == "logistics" for row in new_sources),
        "prospects": sum(event_type_for(row["row_number"]) == "prospect" for row in new_sources),
        "new_event_records": len(new_sources) + 2,
        "expected_source_links": len(rows) + 4,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa una agenda conciliada a booking_events sin duplicar Booking.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--secret-file", type=Path, default=DEFAULT_SECRET)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    load_secret_env(args.secret_file)
    rows = read_candu_rows(args.workbook)
    plan = build_plan(rows)
    print(json.dumps({"mode": "apply" if args.apply else "dry_run", **plan}, ensure_ascii=False, indent=2))

    with operational_connect() as conn:
        artist_rows = conn.execute(
            "SELECT id, stage_name FROM artists WHERE id IN (7, 11) ORDER BY id"
        ).fetchall()
        artists = {int(row["id"]): dict(row) for row in artist_rows}
        if set(artists) != {7, 11}:
            raise RuntimeError("No se encontraron los artistas canonicos Candu y G Sony")

        existing_ids = sorted({event_id for values in EXISTING_EVENT_IDS.values() for event_id in values})
        existing_rows = conn.execute(
            "SELECT id, event_date, venue, booking_mode FROM booking_events WHERE id = ANY(%s) ORDER BY id",
            (existing_ids,),
        ).fetchall()
        if len(existing_rows) != len(existing_ids):
            found = {int(row["id"]) for row in existing_rows}
            raise RuntimeError(f"Faltan eventos existentes: {sorted(set(existing_ids) - found)}")

        before_counts = dict(
            conn.execute(
                """
                SELECT
                    (SELECT count(*) FROM booking_events) AS booking_events,
                    (SELECT count(*) FROM booking_shows) AS booking_shows,
                    (SELECT count(*) FROM booking_composite_events) AS composite_events,
                    (SELECT count(*) FROM booking_event_deposits) AS deposits
                """
            ).fetchone()
        )
        if not args.apply:
            print(json.dumps({"database_before": before_counts, "verified_existing_events": len(existing_rows)}, default=json_default, indent=2))
            return

        backup_dir = args.workbook.parent / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backup_dir / f"candu_agenda_preimport_{stamp}.json"
        backup_path.write_text(
            json.dumps(
                {
                    "created_at": datetime.now().isoformat(),
                    "workbook": str(args.workbook),
                    "database_counts": before_counts,
                    "existing_events": [dict(row) for row in existing_rows],
                    "plan": plan,
                },
                ensure_ascii=False,
                indent=2,
                default=json_default,
            ),
            encoding="utf-8",
        )

        imported_event_ids: list[int] = []
        linked_existing_ids: list[int] = []
        skipped_sources: list[str] = []
        for source_row in rows:
            row_number = source_row["row_number"]
            reference = source_reference(args.workbook, row_number)
            already = conn.execute(
                """
                SELECT event_id FROM booking_event_source_links
                WHERE source_system = %s AND source_reference = %s
                ORDER BY event_id
                """,
                (SOURCE_SYSTEM, reference),
            ).fetchall()
            if already:
                skipped_sources.append(reference)
                continue

            payload = source_payload(source_row)
            if row_number in EXISTING_EVENT_IDS:
                for event_id in EXISTING_EVENT_IDS[row_number]:
                    insert_source_link(
                        conn,
                        event_id=event_id,
                        reference=reference,
                        role="linked_existing",
                        source_text=source_row["source_text"],
                        payload=payload,
                    )
                    linked_existing_ids.append(event_id)
                continue

            event_type = event_type_for(row_number)
            selected_artists = [artists[7], artists[11]] if row_number in SHARED_SHOW_ROWS else [artists[7]]
            venue = DISPLAY_NAMES.get(row_number, source_row["source_text"] or "Agenda Candu")
            notes = source_row["observations"] or None

            if event_type == "show_group":
                group_id = insert_event(
                    conn,
                    event_type="show_group",
                    event_date=source_row["event_date"],
                    venue=venue,
                    artists=[artists[7]],
                    cachet=CACHETS[row_number],
                    notes="Dos shows relacionados; total del grupo. Cada hijo se liquida por separado.",
                )
                imported_event_ids.append(group_id)
                insert_source_link(
                    conn,
                    event_id=group_id,
                    reference=reference,
                    role="imported_group",
                    source_text=source_row["source_text"],
                    payload=payload,
                )
                for position in (1, 2):
                    child_id = insert_event(
                        conn,
                        event_type="show",
                        event_date=source_row["event_date"],
                        venue=f"Teodolina - Show {position}",
                        artists=[artists[7]],
                        cachet=Decimal("9500000"),
                        group_event_id=group_id,
                        group_position=position,
                        notes="Parte del grupo Teodolina las 2.",
                    )
                    imported_event_ids.append(child_id)
                    insert_source_link(
                        conn,
                        event_id=child_id,
                        reference=reference,
                        role=f"group_child_{position}",
                        source_text=source_row["source_text"],
                        payload=payload,
                    )
                continue

            event_id = insert_event(
                conn,
                event_type=event_type,
                event_date=source_row["event_date"],
                venue=venue,
                artists=selected_artists,
                cachet=CACHETS.get(row_number, Decimal("0")),
                notes=notes,
            )
            imported_event_ids.append(event_id)
            insert_source_link(
                conn,
                event_id=event_id,
                reference=reference,
                role="imported",
                source_text=source_row["source_text"],
                payload=payload,
            )

        after_counts = dict(
            conn.execute(
                """
                SELECT
                    (SELECT count(*) FROM booking_events) AS booking_events,
                    (SELECT count(*) FROM booking_shows) AS booking_shows,
                    (SELECT count(*) FROM booking_composite_events) AS composite_events,
                    (SELECT count(*) FROM booking_event_deposits) AS deposits,
                    (SELECT count(*) FROM booking_event_source_links WHERE source_system = %s) AS source_links
                """,
                (SOURCE_SYSTEM,),
            ).fetchone()
        )
        if not skipped_sources:
            if after_counts["booking_events"] - before_counts["booking_events"] != plan["new_event_records"]:
                raise RuntimeError("La cantidad de eventos nuevos no coincide con el plan")
            if after_counts["booking_shows"] != before_counts["booking_shows"]:
                raise RuntimeError("La importacion modifico booking_shows")
            if after_counts["composite_events"] != before_counts["composite_events"]:
                raise RuntimeError("La importacion modifico booking_composite_events")
            if after_counts["deposits"] != before_counts["deposits"]:
                raise RuntimeError("La importacion genero movimientos de seña")

        conn.execute(
            """
            INSERT INTO app_audit_log (
                actor_username, module_key, action, entity_table, entity_id,
                after_json, source, notes
            )
            VALUES (%s, 'booking', 'import', 'booking_events', %s, %s::jsonb, 'validated_workbook', %s)
            """,
            (
                ACTOR,
                args.workbook.stem,
                json.dumps(
                    {
                        "new_event_ids": imported_event_ids,
                        "linked_existing_ids": linked_existing_ids,
                        "skipped_sources": skipped_sources,
                        "counts_before": before_counts,
                        "counts_after": after_counts,
                    },
                    ensure_ascii=False,
                    default=json_default,
                ),
                "Importacion conciliada de la Agenda actual de Candu; sin hechos financieros.",
            ),
        )

        print(
            json.dumps(
                {
                    "backup": str(backup_path),
                    "new_event_ids": imported_event_ids,
                    "linked_existing_ids": linked_existing_ids,
                    "skipped_sources": skipped_sources,
                    "database_before": before_counts,
                    "database_after": after_counts,
                },
                ensure_ascii=False,
                indent=2,
                default=json_default,
            )
        )


if __name__ == "__main__":
    main()
