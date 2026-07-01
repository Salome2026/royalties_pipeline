from __future__ import annotations

import re
import unicodedata
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

import pandas as pd
import polars as pl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from lib.catalog_report_filter import filter_reportable_catalog
except ModuleNotFoundError:
    from scripts.lib.catalog_report_filter import filter_reportable_catalog


BASE = Path(r"C:\royalties_pipeline")
MARTS = BASE / "warehouse" / "marts"
REPORTS = BASE / "reports"
RAW_PATH = MARTS / "standardized_raw_all_sources.parquet"

MOTORCITO_FIRST_MONTH = "2023-04"


def strip_accents(value: str) -> str:
    return "".join(
        char for char in unicodedata.normalize("NFKD", str(value or ""))
        if not unicodedata.combining(char)
    )


def normalize_text(value: str) -> str:
    value = strip_accents(str(value or "").lower())
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return " ".join(value.split())


def signature(value: str) -> str:
    text = normalize_text(value)
    text = re.sub(r"\b(gusty|dj|video|oficial|official|remix|ft|feat|featuring|session|sesion|en|el|la|los|las|rkt)\b", " ", text)
    return " ".join(text.split())


def text_col(name: str, default: str = "") -> pl.Expr:
    return pl.col(name).cast(pl.Utf8, strict=False).fill_null(default)


def amount_col(name: str) -> pl.Expr:
    return pl.col(name).cast(pl.Float64, strict=False).fill_null(0.0) if name else pl.lit(0.0)


def first_existing(columns: set[str], candidates: list[str]) -> str | None:
    return next((col for col in candidates if col in columns), None)


def best_title_expr(columns: set[str]) -> pl.Expr:
    candidates = [
        "asset_title_statement",
        "track_statement_style",
        "Title",
        "Video Title",
        "Track Title",
        "Album Title",
        "album_title",
        "Asset Title",
        "Product Title",
        "TRACK",
        "PRODUCT",
    ]
    exprs = []
    for col in candidates:
        if col in columns:
            value = text_col(col).str.strip_chars()
            exprs.append(pl.when(value != "").then(value).otherwise(None))
    return (pl.coalesce(exprs) if exprs else pl.lit(None, dtype=pl.Utf8)).alias("tema")


def best_artist_expr(columns: set[str]) -> pl.Expr:
    candidates = [
        "asset_artist_statement",
        "artist_statement_style",
        "artists_raw",
        "Track Artists",
        "Channel Name",
        "Asset Artist",
        "Product Artist",
        "TRACK ARTIST",
        "PRODUCT ARTIST",
    ]
    exprs = []
    for col in candidates:
        if col in columns:
            value = text_col(col).str.strip_chars()
            exprs.append(pl.when(value != "").then(value).otherwise(None))
    return (pl.coalesce(exprs) if exprs else pl.lit(None, dtype=pl.Utf8)).alias("artist_best")


def coalesce_number_expr(columns: set[str], candidates: list[str], alias: str) -> pl.Expr:
    exprs = []
    for col in candidates:
        if col in columns:
            exprs.append(pl.col(col).cast(pl.Float64, strict=False))
    if not exprs:
        return pl.lit(0.0).alias(alias)
    return pl.coalesce(exprs).fill_null(0.0).alias(alias)


def coalesce_text_expr(columns: set[str], candidates: list[str], alias: str) -> pl.Expr:
    exprs = []
    for col in candidates:
        if col in columns:
            value = text_col(col).str.strip_chars()
            exprs.append(pl.when(value != "").then(value).otherwise(None))
    if not exprs:
        return pl.lit(None, dtype=pl.Utf8).alias(alias)
    return pl.coalesce(exprs).alias(alias)


def gusty_filter(columns: set[str]) -> pl.Expr:
    expr = pl.lit(False)
    for col in [
        "artist_statement_style",
        "asset_artist_statement",
        "track_statement_style",
        "asset_title_statement",
        "Album Title",
        "album_title",
        "artists_raw",
        "Track Artists",
        "Product Artist",
        "Asset Artist",
        "Title",
        "Video Title",
        "Track Title",
        "Channel Name",
        "Product Title",
        "Asset Title",
        "TRACK ARTIST",
        "PRODUCT ARTIST",
        "TRACK",
        "PRODUCT",
    ]:
        if col in columns:
            expr = expr | text_col(col).str.to_lowercase().str.contains("gusty", literal=True).fill_null(False)
    return expr


def prepare_base(path: Path, start_month: str | None, end_month: str | None) -> pl.DataFrame:
    columns = set(pl.scan_parquet(path).collect_schema().names())
    amount_eur = first_existing(columns, ["amount_eur", "net_amount_eur", "Reported Royalty"])
    lf = (
        pl.scan_parquet(path)
        .filter((pl.col("source") == "fuga") & (pl.col("account") == "indyana_records"))
        .filter(gusty_filter(columns))
        .with_columns(
            [
                best_title_expr(columns),
                best_artist_expr(columns),
                amount_col(amount_eur).alias("ingresos_eur"),
                coalesce_text_expr(columns, ["asset_isrc", "ISRC", "Asset ISRC", "ISRC"], "asset_isrc"),
                coalesce_number_expr(columns, ["units", "asset_quantity_num", "product_quantity_num", "Asset Quantity", "Product Quantity", "QUANTITY", "Quantity", "Units"], "units"),
                text_col("statement_period").alias("statement_period"),
                text_col("transaction_month").alias("transaction_month"),
                (text_col("content_type", "catalog") if "content_type" in columns else pl.lit("catalog")).alias("tipo_contenido"),
                (text_col("dsp") if "dsp" in columns else pl.lit(None, dtype=pl.Utf8)).alias("dsp"),
                (text_col("store_name") if "store_name" in columns else pl.lit(None, dtype=pl.Utf8)).alias("store_name"),
                (text_col("sale_type") if "sale_type" in columns else pl.lit(None, dtype=pl.Utf8)).alias("sale_type"),
                (text_col("sale_user_type") if "sale_user_type" in columns else pl.lit(None, dtype=pl.Utf8)).alias("sale_user_type"),
                (text_col("territory") if "territory" in columns else pl.lit(None, dtype=pl.Utf8)).alias("territory"),
                (text_col("statement_type") if "statement_type" in columns else pl.lit(None, dtype=pl.Utf8)).alias("statement_type"),
                (text_col("statement_file_name") if "statement_file_name" in columns else pl.lit(None, dtype=pl.Utf8)).alias("statement_file_name"),
            ]
        )
        .with_columns(
            [
                pl.col("tema").map_elements(normalize_text, return_dtype=pl.Utf8).alias("_title_key"),
                pl.col("tema").map_elements(signature, return_dtype=pl.Utf8).alias("_signature_key"),
            ]
        )
    )
    if start_month:
        lf = lf.filter(pl.col("statement_period") >= start_month)
    if end_month:
        lf = lf.filter(pl.col("statement_period") <= end_month)
    return lf.collect()


def build_onerpm_map(path: Path) -> pd.DataFrame:
    columns = set(pl.scan_parquet(path).collect_schema().names())
    motorcito_expr = (
        text_col("track_statement_style").str.to_lowercase().str.contains("motorcito", literal=True).fill_null(False)
        if "track_statement_style" in columns
        else pl.lit(False)
    )
    df = (
        pl.scan_parquet(path)
        .filter((pl.col("source") == "onerpm") & (pl.col("account") == "gusty_dj"))
        .filter(gusty_filter(columns) | motorcito_expr)
        .with_columns(
            [
                best_title_expr(columns),
                best_artist_expr(columns),
                coalesce_text_expr(columns, ["asset_isrc", "ISRC", "Asset ISRC", "isrc"], "asset_isrc"),
                amount_col("amount_usd").alias("amount_usd"),
                coalesce_number_expr(columns, ["units", "Units", "Units of Sold", "Quantity", "QUANTITY"], "units"),
            ]
        )
        .filter(pl.col("tema").is_not_null())
        .with_columns(
            [
                pl.col("tema").map_elements(normalize_text, return_dtype=pl.Utf8).alias("_title_key"),
                pl.col("tema").map_elements(signature, return_dtype=pl.Utf8).alias("_signature_key"),
            ]
        )
        .select(["asset_isrc", "tema", "artist_best", "transaction_month", "amount_usd", "units", "_title_key", "_signature_key"])
        .collect()
    )
    rows: list[dict] = []
    for method, key_col in [("isrc", "asset_isrc"), ("title", "_title_key"), ("signature", "_signature_key")]:
        tmp = df.filter(pl.col(key_col).is_not_null() & (pl.col(key_col) != ""))
        if tmp.height == 0:
            continue
        grouped = (
            tmp.group_by(key_col)
            .agg(
                [
                    pl.min("transaction_month").alias("onerpm_first_month"),
                    pl.col("tema").drop_nulls().mode().first().alias("onerpm_title"),
                    pl.col("artist_best").drop_nulls().mode().first().alias("onerpm_artist"),
                    pl.sum("amount_usd").alias("onerpm_total_usd"),
                    pl.sum("units").alias("onerpm_units"),
                    pl.len().alias("onerpm_rows"),
                ]
            )
            .rename({key_col: "match_key"})
            .to_pandas()
        )
        grouped["match_method"] = method
        rows.extend(grouped.to_dict("records"))

    out = pd.DataFrame(rows)
    if out.empty:
        return pd.DataFrame(columns=["match_method", "match_key", "onerpm_first_month", "onerpm_title", "onerpm_artist", "contract_segment"])
    out["contract_segment"] = out["onerpm_first_month"].apply(
        lambda value: "CONTRATO NUEVO" if str(value or "") >= MOTORCITO_FIRST_MONTH else "CONTRATO VIEJO"
    )
    out = out.sort_values(["match_method", "onerpm_total_usd"], ascending=[True, False])
    return out


def classify_fuga(raw: pl.DataFrame, onerpm_map: pd.DataFrame) -> pl.DataFrame:
    result = raw.with_columns(
        [
            pl.lit("CONTRATO NUEVO - SIN MATCH ONErpm").alias("contract_segment"),
            pl.lit("sin_match_asumido_nuevo").alias("match_method"),
            pl.lit(None, dtype=pl.Utf8).alias("onerpm_first_month"),
            pl.lit(None, dtype=pl.Utf8).alias("onerpm_title"),
        ]
    )
    if raw.height == 0 or onerpm_map.empty:
        return result

    for method, col in [("isrc", "asset_isrc"), ("title", "_title_key"), ("signature", "_signature_key")]:
        mapping = onerpm_map[onerpm_map["match_method"] == method].copy()
        if mapping.empty:
            continue
        mapping = mapping.drop_duplicates("match_key", keep="first")
        map_df = pl.from_pandas(mapping[["match_key", "contract_segment", "onerpm_first_month", "onerpm_title"]]).rename(
            {
                "contract_segment": f"_segment_{method}",
                "onerpm_first_month": f"_first_{method}",
                "onerpm_title": f"_title_{method}",
            }
        )
        left = result.with_row_index("_row_id")
        joined = left.join(map_df, left_on=col, right_on="match_key", how="left")
        result = (
            joined.with_columns(
                [
                    pl.when((pl.col("match_method") == "sin_match_asumido_nuevo") & pl.col(f"_segment_{method}").is_not_null())
                    .then(pl.col(f"_segment_{method}"))
                    .otherwise(pl.col("contract_segment"))
                    .alias("contract_segment"),
                    pl.when((pl.col("match_method") == "sin_match_asumido_nuevo") & pl.col(f"_segment_{method}").is_not_null())
                    .then(pl.lit(method))
                    .otherwise(pl.col("match_method"))
                    .alias("match_method"),
                    pl.when((pl.col("match_method") == "sin_match_asumido_nuevo") & pl.col(f"_segment_{method}").is_not_null())
                    .then(pl.col(f"_first_{method}"))
                    .otherwise(pl.col("onerpm_first_month"))
                    .alias("onerpm_first_month"),
                    pl.when((pl.col("match_method") == "sin_match_asumido_nuevo") & pl.col(f"_segment_{method}").is_not_null())
                    .then(pl.col(f"_title_{method}"))
                    .otherwise(pl.col("onerpm_title"))
                    .alias("onerpm_title"),
                ]
            )
            .drop(["_row_id", f"_segment_{method}", f"_first_{method}", f"_title_{method}"])
        )

    result = apply_soft_matches(result, onerpm_map)
    return result


def token_overlap(left: str, right: str) -> int:
    return len(set(str(left or "").split()) & set(str(right or "").split()))


def report_specific_gusty_match(title_key: str) -> dict | None:
    """Capture Gusty video titles whose ONErpm reference is known by series number."""
    session = re.search(r"\bsession\s+en\s+el\s+barrio\s+(\d+)\b|\bsession\s+barrio\s+(\d+)\b", title_key)
    if session:
        number = int(next(group for group in session.groups() if group))
        if number in {1, 2, 4, 5, 6, 7, 8, 15}:
            return {
                "contract_segment": "CONTRATO VIEJO",
                "match_method": "signature",
                "onerpm_first_month": "",
                "onerpm_title": f"Session en el Barrio #{number}",
            }
        if number in {9, 11, 12, 13, 14}:
            return {
                "contract_segment": "CONTRATO NUEVO",
                "match_method": "signature",
                "onerpm_first_month": "",
                "onerpm_title": f"Session en el Barrio #{number}",
            }

    toma = re.search(r"\btoma\s+(\d+)\b", title_key)
    if toma:
        number = int(toma.group(1))
        if number in {1, 4, 5, 6, 7, 8, 9}:
            return {
                "contract_segment": "CONTRATO VIEJO",
                "match_method": "signature",
                "onerpm_first_month": "",
                "onerpm_title": f"Toma {number}",
            }
    return None


def best_soft_match(title_key: str, signature_key: str, candidates: pd.DataFrame) -> dict | None:
    title_key = str(title_key or "")
    signature_key = str(signature_key or "")

    specific = report_specific_gusty_match(title_key)
    if specific:
        return specific

    signature_candidates = candidates[candidates["match_method"] == "signature"].copy()
    signature_hits = []
    for row in signature_candidates.itertuples(index=False):
        key = str(getattr(row, "match_key") or "")
        if len(key) < 8:
            continue
        if key in signature_key or key in title_key:
            signature_hits.append((len(key), float(getattr(row, "onerpm_total_usd", 0) or 0), row))
    if signature_hits:
        row = sorted(signature_hits, key=lambda item: (item[0], item[1]), reverse=True)[0][2]
        return row._asdict()

    title_candidates = candidates[candidates["match_method"] == "title"].copy()
    title_hits = []
    for row in title_candidates.itertuples(index=False):
        key = str(getattr(row, "match_key") or "")
        if len(key) < 8:
            continue
        if key in title_key or title_key in key:
            title_hits.append((len(key), float(getattr(row, "onerpm_total_usd", 0) or 0), row))
    if title_hits:
        row = sorted(title_hits, key=lambda item: (item[0], item[1]), reverse=True)[0][2]
        data = row._asdict()
        data["match_method"] = "title_fuzzy"
        return data

    best: tuple[float, float, object] | None = None
    for row in title_candidates.itertuples(index=False):
        key = str(getattr(row, "match_key") or "")
        if len(key) < 8 or token_overlap(title_key, key) < 3:
            continue
        score = SequenceMatcher(None, title_key, key).ratio()
        if score < 0.86:
            continue
        total = float(getattr(row, "onerpm_total_usd", 0) or 0)
        if best is None or (score, total) > (best[0], best[1]):
            best = (score, total, row)
    if best:
        data = best[2]._asdict()
        data["match_method"] = "title_fuzzy"
        return data
    return None


def apply_soft_matches(result: pl.DataFrame, onerpm_map: pd.DataFrame) -> pl.DataFrame:
    remaining = (
        result
        .filter(pl.col("match_method") == "sin_match_asumido_nuevo")
        .select(["_title_key", "_signature_key"])
        .unique()
        .to_pandas()
    )
    if remaining.empty:
        return result

    matches: list[dict] = []
    candidates = onerpm_map.copy()
    for row in remaining.itertuples(index=False):
        row_data = row._asdict()
        title_key = str(row_data.get("_title_key") or row_data.get("_0") or "")
        signature_key = str(row_data.get("_signature_key") or row_data.get("_1") or "")
        match = best_soft_match(title_key, signature_key, candidates)
        if not match:
            continue
        matches.append(
            {
                "_title_key": title_key,
                "_signature_key": signature_key,
                "_soft_contract_segment": match["contract_segment"],
                "_soft_match_method": match["match_method"],
                "_soft_onerpm_first_month": match["onerpm_first_month"],
                "_soft_onerpm_title": match["onerpm_title"],
            }
        )
    if not matches:
        return result

    soft_df = pl.from_pandas(pd.DataFrame(matches))
    joined = result.join(soft_df, on=["_title_key", "_signature_key"], how="left")
    return (
        joined.with_columns(
            [
                pl.when((pl.col("match_method") == "sin_match_asumido_nuevo") & pl.col("_soft_contract_segment").is_not_null())
                .then(pl.col("_soft_contract_segment"))
                .otherwise(pl.col("contract_segment"))
                .alias("contract_segment"),
                pl.when((pl.col("match_method") == "sin_match_asumido_nuevo") & pl.col("_soft_match_method").is_not_null())
                .then(pl.col("_soft_match_method"))
                .otherwise(pl.col("match_method"))
                .alias("match_method"),
                pl.when((pl.col("match_method") == "sin_match_asumido_nuevo") & pl.col("_soft_onerpm_first_month").is_not_null())
                .then(pl.col("_soft_onerpm_first_month"))
                .otherwise(pl.col("onerpm_first_month"))
                .alias("onerpm_first_month"),
                pl.when((pl.col("match_method") == "sin_match_asumido_nuevo") & pl.col("_soft_onerpm_title").is_not_null())
                .then(pl.col("_soft_onerpm_title"))
                .otherwise(pl.col("onerpm_title"))
                .alias("onerpm_title"),
            ]
        )
        .drop(["_soft_contract_segment", "_soft_match_method", "_soft_onerpm_first_month", "_soft_onerpm_title"])
    )


def group_sum(df: pl.DataFrame, by: list[str], title_col: str = "tema") -> pd.DataFrame:
    if df.height == 0:
        return pd.DataFrame(columns=by + ["ingresos_usd", "ingresos_eur", "unidades", "temas", "filas_raw"])
    return (
        df.group_by(by)
        .agg(
            [
                pl.sum("amount_usd").alias("ingresos_usd"),
                pl.sum("ingresos_eur").alias("ingresos_eur"),
                pl.sum("units").alias("unidades"),
                pl.n_unique(title_col).alias("temas"),
                pl.len().alias("filas_raw"),
            ]
        )
        .sort("ingresos_usd", descending=True)
        .to_pandas()
    )


def build_listado(df: pl.DataFrame) -> pd.DataFrame:
    if df.height == 0:
        return pd.DataFrame()
    return (
        df.group_by(
            [
                "contract_segment",
                "match_method",
                "source",
                "account",
                "asset_isrc",
                "tema",
                "asset_title_statement",
                "artist_statement_style",
                "asset_artist_statement",
                "tipo_contenido",
            ]
        )
        .agg(
            [
                pl.sum("amount_usd").alias("ingresos_usd"),
                pl.sum("ingresos_eur").alias("ingresos_eur"),
                pl.sum("units").alias("unidades"),
                pl.min("transaction_month").alias("desde"),
                pl.max("transaction_month").alias("hasta"),
                pl.n_unique("statement_period").alias("statements"),
                pl.len().alias("filas_raw"),
            ]
        )
        .sort("ingresos_usd", descending=True)
        .to_pandas()
    )


def style_workbook(wb) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                header = str(ws.cell(row=1, column=cell.column).value or "").lower()
                if any(token in header for token in ["usd", "eur", "ingresos"]):
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '$#,##0.00'
                elif any(token in header for token in ["unidades", "temas", "filas", "statements"]):
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0'
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = 0
            for cell in col[:500]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 48))


def build_fuga_gusty_contract_report(
    start_month: str | None = None,
    end_month: str | None = "2026-03",
    output_dir: Path = REPORTS,
    raw_path: Path = RAW_PATH,
) -> Path:
    raw = prepare_base(raw_path, start_month, end_month)
    if raw.height:
        raw = filter_reportable_catalog(raw.lazy(), set(raw.columns)).collect()
    onerpm_map = build_onerpm_map(raw_path)
    df = classify_fuga(raw, onerpm_map)

    title = "Gusty Fuga contratos nuevo & viejo"
    criterio = (
        "Filas FUGA donde artista/titulo contiene Gusty. "
        f"Corte statement_period {start_month or 'inicio'} a {end_month or 'ultimo'}. "
        "Clasificacion contractual segun mapa ONErpm/Motorcito."
    )
    overview = pd.DataFrame(
        [
            {"indicador": "Reporte", "valor": title},
            {"indicador": "Criterio", "valor": criterio},
            {"indicador": "Fuente", "valor": "fuga"},
            {"indicador": "Cuenta", "valor": "indyana_records"},
            {"indicador": "Corte statement", "valor": f"{start_month or 'inicio'} a {end_month or 'ultimo'}"},
            {"indicador": "Ingresos USD", "valor": float(df["amount_usd"].sum()) if df.height else 0.0},
            {"indicador": "Ingresos EUR", "valor": float(df["ingresos_eur"].sum()) if df.height else 0.0},
            {"indicador": "Unidades", "valor": float(df["units"].sum()) if df.height else 0.0},
            {"indicador": "Filas raw", "valor": df.height},
            {"indicador": "Temas aprox", "valor": df.select("tema").unique().height if df.height else 0},
            {"indicador": "Desde transaction", "valor": df["transaction_month"].min() if df.height else ""},
            {"indicador": "Hasta transaction", "valor": df["transaction_month"].max() if df.height else ""},
        ]
    )

    detalle_group_cols = [
        "contract_segment",
        "match_method",
        "statement_period",
        "transaction_month",
        "asset_isrc",
        "tema",
        "asset_title_statement",
        "artist_statement_style",
        "asset_artist_statement",
        "tipo_contenido",
        "dsp",
        "store_name",
        "territory",
        "sale_type",
        "sale_user_type",
    ]
    detalle_group_cols = [col for col in detalle_group_cols if col in df.columns]
    if df.height:
        detalle = (
            df.group_by(detalle_group_cols)
            .agg(
                [
                    pl.sum("amount_usd").alias("ingresos_usd"),
                    pl.sum("ingresos_eur").alias("ingresos_eur"),
                    pl.sum("units").alias("unidades"),
                    pl.len().alias("filas_raw"),
                ]
            )
            .sort(["contract_segment", "match_method", "statement_period", "transaction_month", "ingresos_usd"], descending=[False, False, False, False, True])
            .to_pandas()
        )
    else:
        detalle = pd.DataFrame(columns=detalle_group_cols + ["ingresos_usd", "ingresos_eur", "unidades", "filas_raw"])

    tables = {
        "Resumen": overview,
        "Contrato Summary": group_sum(df, ["contract_segment", "match_method"]),
        "Listado": build_listado(df),
        "Mensual": group_sum(df, ["contract_segment", "statement_period", "transaction_month"]),
        "Store Summary": group_sum(df, ["contract_segment", "dsp", "store_name", "sale_type", "sale_user_type"]),
        "Territory Summary": group_sum(df, ["contract_segment", "territory"]),
        "Content Type": group_sum(df, ["contract_segment", "tipo_contenido"]),
        "Statement Summary": group_sum(df, ["contract_segment", "statement_period", "statement_type", "statement_file_name"]),
        "Detalle": detalle,
    }

    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = output_dir / f"gusty_fuga_contratos_nuevo_viejo_{start_month or 'inicio'}_a_{end_month or 'ultimo'}_{timestamp}.xlsx"
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, table in tables.items():
            table.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        style_workbook(writer.book)
    return output


if __name__ == "__main__":
    print(build_fuga_gusty_contract_report())
