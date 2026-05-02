from pathlib import Path

import pandas as pd
import polars as pl

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


MARTS_DIR = Path(r"C:\royalties_pipeline\warehouse\marts")
OUTPUT_PATH = Path(r"C:\royalties_pipeline\reports\reporte_ingresos_digitales_por_mes_de_statement_marts.xlsx")

FUGA_STATEMENT_FACTOR = 0.977832


STANDARDIZED_FILES = [
    MARTS_DIR / "standardized_raw_dashgo.parquet",
    MARTS_DIR / "standardized_raw_fuga.parquet",
    MARTS_DIR / "standardized_raw_onerpm.parquet",
    MARTS_DIR / "standardized_raw_orchard.parquet",
    MARTS_DIR / "standardized_raw_soundon.parquet",
]


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ARTIST_FILL = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
TITLE_FILL = PatternFill(start_color="EAF2F8", end_color="EAF2F8", fill_type="solid")
SHARE_ALERT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

TITLE_FONT = Font(size=16, bold=True, color="1F4E78")
TOTAL_FONT = Font(bold=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")


def has_col(schema: dict[str, pl.DataType], col: str) -> bool:
    return col in schema


def col_or_null(schema: dict[str, pl.DataType], col: str, dtype=pl.Utf8) -> pl.Expr:
    if has_col(schema, col):
        return pl.col(col).cast(dtype, strict=False)
    return pl.lit(None).cast(dtype)


def amount_expr(schema: dict[str, pl.DataType]) -> pl.Expr:
    candidates = []

    for col in ["amount_usd", "net_amount_usd", "net_amount"]:
        if has_col(schema, col):
            candidates.append(pl.col(col).cast(pl.Float64, strict=False))

    if not candidates:
        return pl.lit(0.0)

    return pl.coalesce(candidates)


def aggregate_file(path: Path) -> tuple[pl.DataFrame, pl.DataFrame]:
    if not path.exists():
        print(f"  - No existe: {path.name}. Se omite.")
        return pl.DataFrame(), pl.DataFrame()

    print(f"  - Agregando {path.name}")

    lf = pl.scan_parquet(path)
    schema = lf.collect_schema()

    source = col_or_null(schema, "source")

    amount = amount_expr(schema)
    adjusted_amount = (
        pl.when(source == "fuga")
        .then(amount * FUGA_STATEMENT_FACTOR)
        .otherwise(amount)
    )

    has_share_expr = (
        pl.col("has_share_in_out").cast(pl.Boolean, strict=False)
        if has_col(schema, "has_share_in_out")
        else pl.lit(False)
    )

    base = lf.with_columns([
        source.alias("_source"),
        col_or_null(schema, "account").alias("_account"),
        col_or_null(schema, "statement_period").alias("_statement_period"),
        col_or_null(schema, "artist_statement_style").alias("_artist_raw"),
        col_or_null(schema, "source_sheet").alias("_source_sheet"),
        adjusted_amount.alias("_amount"),
        has_share_expr.alias("_has_share_in_out"),
    ])

    base = base.filter(
        pl.col("_statement_period").is_not_null()
        & (pl.col("_statement_period").str.strip_chars() != "")
    )

    if path.name == "standardized_raw_soundon.parquet":
        base = base.filter(pl.col("_source_sheet") == "my_royalty")

    totals = (
        base
        .with_columns(
            pl.when(
                pl.col("_artist_raw").is_null()
                | (pl.col("_artist_raw").str.strip_chars() == "")
            )
            .then(pl.lit("SIN ARTISTA"))
            .otherwise(pl.col("_artist_raw").str.strip_chars())
            .alias("artist")
        )
        .group_by(["_source", "_account", "artist", "_statement_period"])
        .agg([
            pl.sum("_amount").round(2).alias("total"),
            pl.max("_has_share_in_out").cast(pl.Int64).alias("has_share_in_out"),
        ])
        .rename({
            "_source": "source",
            "_account": "account",
            "_statement_period": "statement_period",
        })
        .collect()
    )

    fuga_eur = pl.DataFrame()
    if path.name == "standardized_raw_fuga.parquet" and has_col(schema, "net_amount"):
        fuga_eur = (
            lf
            .filter(
                pl.col("statement_period").is_not_null()
                & (pl.col("statement_period").cast(pl.Utf8).str.strip_chars() != "")
            )
            .group_by(["source", "account", "statement_period"])
            .agg(pl.sum("net_amount").round(2).alias("total_eur"))
            .collect()
        )

    return totals, fuga_eur


def format_sheet(ws, pivot, company_name, share_flags=None):
    max_col = len(pivot.columns) + 1
    max_row = len(pivot) + 2

    for col in range(2, max_col + 1):
        ws.cell(row=1, column=col).value = None

    ws["A1"] = company_name
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = LEFT_ALIGN
    ws.row_dimensions[1].height = 24

    ws["A2"] = "ARTISTA"
    ws["A2"].fill = HEADER_FILL
    ws["A2"].font = HEADER_FONT
    ws["A2"].alignment = CENTER_ALIGN

    for col_idx, col_name in enumerate(pivot.columns, start=2):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = col_name
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN

    for row in range(3, max_row + 1):
        label = str(ws.cell(row=row, column=1).value or "")
        artist_cell = ws.cell(row=row, column=1)
        artist_cell.fill = ARTIST_FILL
        artist_cell.alignment = LEFT_ALIGN

        is_total_row = label.upper().startswith("TOTAL")

        if is_total_row:
            artist_cell.fill = TOTAL_FILL
            artist_cell.font = TOTAL_FONT

        for col in range(2, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.number_format = '#,##0.00'
            cell.alignment = CENTER_ALIGN

            if is_total_row:
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT

    if share_flags is not None and not share_flags.empty:
        for artist_idx, artist_name in enumerate(pivot.index, start=3):
            if str(artist_name).upper().startswith("TOTAL"):
                continue

            if artist_name not in share_flags.index:
                continue

            for col_idx, period in enumerate(pivot.columns, start=2):
                if period == "TOTAL" or period not in share_flags.columns:
                    continue

                try:
                    has_alert = int(share_flags.loc[artist_name, period]) == 1
                except Exception:
                    has_alert = False

                if has_alert:
                    ws.cell(row=artist_idx, column=col_idx).fill = SHARE_ALERT_FILL

    total_col = max_col
    for row in range(2, max_row + 1):
        cell = ws.cell(row=row, column=total_col)
        if row == 2:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        else:
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT

    ws.column_dimensions["A"].width = 30
    for col in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A2:{get_column_letter(max_col)}{max_row}"


def add_totals_and_clean(pivot):
    pivot = pivot.reindex(sorted(pivot.columns), axis=1)
    pivot["TOTAL"] = pivot.sum(axis=1)

    before_rows = len(pivot)
    pivot = pivot[pivot["TOTAL"].round(2) != 0]
    removed_rows = before_rows - len(pivot)

    return pivot, removed_rows


def main():
    print("Generando reporte ingresos por statement desde marts...")

    frames = []
    fuga_eur_frames = []

    for path in STANDARDIZED_FILES:
        totals, fuga_eur = aggregate_file(path)

        if totals.height > 0:
            frames.append(totals)

        if fuga_eur.height > 0:
            fuga_eur_frames.append(fuga_eur)

    if not frames:
        print("ERROR: No hay datos para generar el reporte.")
        raise SystemExit(1)

    df = pl.concat(frames, how="diagonal_relaxed").to_pandas()
    fuga_eur_df = (
        pl.concat(fuga_eur_frames, how="diagonal_relaxed").to_pandas()
        if fuga_eur_frames
        else pd.DataFrame(columns=["source", "account", "statement_period", "total_eur"])
    )

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        pivot_total = df.pivot_table(
            index="artist",
            columns="statement_period",
            values="total",
            aggfunc="sum",
            fill_value=0,
        )

        pivot_total, _ = add_totals_and_clean(pivot_total)

        if not pivot_total.empty:
            total_general = pivot_total.sum(axis=0)
            total_general.name = "TOTAL USD"
            pivot_total = pd.concat([pivot_total, total_general.to_frame().T])
            pivot_total.to_excel(writer, sheet_name="TOTAL", startrow=1)
            format_sheet(writer.sheets["TOTAL"], pivot_total, "SALDO TOTAL ARTISTAS")

        for (source, account), group in df.groupby(["source", "account"], dropna=False):
            pivot = group.pivot_table(
                index="artist",
                columns="statement_period",
                values="total",
                aggfunc="sum",
                fill_value=0,
            )

            pivot, removed_rows = add_totals_and_clean(pivot)

            if removed_rows > 0:
                print(f"  - {source}/{account}: artistas eliminados por TOTAL = 0: {removed_rows}")

            if pivot.empty:
                print(f"  - {source}/{account}: sin artistas con total distinto de 0. Se omite hoja.")
                continue

            total_usd = pivot.sum(axis=0)
            total_usd.name = "TOTAL USD"
            rows_to_add = [total_usd.to_frame().T]

            if source == "fuga":
                matches = fuga_eur_df[
                    (fuga_eur_df["source"] == source)
                    & (fuga_eur_df["account"] == account)
                ]

                total_eur = pd.Series(0.0, index=pivot.columns)
                total_eur.name = "TOTAL EUR"

                for col in pivot.columns:
                    if col == "TOTAL":
                        continue

                    period_matches = matches.loc[matches["statement_period"] == col, "total_eur"]

                    if not period_matches.empty:
                        total_eur[col] = float(period_matches.iloc[0])

                total_eur["TOTAL"] = total_eur.drop(labels=["TOTAL"], errors="ignore").sum()
                rows_to_add.append(total_eur.to_frame().T)

            share_flags = None

            if source == "onerpm" and account == "gusty_dj" and "has_share_in_out" in group.columns:
                share_flags = group.pivot_table(
                    index="artist",
                    columns="statement_period",
                    values="has_share_in_out",
                    aggfunc="max",
                    fill_value=0,
                )

                share_flags = share_flags.reindex(
                    index=pivot.index,
                    columns=[col for col in pivot.columns if col != "TOTAL"],
                    fill_value=0,
                )

                flagged_cells = int(share_flags.sum().sum()) if not share_flags.empty else 0
                print(f"  - onerpm/gusty_dj: celdas con alerta Shares In/Out: {flagged_cells}")

            pivot = pd.concat([pivot] + rows_to_add)

            sheet_name = f"{source}_{account}"[:31]
            pivot.to_excel(writer, sheet_name=sheet_name, startrow=1)
            format_sheet(writer.sheets[sheet_name], pivot, f"{str(source).upper()} - {account}", share_flags)

    print("Reporte generado en:")
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
