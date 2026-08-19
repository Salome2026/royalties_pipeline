from __future__ import annotations

import argparse
import shutil
import re
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
import polars as pl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from lib.catalog_report_filter import filter_reportable_generation
except ModuleNotFoundError:
    from scripts.lib.catalog_report_filter import filter_reportable_generation


BASE = Path(r"C:\royalties_pipeline")
MARTS_DIR = BASE / "warehouse" / "marts"
REPORTS_DIR = BASE / "reports" / "api"
DOWNLOADS_DIR = Path(r"C:\Users\ruben\Downloads")

RAW_ALL_PATH = MARTS_DIR / "standardized_raw_all_sources.parquet"
CATALOG_MASTER_PATH = MARTS_DIR / "catalog_master.parquet"

CLIENT_NET_RATE = 0.80
TITLE_PATTERN = "la juntada de los artistas"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOTAL_FILL = PatternFill("solid", fgColor="D9EAF7")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
TITLE_FONT = Font(size=14, bold=True, color="1F4E78")
LEFT = Alignment(horizontal="left", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")

ILLEGAL_EXCEL_CHARS = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(text.split())


def has_col(schema: dict[str, pl.DataType], col: str) -> bool:
    return col in schema


def col_text(schema: dict[str, pl.DataType], col: str) -> pl.Expr:
    if has_col(schema, col):
        return pl.col(col).cast(pl.Utf8, strict=False).fill_null("")
    return pl.lit("")


def first_non_blank(schema: dict[str, pl.DataType], columns: list[str]) -> pl.Expr:
    exprs: list[pl.Expr] = []
    for col in columns:
        if has_col(schema, col):
            text = pl.col(col).cast(pl.Utf8, strict=False).str.strip_chars()
            exprs.append(pl.when((text.is_not_null()) & (text != "")).then(text).otherwise(None))
    if not exprs:
        return pl.lit(None).cast(pl.Utf8)
    return pl.coalesce(exprs)


def units_expr(schema: dict[str, pl.DataType]) -> pl.Expr:
    candidates = [
        pl.col(col).cast(pl.Float64, strict=False)
        for col in [
            "Quantity",
            "units",
            "Units",
            "streams",
            "Product Quantity",
            "Asset Quantity",
            "product_quantity_num",
            "asset_quantity_num",
            "QUANTITY",
        ]
        if has_col(schema, col)
    ]
    return pl.coalesce(candidates) if candidates else pl.lit(0.0)


def load_la_juntada_catalog_keys() -> tuple[set[str], set[str]]:
    if not CATALOG_MASTER_PATH.exists():
        return set(), set()
    catalog = pl.read_parquet(CATALOG_MASTER_PATH)
    schema = catalog.schema
    text = pl.concat_str(
        [
            col_text(schema, "track_title"),
            col_text(schema, "artist_statement"),
            col_text(schema, "identity_title"),
            col_text(schema, "identity_artist"),
        ],
        separator=" | ",
    ).str.to_lowercase()
    matched = catalog.with_columns(text.alias("_match_text")).filter(
        pl.col("_match_text").str.contains(TITLE_PATTERN)
    )
    isrcs: set[str] = set()
    videos: set[str] = set()
    if "asset_isrc" in matched.columns:
        isrcs = {
            str(value).strip().upper()
            for value in matched.select("asset_isrc").drop_nulls().to_series().to_list()
            if str(value).strip()
        }
    if "video_id" in matched.columns:
        videos = {
            str(value).strip()
            for value in matched.select("video_id").drop_nulls().to_series().to_list()
            if str(value).strip()
        }
    return isrcs, videos


def prepared_rows(end_month: str | None = None) -> pl.DataFrame:
    raw = pl.read_parquet(RAW_ALL_PATH)
    schema = raw.schema
    isrcs, videos = load_la_juntada_catalog_keys()

    title_artist_text = pl.concat_str(
        [
            col_text(schema, "artist_statement_style"),
            col_text(schema, "asset_artist_statement"),
            col_text(schema, "product_artist_statement"),
            col_text(schema, "Track Artist"),
            col_text(schema, "Artist"),
            col_text(schema, "ACCOUNT NAME"),
            col_text(schema, "asset_title_statement"),
            col_text(schema, "track_statement_style"),
            col_text(schema, "Track Title"),
            col_text(schema, "Title"),
            col_text(schema, "Product Title"),
            col_text(schema, "Asset Title"),
        ],
        separator=" | ",
    ).str.to_lowercase()
    isrc_text = pl.concat_str(
        [col_text(schema, "asset_isrc"), col_text(schema, "ISRC"), col_text(schema, "Asset ISRC")],
        separator="|",
    ).str.to_uppercase()
    video_text = pl.concat_str(
        [col_text(schema, "video_id"), col_text(schema, "VideoId"), col_text(schema, "Video ID"), col_text(schema, "YOUTUBE VIDEO ID")],
        separator="|",
    )

    filtered = raw.with_columns(
        [
            title_artist_text.alias("_match_text"),
            isrc_text.alias("_isrc_text"),
            video_text.alias("_video_text"),
        ]
    ).filter(
        pl.col("_match_text").str.contains(TITLE_PATTERN)
        | pl.col("_isrc_text").str.split("|").list.eval(pl.element().is_in(list(isrcs))).list.any()
        | pl.col("_video_text").str.split("|").list.eval(pl.element().is_in(list(videos))).list.any()
    )
    filtered = filter_reportable_generation(filtered.lazy(), set(raw.columns)).collect()
    if end_month:
        filtered = filtered.filter(col_text(schema, "statement_period") <= end_month)

    source = col_text(schema, "source")
    amount_real = pl.col("amount_usd").cast(pl.Float64, strict=False).fill_null(0.0)
    adjusted = amount_real * CLIENT_NET_RATE
    gross_usd = pl.when(source == "fuga").then(
        pl.col("converted_gross_income_num").cast(pl.Float64, strict=False).fill_null(0.0)
        * pl.col("fx_eur_usd_rate").cast(pl.Float64, strict=False).fill_null(1.0)
    ).otherwise(None)

    out = filtered.with_columns(
        [
            first_non_blank(schema, ["source"]).alias("Distribuidora"),
            first_non_blank(schema, ["account"]).alias("Cuenta"),
            first_non_blank(schema, ["statement_period"]).alias("Statement"),
            first_non_blank(schema, ["transaction_month", "Transaction Month"]).alias("Transaction"),
            first_non_blank(schema, ["source_sheet", "statement_type"]).alias("Hoja origen"),
            first_non_blank(schema, ["asset_title_statement", "track_statement_style", "Track Title", "Title", "Asset Title", "Product Title"]).alias("Tema"),
            first_non_blank(schema, ["asset_artist_statement", "artist_statement_style", "Track Artist", "Artist", "Asset Artist", "Product Artist"]).alias("Artista"),
            first_non_blank(schema, ["asset_isrc", "ISRC", "Asset ISRC"]).alias("ISRC"),
            first_non_blank(schema, ["video_id", "VideoId", "Video ID", "YOUTUBE VIDEO ID"]).alias("Video ID"),
            first_non_blank(schema, ["territory", "Territory", "Country", "COUNTRY"]).alias("Pais"),
            first_non_blank(schema, ["dsp_normalized", "dsp", "DSP", "Store", "STORE"]).alias("DSP"),
            first_non_blank(schema, ["store_report_label", "store_name", "Sale Store Name", "Store Name", "service_detail", "SERVICE DETAIL"]).alias("Store"),
            first_non_blank(schema, ["monetization_normalized"]).alias("Monetizacion"),
            first_non_blank(schema, ["content_origin_normalized"]).alias("Origen"),
            first_non_blank(schema, ["plan_normalized"]).alias("Plan"),
            units_expr(schema).fill_null(0.0).alias("Unidades"),
            gross_usd.alias("Gross FUGA USD"),
            amount_real.alias("Ingreso real USD"),
            adjusted.alias("Ingreso ajustado USD"),
            (adjusted - amount_real).alias("Ajuste comision USD"),
        ]
    ).select(
        [
            "Distribuidora",
            "Cuenta",
            "Statement",
            "Transaction",
            "Hoja origen",
            "Tema",
            "Artista",
            "ISRC",
            "Video ID",
            "Pais",
            "DSP",
            "Store",
            "Monetizacion",
            "Origen",
            "Plan",
            "Unidades",
            "Gross FUGA USD",
            "Ingreso real USD",
            "Ingreso ajustado USD",
            "Ajuste comision USD",
        ]
    )
    return out


def to_pandas_grouped(df: pl.DataFrame, group_cols: list[str]) -> pd.DataFrame:
    return (
        df.group_by(group_cols)
        .agg(
            [
                pl.col("Unidades").sum(),
                pl.col("Gross FUGA USD").sum(),
                pl.col("Ingreso real USD").sum(),
                pl.col("Ingreso ajustado USD").sum(),
                pl.col("Ajuste comision USD").sum(),
                pl.len().alias("Filas"),
            ]
        )
        .sort("Ingreso ajustado USD", descending=True)
        .to_pandas()
    )


def summary_tables(df: pl.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    by_distributor = to_pandas_grouped(df, ["Distribuidora"])
    by_account = to_pandas_grouped(df, ["Distribuidora", "Cuenta"])
    by_month = (
        df.group_by(["Statement", "Distribuidora"])
        .agg(
            [
                pl.col("Unidades").sum(),
                pl.col("Gross FUGA USD").sum(),
                pl.col("Ingreso real USD").sum(),
                pl.col("Ingreso ajustado USD").sum(),
                pl.col("Ajuste comision USD").sum(),
            ]
        )
        .sort(["Statement", "Distribuidora"])
        .to_pandas()
    )
    by_title = (
        df.group_by(["Tema", "Artista", "ISRC", "Video ID"])
        .agg(
            [
                pl.col("Statement").min().alias("Primer statement"),
                pl.col("Statement").max().alias("Ultimo statement"),
                pl.col("Unidades").sum(),
                pl.col("Gross FUGA USD").sum(),
                pl.col("Ingreso real USD").sum(),
                pl.col("Ingreso ajustado USD").sum(),
                pl.col("Ajuste comision USD").sum(),
                pl.len().alias("Filas"),
            ]
        )
        .sort("Ingreso ajustado USD", descending=True)
        .to_pandas()
    )
    by_country_dsp = (
        df.group_by(["Tema", "Artista", "ISRC", "Video ID", "Pais", "DSP", "Store", "Monetizacion", "Origen", "Plan"])
        .agg(
            [
                pl.col("Statement").min().alias("Primer statement"),
                pl.col("Statement").max().alias("Ultimo statement"),
                pl.col("Unidades").sum(),
                pl.col("Gross FUGA USD").sum(),
                pl.col("Ingreso real USD").sum(),
                pl.col("Ingreso ajustado USD").sum(),
                pl.col("Ajuste comision USD").sum(),
                pl.len().alias("Filas"),
            ]
        )
        .sort("Ingreso ajustado USD", descending=True)
        .to_pandas()
    )
    return by_distributor, by_account, by_month, by_title, by_country_dsp


def make_summary_sheet(df: pl.DataFrame, by_distributor: pd.DataFrame) -> pd.DataFrame:
    if df.is_empty():
        return pd.DataFrame([{"Indicador": "Sin datos", "Valor": 0}])
    first_statement = df.select(pl.col("Statement").min()).item()
    last_statement = df.select(pl.col("Statement").max()).item()
    rows = [
        {"Indicador": "Proyecto", "Valor": "La Juntada de los Artistas"},
        {"Indicador": "Periodo statement", "Valor": f"{first_statement} a {last_statement}"},
        {"Indicador": "Ingreso USD", "Valor": float(df.select(pl.col("Ingreso ajustado USD").sum()).item())},
        {"Indicador": "Unidades", "Valor": float(df.select(pl.col("Unidades").sum()).item())},
        {"Indicador": "Distribuidoras", "Valor": ", ".join(by_distributor["Distribuidora"].astype(str).tolist())},
    ]
    return pd.DataFrame(rows)


def clean_excel_text(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in out.columns:
        if out[col].dtype == "object":
            out[col] = out[col].map(
                lambda value: ILLEGAL_EXCEL_CHARS.sub("", value) if isinstance(value, str) else value
            )
    return out


def client_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in ["Gross FUGA USD", "Ingreso real USD", "Ajuste comision USD"]:
        if col in out.columns:
            out = out.drop(columns=[col])
    if "Ingreso ajustado USD" in out.columns:
        out = out.rename(columns={"Ingreso ajustado USD": "Ingreso USD"})
    return clean_excel_text(out)


def write_client_workbook(path: Path, sheets: list[tuple[str, pd.DataFrame]]) -> None:
    with pd.ExcelWriter(path, engine="xlsxwriter") as writer:
        workbook = writer.book
        header_fmt = workbook.add_format({
            "bold": True,
            "font_color": "white",
            "bg_color": "#1F4E78",
            "align": "center",
            "valign": "vcenter",
            "border": 1,
        })
        money_fmt = workbook.add_format({"num_format": "$#,##0.00", "border": 1})
        units_fmt = workbook.add_format({"num_format": "#,##0", "border": 1})
        text_fmt = workbook.add_format({"border": 1})
        title_fmt = workbook.add_format({"bold": True, "font_color": "#1F4E78", "font_size": 12})

        for sheet_name, frame in sheets:
            frame = clean_excel_text(frame)
            frame.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]
            rows, cols = frame.shape
            worksheet.hide_gridlines(2)
            if rows > 0 and cols > 0:
                worksheet.autofilter(0, 0, rows, cols - 1)
                if sheet_name != "Resumen":
                    worksheet.freeze_panes(1, 0)
            for col_idx, col_name in enumerate(frame.columns):
                worksheet.write(0, col_idx, col_name, header_fmt)
                sample = frame[col_name].head(300).map(lambda value: len(str(value)) if value is not None else 0).max() if rows else 0
                width = max(12, min(max(len(str(col_name)), int(sample or 0)) + 2, 46))
                fmt = text_fmt
                if "USD" in str(col_name) or col_name == "Valor":
                    fmt = money_fmt
                elif "Unidades" in str(col_name) or col_name == "Filas":
                    fmt = units_fmt
                worksheet.set_column(col_idx, col_idx, width, fmt)
            if sheet_name == "Resumen":
                worksheet.set_column(0, 0, 24, title_fmt)
                worksheet.set_column(1, 1, 34, money_fmt)


def build_report(output_dir: Path = REPORTS_DIR, end_month: str | None = None) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    df = prepared_rows(end_month=end_month)
    by_distributor, by_account, by_month, by_title, by_country_dsp = summary_tables(df)
    summary = make_summary_sheet(df, by_distributor)
    detail = (
        df.group_by(["Distribuidora", "Cuenta", "Statement", "Transaction", "Hoja origen", "Tema", "Artista", "ISRC", "Video ID", "Pais", "DSP", "Store"])
        .agg(
            [
                pl.col("Unidades").sum(),
                pl.col("Gross FUGA USD").sum(),
                pl.col("Ingreso real USD").sum(),
                pl.col("Ingreso ajustado USD").sum(),
                pl.col("Ajuste comision USD").sum(),
                pl.len().alias("Filas"),
            ]
        )
        .sort(["Statement", "Distribuidora", "Ingreso ajustado USD"], descending=[False, False, True])
        .to_pandas()
    )

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    last_statement = str(df.select(pl.col("Statement").max()).item() or end_month or "ultimo")
    output = output_dir / f"la_juntada_artistas_cliente_inicio_a_{last_statement}_{stamp}.xlsx"
    write_client_workbook(
        output,
        [
            ("Resumen", client_columns(summary)),
            ("Distribuidora", client_columns(by_distributor)),
            ("Cuenta", client_columns(by_account)),
            ("Mensual", client_columns(by_month)),
            ("Temas", client_columns(by_title)),
            ("Pais DSP", client_columns(by_country_dsp)),
            ("Detalle", client_columns(detail)),
        ],
    )
    return output


def main() -> None:
    parser = argparse.ArgumentParser(description="Reporte La Juntada de los Artistas listo para cliente.")
    parser.add_argument("--end-month", default=None)
    parser.add_argument("--copy-downloads", action="store_true")
    args = parser.parse_args()
    output = build_report(end_month=args.end_month)
    if args.copy_downloads:
        DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)
        shutil.copy2(output, DOWNLOADS_DIR / output.name)
    print(output)


if __name__ == "__main__":
    main()
