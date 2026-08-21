import argparse
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
import polars as pl

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from lib.catalog_report_filter import apply_report_net_personalization, filter_reportable_generation
    from lib.store_taxonomy import build_normalized_store_summary
except ModuleNotFoundError:
    from scripts.lib.catalog_report_filter import apply_report_net_personalization, filter_reportable_generation
    from scripts.lib.store_taxonomy import build_normalized_store_summary


BASE = Path(r"C:\royalties_pipeline")
MARTS = BASE / "warehouse" / "marts"
REPORTS = BASE / "reports"

SONG_PATH = MARTS / "song_level_all_sources.parquet"
STANDARDIZED_PATH = MARTS / "standardized_raw_all_sources.parquet"


SEARCH_COLUMNS_SONG = [
    "asset_isrc",
    "track_statement_style",
    "asset_title_statement",
    "artist_statement_style",
    "asset_artist_statement",
    "source",
    "account",
    "content_type",
    "source_sheet",
    "revenue_basis",
]

SEARCH_COLUMNS_STANDARDIZED = [
    "asset_isrc",
    "ISRC",
    "Track Title",
    "Title",
    "Video Title",
    "Asset Title",
    "Product Title",
    "track_statement_style",
    "asset_title_statement",
    "artist_statement_style",
    "asset_artist_statement",
    "Track Artists",
    "Artist Name",
    "Channel Name",
    "Product Artist",
    "TRACK ARTIST",
    "TRACK",
    "artists_raw",
    "Composers",
    "composers",
    "Payee",
    "payee",
    "source",
    "account",
    "store_name",
    "Store",
    "DSP",
    "territory",
    "Territory",
    "SALE COUNTRY",
    "Region",
    "Sales Region",
]


TERRITORY_CANDIDATES = [
    "territory",
    "Territory",
    "SALE COUNTRY",
    "Region",
    "Sales Region",
]


CODE_CANDIDATES = [
    ("asset_isrc", "ISRC"),
    ("ISRC", "ISRC"),
    ("Asset ISRC", "ISRC"),
    ("track_id", "Track ID"),
    ("Track ID", "Track ID"),
    ("Label Track ID", "Label Track ID"),
    ("label_track_id", "Label Track ID"),
    ("Asset Reference", "Asset Reference"),
    ("asset_reference", "Asset Reference"),
    ("FUGA Asset ID", "FUGA Asset ID"),
    ("Video ID", "Video ID"),
    ("VideoId", "Video ID"),
    ("video_id", "Video ID"),
    ("YOUTUBE VIDEO ID", "Video ID"),
    ("Product Reference", "Product Reference"),
    ("product_reference", "Product Reference"),
    ("DSP Unit ID", "DSP Unit ID"),
    ("DSP Container ID", "DSP Container ID"),
    ("ID", "ID"),
    ("Sale ID", "Sale ID"),
    ("sale_id", "Sale ID"),
]


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOTAL_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
TOTAL_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

DISPLAY_HEADERS = {
    "distributor": "Distribuidora",
    "keywords": "Filtro",
    "mode": "Coincidencia",
    "period_basis": "Criterio periodo",
    "start_month": "Desde",
    "end_month": "Hasta",
    "excluded_isrcs": "ISRC excluidos",
    "song_level_rows": "Filas resultado",
    "song_level_amount_usd": "Ingresos USD",
    "song_level_units": "Unidades",
    "raw_sample_rows": "Filas raw",
    "generated_at": "Generado el",
    "source": "Fuente",
    "account": "Cuenta",
    "transaction_month": "Mes",
    "period_month": "Mes",
    "amount_usd": "Ingresos USD",
    "units": "Unidades",
    "rows": "Filas",
    "report_territory": "Territorio",
    "report_code": "Codigo",
    "report_code_source": "Tipo codigo",
    "original_codes": "Codigos originales",
    "content_types": "Tipo de contenido",
    "asset_isrc": "ISRC",
    "track_statement_style": "Tema",
    "asset_title_statement": "Asset title",
    "artist_statement_style": "Artista",
    "asset_artist_statement": "Asset artist",
    "content_type": "Tipo de contenido",
    "store_raw": "Store original",
    "dsp_normalized": "DSP / Store",
    "monetization_normalized": "Monetizacion",
    "content_origin_normalized": "Origen",
    "plan_normalized": "Plan",
    "classification_status": "Clasificacion",
    "store_report_label": "Store normalizado",
    "usage_type": "Tipo",
    "first_month": "Desde",
    "last_month": "Hasta",
    "source_sheet": "Hoja origen",
    "revenue_basis": "Base ingreso",
    "match_text": "Texto coincidente",
    "statement_period": "Periodo statement",
    "net_amount": "Importe neto original",
    "currency_original": "Moneda original",
    "fx_to_usd_rate": "FX USD",
    "store_name": "Tienda",
    "territory": "Territorio",
    "statement_file_name": "Archivo statement",
    "hoja": "Hoja",
    "contenido": "Contenido",
    "resultado": "Resultado",
}

PERIOD_BASIS_LABELS = {
    "transaction_month": "Performance / mes de consumo",
    "statement_period": "Liquidacion / mes de statement",
}


def parse_args():
    parser = argparse.ArgumentParser(
        description="Genera un reporte dinamico de royalties por palabras clave."
    )
    parser.add_argument(
        "keywords",
        nargs="*",
        help="Palabras clave. Si no se pasan, el script las pide por consola.",
    )
    parser.add_argument(
        "--mode",
        choices=["any", "all"],
        default="any",
        help="any = matchea cualquier keyword; all = exige todas.",
    )
    parser.add_argument(
        "--raw-limit",
        type=int,
        default=5000,
        help="Maximo de filas en la hoja detalle.",
    )
    parser.add_argument(
        "--start-month",
        help="Mes inicial YYYY-MM. Filtra por transaction_month.",
    )
    parser.add_argument(
        "--end-month",
        help="Mes final YYYY-MM.",
    )
    parser.add_argument(
        "--period-basis",
        choices=["transaction_month", "statement_period"],
        default="transaction_month",
        help="transaction_month = performance; statement_period = liquidacion.",
    )
    parser.add_argument(
        "--exclude-isrc",
        action="append",
        default=[],
        help="ISRC a excluir del reporte. Se puede repetir o pasar separados por coma/semicolon.",
    )
    return parser.parse_args()


def normalize_keywords(raw_keywords: list[str]) -> list[str]:
    keywords = []

    for item in raw_keywords:
        parts = [part.strip() for part in re.split(r"[;,]", item) if part.strip()]
        keywords.extend(parts)

    return keywords


def prompt_keywords() -> list[str]:
    raw = input("Palabras clave a buscar, separadas por coma: ").strip()
    return normalize_keywords([raw])


def existing_columns(path: Path) -> set[str]:
    return set(pl.scan_parquet(path).collect_schema().names())


def normalize_isrcs(raw_isrcs: list[str] | None) -> list[str]:
    if not raw_isrcs:
        return []

    isrcs = []
    for item in raw_isrcs:
        parts = [part.strip().upper() for part in re.split(r"[;,\s]+", item) if part.strip()]
        isrcs.extend(parts)

    return sorted(set(isrcs))


def exclude_isrc_expr(columns: set[str], excluded_isrcs: list[str]) -> pl.Expr:
    if not excluded_isrcs:
        return pl.lit(True)

    exprs = []
    for col in ["asset_isrc", "ISRC"]:
        if col in columns:
            exprs.append(
                pl.col(col)
                .cast(pl.Utf8)
                .str.strip_chars()
                .str.to_uppercase()
                .is_in(excluded_isrcs)
                .fill_null(False)
            )

    if not exprs:
        return pl.lit(True)

    excluded = exprs[0]
    for expr in exprs[1:]:
        excluded = excluded | expr

    return ~excluded


def contains_expr(columns: set[str], search_columns: list[str], keyword: str) -> pl.Expr:
    exprs = []
    normalized_keyword = re.sub(r"\s+", " ", keyword.strip().lower())
    compact_keyword = re.sub(r"[\s_-]+", "", normalized_keyword)

    for col in search_columns:
        if col in columns:
            normalized_value = (
                pl.col(col)
                .cast(pl.Utf8)
                .str.to_lowercase()
                .str.replace_all(r"\s+", " ")
                .str.strip_chars()
            )
            match = normalized_value.str.contains(normalized_keyword, literal=True)
            if compact_keyword:
                match = match | (
                    normalized_value
                    .str.replace_all(r"[\s_-]+", "")
                    .str.contains(compact_keyword, literal=True)
                )
            exprs.append(match.fill_null(False))

    if not exprs:
        return pl.lit(False)

    result = exprs[0]
    for expr in exprs[1:]:
        result = result | expr

    return result


def build_filter(columns: set[str], search_columns: list[str], keywords: list[str], mode: str) -> pl.Expr:
    exprs = [contains_expr(columns, search_columns, keyword) for keyword in keywords]

    if not exprs:
        return pl.lit(True)

    result = exprs[0]

    for expr in exprs[1:]:
        if mode == "all":
            result = result & expr
        else:
            result = result | expr

    return result


def add_match_text(lf: pl.LazyFrame, columns: set[str], search_columns: list[str]) -> pl.LazyFrame:
    usable = [col for col in search_columns if col in columns]

    if not usable:
        return lf.with_columns(pl.lit("").alias("match_text"))

    return lf.with_columns(
        pl.concat_str(
            [pl.col(col).cast(pl.Utf8).fill_null("") for col in usable],
            separator=" | ",
        ).alias("match_text")
    )


def safe_select(df: pl.DataFrame, columns: list[str]) -> pl.DataFrame:
    existing = [col for col in columns if col in df.columns]
    return df.select(existing)


def non_empty_code_expr(col_name: str) -> pl.Expr:
    value = pl.col(col_name).cast(pl.Utf8, strict=False).str.strip_chars()
    return pl.when(value.is_not_null() & (value != "")).then(value).otherwise(None)


def report_code_expr(columns: set[str]) -> pl.Expr:
    candidates = [non_empty_code_expr(col) for col, _label in CODE_CANDIDATES if col in columns]

    if not candidates:
        return pl.lit(None).cast(pl.Utf8)

    return pl.coalesce(candidates)


def report_code_source_expr(columns: set[str]) -> pl.Expr:
    expr = pl.lit(None).cast(pl.Utf8)

    for col, label in reversed([(col, label) for col, label in CODE_CANDIDATES if col in columns]):
        value = pl.col(col).cast(pl.Utf8, strict=False).str.strip_chars()
        expr = pl.when(value.is_not_null() & (value != "")).then(pl.lit(label)).otherwise(expr)

    return expr


def add_report_code(lf: pl.LazyFrame, columns: set[str]) -> pl.LazyFrame:
    return lf.with_columns([
        report_code_expr(columns).alias("report_code"),
        report_code_source_expr(columns).alias("report_code_source"),
        report_territory_expr(columns).alias("report_territory"),
    ])


def _summary_text_expr(columns: set[str], names: list[str]) -> pl.Expr:
    candidates = []
    for name in names:
        if name not in columns:
            continue
        value = pl.col(name).cast(pl.Utf8, strict=False).str.strip_chars()
        candidates.append(pl.when(value.is_not_null() & (value != "")).then(value).otherwise(None))
    if not candidates:
        return pl.lit(None).cast(pl.Utf8)
    return pl.coalesce(candidates)


def build_track_summary_lf(base_lf: pl.LazyFrame) -> pl.LazyFrame:
    """Build one economic row per resolved catalog identity."""
    columns = set(base_lf.collect_schema().names())
    title = _summary_text_expr(columns, ["track_statement_style", "asset_title_statement"])
    raw_artist = _summary_text_expr(columns, ["artist_statement_style", "asset_artist_statement"])
    artist = (
        pl.when(
            raw_artist.fill_null("").str.to_lowercase().str.replace_all(r"\s+", " ").str.strip_chars()
            == title.fill_null("").str.to_lowercase().str.replace_all(r"\s+", " ").str.strip_chars()
        )
        .then(None)
        .otherwise(raw_artist)
    )
    catalog_key = _summary_text_expr(columns, ["catalog_key"])
    asset_isrc = _summary_text_expr(columns, ["asset_isrc"])
    report_code = _summary_text_expr(columns, ["report_code"])
    report_code_source = _summary_text_expr(columns, ["report_code_source"])

    normalized_title = (
        title.fill_null("").str.to_lowercase().str.replace_all(r"\s+", " ").str.strip_chars()
    )
    normalized_artist = (
        artist.fill_null("").str.to_lowercase().str.replace_all(r"\s+", " ").str.strip_chars()
    )
    fallback_identity = (
        pl.when(asset_isrc.is_not_null())
        .then(pl.concat_str([pl.lit("ISRC:"), asset_isrc.str.to_uppercase()]))
        .when(report_code.is_not_null())
        .then(pl.concat_str([
            pl.lit("CODE:"),
            report_code_source.fill_null("unknown"),
            pl.lit(":"),
            report_code,
        ]))
        .otherwise(pl.concat_str([
            pl.lit("TEXT:"),
            normalized_title,
            pl.lit("|"),
            normalized_artist,
        ]))
    )
    trusted_code = (
        pl.when(
            report_code.is_not_null()
            & ~report_code_source.fill_null("").is_in(["Sale ID", "ID"])
        )
        .then(report_code)
        .otherwise(None)
    )

    grouped = (
        base_lf
        .with_columns([
            pl.coalesce([catalog_key, fallback_identity]).alias("_report_identity"),
            title.alias("_summary_title"),
            artist.alias("_summary_artist"),
            trusted_code.alias("_trusted_original_code"),
            report_code_source.alias("_summary_code_source"),
            _summary_text_expr(columns, ["content_type"]).alias("_summary_content_type"),
        ])
        .group_by("_report_identity")
        .agg([
            pl.col("_summary_title")
            .sort_by(pl.col("amount_usd").abs(), descending=True)
            .drop_nulls()
            .first()
            .alias("track_statement_style"),
            pl.col("_summary_artist")
            .sort_by(pl.col("amount_usd").abs(), descending=True)
            .drop_nulls()
            .first()
            .alias("artist_statement_style"),
            pl.col("_trusted_original_code")
            .drop_nulls()
            .unique()
            .sort()
            .str.join(" | ")
            .alias("original_codes"),
            pl.col("_summary_content_type")
            .drop_nulls()
            .unique()
            .sort()
            .str.join(" | ")
            .alias("content_types"),
            pl.col("_trusted_original_code")
            .sort_by(pl.col("amount_usd").abs(), descending=True)
            .drop_nulls()
            .first()
            .alias("_best_original_code"),
            pl.col("_summary_code_source")
            .sort_by(pl.col("amount_usd").abs(), descending=True)
            .drop_nulls()
            .first()
            .alias("_best_code_source"),
            pl.sum("amount_usd").alias("amount_usd"),
            pl.sum("units").alias("units"),
            pl.min("period_month").alias("first_month"),
            pl.max("period_month").alias("last_month"),
        ])
        .with_columns([
            pl.when(pl.col("_report_identity").str.starts_with("ISRC:"))
            .then(pl.col("_report_identity").str.replace(r"^ISRC:", ""))
            .when(pl.col("_report_identity").str.starts_with("VIDEO:"))
            .then(pl.col("_report_identity").str.replace(r"^VIDEO:", ""))
            .when(pl.col("_report_identity").str.starts_with("UPC:"))
            .then(pl.col("_report_identity").str.replace(r"^UPC:", ""))
            .otherwise(pl.col("_best_original_code"))
            .alias("report_code"),
            pl.when(pl.col("_report_identity").str.starts_with("ISRC:"))
            .then(pl.lit("ISRC"))
            .when(pl.col("_report_identity").str.starts_with("VIDEO:"))
            .then(pl.lit("Video ID"))
            .when(pl.col("_report_identity").str.starts_with("UPC:"))
            .then(pl.lit("UPC"))
            .otherwise(pl.col("_best_code_source"))
            .alias("report_code_source"),
            pl.col("track_statement_style")
            .fill_null("")
            .str.to_lowercase()
            .alias("_sort_tema"),
            pl.col("track_statement_style")
            .fill_null("")
            .str.strip_chars()
            .eq("")
            .cast(pl.Int8)
            .alias("_sort_tema_empty"),
        ])
        .sort([
            "_sort_tema_empty",
            "_sort_tema",
            "track_statement_style",
            "artist_statement_style",
            "report_code",
        ])
        .select([
            "track_statement_style",
            "artist_statement_style",
            "report_code",
            "report_code_source",
            "original_codes",
            "content_types",
            "amount_usd",
            "units",
            "first_month",
            "last_month",
        ])
    )
    return grouped


def report_territory_expr(columns: set[str]) -> pl.Expr:
    candidates = []
    for col in TERRITORY_CANDIDATES:
        if col in columns:
            value = pl.col(col).cast(pl.Utf8, strict=False).str.strip_chars()
            candidates.append(pl.when(value.is_not_null() & (value != "")).then(value).otherwise(None))

    if not candidates:
        return pl.lit(None).cast(pl.Utf8)

    return pl.coalesce(candidates)


def add_report_territory(lf: pl.LazyFrame, columns: set[str]) -> pl.LazyFrame:
    return lf.with_columns(report_territory_expr(columns).alias("report_territory"))


def normalize_report_territory_expr() -> pl.Expr:
    value = pl.col("report_territory").cast(pl.Utf8, strict=False).str.strip_chars()
    return (
        pl.when(value.is_null() | (value == ""))
        .then(pl.lit("Sin territorio"))
        .otherwise(value)
    )


def report_units_expr(columns: set[str]) -> pl.Expr:
    candidates = []
    for col in [
        "units",
        "streams",
        "asset_quantity_num",
        "Asset Quantity",
        "product_quantity_num",
        "Product Quantity",
        "Quantity",
        "QUANTITY",
        "Units of Sold",
    ]:
        if col in columns:
            candidates.append(pl.col(col).cast(pl.Float64, strict=False))

    if not candidates:
        return pl.lit(None).cast(pl.Float64)

    return pl.coalesce(candidates)


def normalize_report_units(lf: pl.LazyFrame, columns: set[str]) -> pl.LazyFrame:
    return lf.with_columns(report_units_expr(columns).alias("units"))


def store_raw_expr(columns: set[str]) -> pl.Expr:
    candidates = []
    for col in [
        "store_name",
        "Store",
        "DSP",
        "Sale Store Name",
        "STORE",
        "service_detail",
        "SERVICE DETAIL",
        "Store Name",
    ]:
        if col in columns:
            candidates.append(pl.col(col).cast(pl.Utf8, strict=False).str.strip_chars())

    if not candidates:
        return pl.lit(None).cast(pl.Utf8)

    return pl.coalesce(candidates)


def normalize_store_expr(raw_store: pl.Expr) -> pl.Expr:
    value = raw_store.fill_null("").str.to_lowercase()

    return (
        pl.when(value.str.contains("spotify", literal=True))
        .then(pl.lit("spotify"))
        .when(value.str.contains("youtube", literal=True) | value.str.contains("yt ", literal=True))
        .then(
            pl.when(
                value.str.contains("art track", literal=True)
                | value.str.contains("youtube music", literal=True)
                | value.str.contains("yt audio", literal=True)
            )
            .then(pl.lit("youtube_music"))
            .otherwise(pl.lit("youtube_video"))
        )
        .when(value.str.contains("tiktok", literal=True))
        .then(pl.lit("tiktok"))
        .when(value.str.contains("apple", literal=True) | value.str.contains("itunes", literal=True))
        .then(pl.lit("apple_music"))
        .when(value.str.contains("amazon", literal=True))
        .then(pl.lit("amazon"))
        .when(value.str.contains("facebook", literal=True) | value.str.contains("instagram", literal=True))
        .then(pl.lit("meta"))
        .when(value.str.contains("deezer", literal=True))
        .then(pl.lit("deezer"))
        .when(value.str.contains("pandora", literal=True))
        .then(pl.lit("pandora"))
        .when(value.str.contains("tidal", literal=True))
        .then(pl.lit("tidal"))
        .when(value.str.contains("audiomack", literal=True))
        .then(pl.lit("audiomack"))
        .when(value.str.contains("qobuz", literal=True))
        .then(pl.lit("qobuz"))
        .when(value.str.contains("soundcloud", literal=True))
        .then(pl.lit("soundcloud"))
        .when(raw_store.is_null() | (raw_store == ""))
        .then(pl.lit("unknown"))
        .otherwise(value.str.replace_all(r"[^a-z0-9]+", "_").str.strip_chars("_"))
    )


def normalize_report_store(lf: pl.LazyFrame, columns: set[str]) -> pl.LazyFrame:
    raw_store = store_raw_expr(columns)
    normalized_store = (
        pl.col("store_report_label").cast(pl.Utf8, strict=False).str.strip_chars()
        if "store_report_label" in columns
        else normalize_store_expr(raw_store)
    )
    return lf.with_columns([
        raw_store.alias("store_raw"),
        normalized_store.alias("store"),
    ])


def usage_raw_expr(columns: set[str]) -> pl.Expr:
    candidates = []
    for col in [
        "Sale Type",
        "sale_type",
        "Sales Type",
        "Use Type",
        "use_type",
        "Product Type",
        "product_type",
        "ROYALTY TYPE",
        "Royalty Type",
        "royalty_type",
        "TRANSACTION TYPE",
        "TRANSACTION SUBTYPE",
        "SERVICE DETAIL",
        "service_detail",
        "source_sheet",
    ]:
        if col in columns:
            candidates.append(pl.col(col).cast(pl.Utf8, strict=False).str.strip_chars())

    if not candidates:
        return pl.lit(None).cast(pl.Utf8)

    return pl.coalesce(candidates)


def normalize_usage_expr(raw_usage: pl.Expr, raw_store: pl.Expr, source_sheet: pl.Expr) -> pl.Expr:
    value = pl.concat_str([
        raw_usage.fill_null(""),
        raw_store.fill_null(""),
        source_sheet.fill_null(""),
    ], separator=" | ").str.to_lowercase()
    store_value = raw_store.fill_null("").str.to_lowercase()
    usage_value = raw_usage.fill_null("").str.to_lowercase()
    sheet_value = source_sheet.fill_null("").str.to_lowercase()

    return (
        pl.when(value.str.contains("short", literal=True))
        .then(pl.lit("short_video"))
        .when(value.str.contains("download", literal=True))
        .then(pl.lit("download"))
        .when(
            store_value.str.contains("youtube premium", literal=True)
            & usage_value.str.contains("stream", literal=True)
            & sheet_value.is_in(["masters", "shares in & out"])
        )
        .then(pl.lit("youtube_premium_stream"))
        .when(
            store_value.str.contains("youtube", literal=True)
            & usage_value.str.contains("stream", literal=True)
            & sheet_value.is_in(["masters", "shares in & out"])
        )
        .then(pl.lit("youtube_master_stream"))
        .when(
            value.str.contains("ugc", literal=True)
            | value.str.contains("content id", literal=True)
            | value.str.contains("channel income", literal=True)
        )
        .then(pl.lit("video_ugc"))
        .when(
            value.str.contains("video", literal=True)
            | value.str.contains("youtube", literal=True)
            | value.str.contains("vevo", literal=True)
        )
        .then(pl.lit("video_stream"))
        .when(value.str.contains("stream", literal=True))
        .then(pl.lit("audio_stream"))
        .when(value.str.contains("tiktok", literal=True))
        .then(pl.lit("short_video"))
        .when(raw_usage.is_null() | (raw_usage == ""))
        .then(pl.lit("unknown"))
        .otherwise(value.str.replace_all(r"[^a-z0-9]+", "_").str.strip_chars("_"))
    )


def normalize_report_usage(lf: pl.LazyFrame, columns: set[str]) -> pl.LazyFrame:
    raw_usage = usage_raw_expr(columns)
    raw_store = store_raw_expr(columns)
    source_sheet = (
        pl.col("source_sheet").cast(pl.Utf8, strict=False).str.strip_chars()
        if "source_sheet" in columns
        else pl.lit(None).cast(pl.Utf8)
    )
    normalized_usage = (
        pl.col("content_origin_normalized").cast(pl.Utf8, strict=False).str.strip_chars()
        if "content_origin_normalized" in columns
        else normalize_usage_expr(raw_usage, raw_store, source_sheet)
    )
    return lf.with_columns([
        raw_usage.alias("usage_raw"),
        normalized_usage.alias("usage_type"),
    ])


def display_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    return dataframe.rename(columns={col: DISPLAY_HEADERS.get(col, col) for col in dataframe.columns})


def prepare_sheet(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    amount_headers = {
        "amount_usd",
        "song_level_amount_usd",
        "net_amount",
        "Ingresos USD",
        "Importe neto",
    }
    integer_headers = {
        "units",
        "rows",
        "song_level_rows",
        "raw_sample_rows",
        "Unidades",
        "Filas",
        "Filas song level",
        "Filas raw",
    }

    header_by_column = {
        cell.column: str(cell.value) if cell.value is not None else ""
        for cell in ws[1]
    }

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            header = header_by_column.get(cell.column, "")
            if header in amount_headers:
                cell.number_format = '$#,##0.00'
            elif header in integer_headers:
                cell.number_format = '#,##0'

    for column_cells in ws.columns:
        col_letter = get_column_letter(column_cells[0].column)
        max_len = 0

        for cell in column_cells[:300]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def add_period_filter(
    lf: pl.LazyFrame,
    columns: set[str],
    start_month: str | None,
    end_month: str | None,
    period_column: str,
) -> pl.LazyFrame:
    if period_column not in columns:
        return lf

    if start_month:
        lf = lf.filter(pl.col(period_column).cast(pl.Utf8) >= start_month)

    if end_month:
        lf = lf.filter(pl.col(period_column).cast(pl.Utf8) <= end_month)

    return lf


def add_source_account_filter(
    lf: pl.LazyFrame,
    columns: set[str],
    source: str | None,
    account: str | None,
) -> pl.LazyFrame:
    source_value = (source or "").strip().lower()
    account_value = (account or "").strip().lower()
    if source_value and "source" in columns:
        lf = lf.filter(
            pl.col("source").cast(pl.Utf8, strict=False).str.to_lowercase().str.strip_chars()
            == source_value
        )
    if account_value and "account" in columns:
        lf = lf.filter(
            pl.col("account").cast(pl.Utf8, strict=False).str.to_lowercase().str.strip_chars()
            == account_value
        )
    return lf


def style_workbook(writer):
    for ws in writer.book.worksheets:
        prepare_sheet(ws)


def report_output_path(
    keywords: list[str],
    start_month: str | None,
    end_month: str | None,
    output_dir: Path,
) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    slug = re.sub(r"[^A-Za-z0-9]+", "_", "_".join(keywords)).strip("_").lower()[:60]
    period_slug = ""
    if start_month or end_month:
        period_slug = f"_{start_month or 'start'}_to_{end_month or 'end'}"
    return output_dir / f"keyword_royalty_report_{slug}{period_slug}_{timestamp}.xlsx"


def build_report_tables(
    keywords: list[str],
    mode: str,
    raw_limit: int,
    start_month: str | None = None,
    end_month: str | None = None,
    period_basis: str = "transaction_month",
    song_path: Path = SONG_PATH,
    standardized_path: Path = STANDARDIZED_PATH,
    exclude_isrcs: list[str] | None = None,
    source: str | None = None,
    account: str | None = None,
) -> dict[str, pd.DataFrame]:
    period_column = "statement_period" if period_basis == "statement_period" else "transaction_month"
    period_label = PERIOD_BASIS_LABELS.get(period_column, period_column)
    excluded_isrcs = normalize_isrcs(exclude_isrcs)

    song_cols = existing_columns(song_path)
    song_filter = build_filter(song_cols, SEARCH_COLUMNS_SONG, keywords, mode)
    standardized_cols = existing_columns(standardized_path)

    if period_column == "statement_period":
        raw_filter = build_filter(standardized_cols, SEARCH_COLUMNS_STANDARDIZED, keywords, mode)

        statement_base_lf = (
            add_report_code(
                normalize_report_usage(
                    normalize_report_store(
                        normalize_report_units(
                            add_period_filter(
                                add_source_account_filter(
                                    pl.scan_parquet(standardized_path),
                                    standardized_cols,
                                    source,
                                    account,
                                ),
                                standardized_cols,
                                start_month,
                                end_month,
                                period_column,
                            ),
                            standardized_cols,
                        ),
                        standardized_cols,
                    ),
                    standardized_cols,
                ),
                standardized_cols,
            )
            .pipe(lambda frame: filter_reportable_generation(frame, standardized_cols))
            .pipe(lambda frame: apply_report_net_personalization(frame))
            .filter(raw_filter)
            .filter(exclude_isrc_expr(standardized_cols, excluded_isrcs))
            .select([
                col for col in [
                    "source",
                    "account",
                    "statement_period",
                    "transaction_month",
                    "report_code",
                    "report_code_source",
                    "asset_isrc",
                    "report_territory",
                    "track_statement_style",
                    "asset_title_statement",
                    "artist_statement_style",
                    "asset_artist_statement",
                    "content_type",
                    "dsp_normalized",
                    "monetization_normalized",
                    "content_origin_normalized",
                    "plan_normalized",
                    "classification_status",
                    "store_report_label",
                    "store_raw",
                    "usage_type",
                    "amount_usd",
                    "units",
                    "territory",
                    "source_sheet",
                    "revenue_basis",
                    "catalog_key",
                    "catalog_business_status",
                ]
                if col in standardized_cols or col in {"store", "store_raw", "usage_type", "usage_raw", "report_code", "report_code_source", "report_territory", "catalog_key", "catalog_business_status"}
            ])
            .with_columns([
                pl.lit(None).cast(pl.Utf8).alias("content_type")
                if "content_type" not in standardized_cols
                else pl.col("content_type"),
                pl.col(period_column).cast(pl.Utf8).alias("period_month"),
            ])
        )

        metrics_lf = statement_base_lf.select([
            pl.len().alias("song_level_rows"),
            pl.sum("amount_usd").alias("song_level_amount_usd"),
            pl.sum("units").alias("song_level_units"),
        ])

        monthly_summary_lf = (
            statement_base_lf
            .group_by(["period_month"])
            .agg([
                pl.sum("amount_usd").alias("amount_usd"),
                pl.sum("units").alias("units"),
                pl.len().alias("rows"),
            ])
            .sort("period_month")
        )

        store_summary_lf = build_normalized_store_summary(
            statement_base_lf,
            set(statement_base_lf.collect_schema().names()),
        ).rename({"source": "distributor"})

        territory_summary_lf = (
            statement_base_lf
            .with_columns(
                normalize_report_territory_expr().alias("report_territory")
            )
            .group_by(["report_territory"])
            .agg([
                pl.sum("amount_usd").alias("amount_usd"),
                pl.sum("units").alias("units"),
            ])
            .sort("report_territory")
        )

        track_summary_lf = build_track_summary_lf(statement_base_lf)

        raw_sample_lf = (
            add_report_code(
                normalize_report_usage(
                    normalize_report_store(
                        normalize_report_units(
                            add_period_filter(
                                add_match_text(
                                    add_source_account_filter(
                                        pl.scan_parquet(standardized_path),
                                        standardized_cols,
                                        source,
                                        account,
                                    ),
                                    standardized_cols,
                                    SEARCH_COLUMNS_STANDARDIZED,
                                ),
                                standardized_cols,
                                start_month,
                                end_month,
                                period_column,
                            ),
                            standardized_cols,
                        ),
                        standardized_cols,
                    ),
                    standardized_cols,
                ),
                standardized_cols,
            )
            .pipe(lambda frame: filter_reportable_generation(frame, standardized_cols))
            .pipe(lambda frame: apply_report_net_personalization(frame))
            .filter(raw_filter)
            .filter(exclude_isrc_expr(standardized_cols, excluded_isrcs))
            .select([
                col for col in [
                    "statement_period",
                    "transaction_month",
                    "artist_statement_style",
                    "track_statement_style",
                    "report_code",
                    "report_code_source",
                    "asset_isrc",
                    "report_territory",
                    "dsp_normalized",
                    "monetization_normalized",
                    "content_origin_normalized",
                    "plan_normalized",
                    "classification_status",
                    "store_report_label",
                    "store_raw",
                    "usage_type",
                    "amount_usd",
                    "net_amount",
                    "currency_original",
                    "fx_to_usd_rate",
                    "units",
                    "territory",
                    "statement_file_name",
                    "match_text",
                    "catalog_key",
                    "catalog_business_status",
                    "source",
                    "account",
                ]
                if col in standardized_cols or col in {"match_text", "store", "store_raw", "usage_type", "usage_raw", "report_code", "report_code_source", "report_territory", "catalog_key", "catalog_business_status"}
            ])
            .limit(raw_limit)
        )

        metrics = metrics_lf.collect()
        song_rows = int(metrics["song_level_rows"][0])
        song_amount_usd = float(metrics["song_level_amount_usd"][0] or 0)
        song_units = float(metrics["song_level_units"][0] or 0)

        if song_rows == 0:
            print("No hubo matches en standardized_raw_all_sources.")
            return {
                "resumen general": display_dataframe(pd.DataFrame([{
                    "keywords": ", ".join(keywords),
                    "mode": mode,
                    "period_basis": period_label,
                    "start_month": start_month or "",
                    "end_month": end_month or "",
                    "excluded_isrcs": ", ".join(excluded_isrcs),
                    "song_level_rows": 0,
                    "song_level_amount_usd": 0,
                    "song_level_units": 0,
                    "raw_sample_rows": 0,
                    "generated_at": datetime.now().isoformat(timespec="seconds"),
                }])),
                "sin resultados": display_dataframe(pd.DataFrame([{
                    "resultado": "Sin coincidencias para los parametros ingresados."
                }])),
            }

        monthly_summary = monthly_summary_lf.collect()
        store_summary = store_summary_lf.collect()
        territory_summary = territory_summary_lf.collect()
        track_summary = track_summary_lf.collect()
        raw_sample = raw_sample_lf.collect()

        tables = {
            "resumen general": display_dataframe(pd.DataFrame([{
                "keywords": ", ".join(keywords),
                "mode": mode,
                "period_basis": period_label,
                "start_month": start_month or "",
                "end_month": end_month or "",
                "excluded_isrcs": ", ".join(excluded_isrcs),
                "song_level_rows": song_rows,
                "song_level_amount_usd": song_amount_usd,
                "song_level_units": song_units,
                "raw_sample_rows": raw_sample.height if raw_sample.height > 0 else 0,
                "generated_at": datetime.now().isoformat(timespec="seconds"),
            }])),
            "resumen mensual": display_dataframe(monthly_summary.to_pandas()),
            "resumen por store": display_dataframe(store_summary.to_pandas()),
            "resumen por territorio": display_dataframe(territory_summary.to_pandas()),
            "resumen por tema": display_dataframe(track_summary.to_pandas()),
        }

        if raw_sample.height > 0:
            tables["detalle"] = display_dataframe(raw_sample.to_pandas())

        return tables
    else:
        song = (
            add_report_code(
                normalize_report_units(
                    add_period_filter(
                        add_match_text(
                            add_source_account_filter(
                                pl.scan_parquet(song_path),
                                song_cols,
                                source,
                                account,
                            ),
                            song_cols,
                            SEARCH_COLUMNS_SONG,
                        ),
                        song_cols,
                        start_month,
                        end_month,
                        period_column,
                    ),
                    song_cols,
                ),
                song_cols,
            )
            .pipe(lambda frame: filter_reportable_generation(frame, song_cols))
            .pipe(lambda frame: apply_report_net_personalization(frame))
            .filter(song_filter)
            .filter(exclude_isrc_expr(song_cols, excluded_isrcs))
            .with_columns(pl.col(period_column).cast(pl.Utf8).alias("period_month"))
            .collect()
        )

    if song.height == 0:
        print("No hubo matches en song_level_all_sources.")
        return {
            "resumen general": display_dataframe(pd.DataFrame([{
                "keywords": ", ".join(keywords),
                "mode": mode,
                "period_basis": period_label,
                "start_month": start_month or "",
                "end_month": end_month or "",
                "excluded_isrcs": ", ".join(excluded_isrcs),
                "song_level_rows": 0,
                "song_level_amount_usd": 0,
                "song_level_units": 0,
                "raw_sample_rows": 0,
                "generated_at": datetime.now().isoformat(timespec="seconds"),
            }])),
            "sin resultados": display_dataframe(pd.DataFrame([{
                "resultado": "Sin coincidencias para los parametros ingresados."
            }])),
        }

    monthly_summary = (
        song
        .group_by(["period_month"])
        .agg([
            pl.sum("amount_usd").alias("amount_usd"),
            pl.sum("units").alias("units"),
            pl.len().alias("rows"),
        ])
        .sort("period_month")
    )

    track_summary = build_track_summary_lf(song.lazy()).collect()

    raw_sample = pl.DataFrame()
    store_summary = pl.DataFrame()
    territory_summary = pl.DataFrame()
    if standardized_path.exists():
        raw_cols = standardized_cols
        raw_filter = build_filter(raw_cols, SEARCH_COLUMNS_STANDARDIZED, keywords, mode)

        raw_matches_lf = (
            add_report_code(
                normalize_report_usage(
                    normalize_report_store(
                        normalize_report_units(
                            add_period_filter(
                                add_match_text(
                                    add_source_account_filter(
                                        pl.scan_parquet(standardized_path),
                                        raw_cols,
                                        source,
                                        account,
                                    ),
                                    raw_cols,
                                    SEARCH_COLUMNS_STANDARDIZED,
                                ),
                                raw_cols,
                                start_month,
                                end_month,
                                period_column,
                            ),
                            raw_cols,
                        ),
                        raw_cols,
                    ),
                    raw_cols,
                ),
                raw_cols,
            )
            .pipe(lambda frame: filter_reportable_generation(frame, raw_cols))
            .pipe(lambda frame: apply_report_net_personalization(frame))
            .filter(raw_filter)
            .filter(exclude_isrc_expr(raw_cols, excluded_isrcs))
        )

        store_summary = (
            build_normalized_store_summary(
                raw_matches_lf,
                set(raw_matches_lf.collect_schema().names()),
            )
            .rename({"source": "distributor"})
            .collect()
        )

        territory_summary = (
            raw_matches_lf
            .with_columns(
                normalize_report_territory_expr().alias("report_territory")
            )
            .group_by(["report_territory"])
            .agg([
                pl.sum("amount_usd").alias("amount_usd"),
                pl.sum("units").alias("units"),
            ])
            .sort("report_territory")
            .collect()
        )

        raw_sample = (
            raw_matches_lf
            .select([
                col for col in [
                    "statement_period",
                    "transaction_month",
                    "artist_statement_style",
                    "track_statement_style",
                    "report_code",
                    "report_code_source",
                    "asset_isrc",
                    "report_territory",
                    "dsp_normalized",
                    "monetization_normalized",
                    "content_origin_normalized",
                    "plan_normalized",
                    "classification_status",
                    "store_report_label",
                    "store_raw",
                    "usage_type",
                    "amount_usd",
                    "net_amount",
                    "currency_original",
                    "fx_to_usd_rate",
                    "units",
                    "territory",
                    "statement_file_name",
                    "match_text",
                    "source",
                    "account",
                ]
                if col in raw_cols or col in {"match_text", "store", "store_raw", "usage_type", "usage_raw", "report_code", "report_code_source", "report_territory"}
            ])
            .limit(raw_limit)
            .collect()
        )

    tables = {
        "resumen general": display_dataframe(pd.DataFrame([{
            "keywords": ", ".join(keywords),
            "mode": mode,
            "period_basis": period_label,
            "start_month": start_month or "",
            "end_month": end_month or "",
            "excluded_isrcs": ", ".join(excluded_isrcs),
            "song_level_rows": song.height,
            "song_level_amount_usd": song["amount_usd"].sum(),
            "song_level_units": song["units"].sum(),
            "raw_sample_rows": raw_sample.height if raw_sample.height > 0 else 0,
            "generated_at": datetime.now().isoformat(timespec="seconds"),
        }])),
        "resumen mensual": display_dataframe(monthly_summary.to_pandas()),
        "resumen por store": display_dataframe(store_summary.to_pandas()),
        "resumen por territorio": display_dataframe(territory_summary.to_pandas()),
        "resumen por tema": display_dataframe(track_summary.to_pandas()),
    }

    if raw_sample.height > 0:
        tables["detalle"] = display_dataframe(raw_sample.to_pandas())

    return tables


def build_report(
    keywords: list[str],
    mode: str,
    raw_limit: int,
    start_month: str | None = None,
    end_month: str | None = None,
    period_basis: str = "transaction_month",
    song_path: Path = SONG_PATH,
    standardized_path: Path = STANDARDIZED_PATH,
    output_dir: Path = REPORTS,
    exclude_isrcs: list[str] | None = None,
) -> Path:
    output_path = report_output_path(keywords, start_month, end_month, output_dir)
    tables = build_report_tables(
        keywords=keywords,
        mode=mode,
        raw_limit=raw_limit,
        start_month=start_month,
        end_month=end_month,
        period_basis=period_basis,
        song_path=song_path,
        standardized_path=standardized_path,
        exclude_isrcs=exclude_isrcs,
    )

    output_dir.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, dataframe in tables.items():
            dataframe.to_excel(writer, index=False, sheet_name=sheet_name)

        style_workbook(writer)

    return output_path


def main():
    args = parse_args()
    keywords = normalize_keywords(args.keywords) if args.keywords else prompt_keywords()

    if not keywords:
        print("No ingresaste palabras clave.")
        raise SystemExit(1)

    print("Buscando:", ", ".join(keywords))
    print("Modo:", args.mode)

    output_path = build_report(
        keywords=keywords,
        mode=args.mode,
        raw_limit=args.raw_limit,
        start_month=args.start_month,
        end_month=args.end_month,
        period_basis=args.period_basis,
        exclude_isrcs=args.exclude_isrc,
    )

    print("\nReporte generado:")
    print(output_path)


if __name__ == "__main__":
    main()
