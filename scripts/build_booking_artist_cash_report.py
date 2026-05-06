from __future__ import annotations

import json
import sqlite3
from datetime import datetime
from pathlib import Path

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


BASE = Path(r"C:\royalties_pipeline")
HISTORICAL_CSV = BASE / "reports" / "booking" / "booking_shows_report_base.csv"
LIVE_DB = BASE / "warehouse" / "booking" / "live" / "booking_live.sqlite"
OUTPUT_DIR = BASE / "reports" / "booking"


MONEY_COLUMNS = {
    "cachet_show",
    "gastos",
    "gastos_directos_show",
    "neto_show",
    "se_lleva_artista",
    "se_lleva_indyana",
    "artist_share_amount",
    "producer_share_amount",
    "recupero_aplicado",
    "pago_artista_sugerido",
    "caja_indyana_sugerida",
    "artist_paid_amount",
    "producer_received_amount",
    "balance_artist_amount",
    "balance_producer_amount",
    "amount",
    "applied_amount",
    "artist_amount",
    "producer_amount",
    "saldo_recuperable",
}


def normalize(value: object) -> str:
    return str(value or "").strip().lower()


def autosize(ws, min_width: int = 10, max_width: int = 46) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min_width
        for cell in col[:500]:
            if cell.value is not None:
                width = max(width, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[letter].width = width


def style_header(ws, row: int, cols: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="D9E2F3"))
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")


def add_table(ws, name: str) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    ws.freeze_panes = "A2"


def write_sheet(wb: Workbook, name: str, rows: list[dict], table_name: str) -> None:
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False

    if not rows:
        ws.append(["Sin datos"])
        return

    headers = list(rows[0].keys())
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for row in rows:
        ws.append([row.get(header) for header in headers])

    for col_idx, header in enumerate(headers, start=1):
        if header in MONEY_COLUMNS or header.endswith("_ars"):
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = '$ #,##0'
        if header in {"porcentaje_artista", "porcentaje_productora", "artist_percent", "producer_percent"}:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "0%"
        if header in {"fecha", "show_date"}:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "yyyy-mm-dd"

    add_table(ws, table_name)
    autosize(ws)


def load_historical(keyword: str) -> pl.DataFrame:
    if not HISTORICAL_CSV.exists():
        return pl.DataFrame()

    df = pl.read_csv(HISTORICAL_CSV)
    return df.filter(pl.col("artista").cast(pl.Utf8).str.to_lowercase().str.contains(keyword))


def load_live(keyword: str) -> tuple[list[dict], list[dict], list[dict], list[dict]]:
    if not LIVE_DB.exists():
        return [], [], [], []

    conn = sqlite3.connect(LIVE_DB)
    conn.row_factory = sqlite3.Row

    show_rows = [dict(row) for row in conn.execute(
        """
        SELECT *
        FROM booking_shows
        WHERE lower(artist) LIKE ?
        ORDER BY show_date, id
        """,
        (f"%{keyword}%",),
    ).fetchall()]

    show_ids = [row["id"] for row in show_rows]
    if not show_ids:
        return [], [], [], []

    placeholders = ",".join("?" for _ in show_ids)
    expenses = [dict(row) for row in conn.execute(
        f"""
        SELECT *
        FROM booking_show_expenses
        WHERE show_id IN ({placeholders})
        ORDER BY show_id, id
        """,
        show_ids,
    ).fetchall()]
    adjustments = [dict(row) for row in conn.execute(
        f"""
        SELECT *
        FROM booking_artist_adjustments
        WHERE show_id IN ({placeholders})
        ORDER BY show_id, id
        """,
        show_ids,
    ).fetchall()]
    movements = [dict(row) for row in conn.execute(
        f"""
        SELECT *
        FROM booking_movements
        WHERE show_id IN ({placeholders})
        ORDER BY show_id, id
        """,
        show_ids,
    ).fetchall()]

    by_show_adjustments: dict[int, list[dict]] = {show_id: [] for show_id in show_ids}
    for adjustment in adjustments:
        by_show_adjustments.setdefault(adjustment["show_id"], []).append(adjustment)

    for show in show_rows:
        applied = sum(float(item["applied_amount"] or 0) for item in by_show_adjustments.get(show["id"], []))
        show["recupero_aplicado"] = applied
        show["pago_artista_sugerido"] = float(show["artist_share_amount"] or 0) - applied
        show["caja_indyana_sugerida"] = float(show["producer_share_amount"] or 0) + applied
        show["balance_artista_post_recupero"] = show["pago_artista_sugerido"] - float(show["artist_paid_amount"] or 0)
        show["balance_indyana_post_recupero"] = show["caja_indyana_sugerida"] - float(show["producer_received_amount"] or 0)
        show["receipt_refs"] = " | ".join(json.loads(show.pop("receipt_refs_json") or "[]"))

    for adjustment in adjustments:
        adjustment["saldo_recuperable"] = max(
            0,
            float(adjustment["artist_amount"] or 0) - float(adjustment["applied_amount"] or 0),
        )

    return show_rows, expenses, adjustments, movements


def build_summary_rows(historical: pl.DataFrame, live_shows: list[dict], live_adjustments: list[dict]) -> list[dict]:
    hist = {
        "shows": historical.height,
        "cachet": historical["cachet_show"].sum() if historical.height else 0,
        "gastos": historical["gastos"].sum() if historical.height else 0,
        "neto": historical["neto_show"].sum() if historical.height else 0,
        "artista": historical["se_lleva_artista"].sum() if historical.height else 0,
        "indyana": historical["se_lleva_indyana"].sum() if historical.height else 0,
    }
    live = {
        "shows": len(live_shows),
        "cachet": sum(float(row["cachet_amount"] or 0) for row in live_shows),
        "gastos": sum(float(row["expenses_amount"] or 0) for row in live_shows),
        "neto": sum(float(row["net_amount"] or 0) for row in live_shows),
        "artista": sum(float(row["artist_share_amount"] or 0) for row in live_shows),
        "indyana": sum(float(row["producer_share_amount"] or 0) for row in live_shows),
        "recupero": sum(float(row["recupero_aplicado"] or 0) for row in live_shows),
        "pago_artista": sum(float(row["artist_paid_amount"] or 0) for row in live_shows),
        "caja_indyana": sum(float(row["producer_received_amount"] or 0) for row in live_shows),
        "ajustes_saldo": sum(float(row["saldo_recuperable"] or 0) for row in live_adjustments),
    }
    return [
        {"Indicador": "Shows historicos", "Valor": hist["shows"]},
        {"Indicador": "Cachet historico", "Valor": hist["cachet"]},
        {"Indicador": "Gastos historicos", "Valor": hist["gastos"]},
        {"Indicador": "Neto historico", "Valor": hist["neto"]},
        {"Indicador": "Artista historico modelado", "Valor": hist["artista"]},
        {"Indicador": "Indyana historico modelado", "Valor": hist["indyana"]},
        {"Indicador": "Shows live", "Valor": live["shows"]},
        {"Indicador": "Cachet live", "Valor": live["cachet"]},
        {"Indicador": "Gastos directos live", "Valor": live["gastos"]},
        {"Indicador": "Neto live", "Valor": live["neto"]},
        {"Indicador": "Artista live generado", "Valor": live["artista"]},
        {"Indicador": "Indyana live generado", "Valor": live["indyana"]},
        {"Indicador": "Recupero aplicado live", "Valor": live["recupero"]},
        {"Indicador": "Caja artista pagada live", "Valor": live["pago_artista"]},
        {"Indicador": "Caja Indyana rendida live", "Valor": live["caja_indyana"]},
        {"Indicador": "Saldo recuperable ajustes live", "Valor": live["ajustes_saldo"]},
    ]


def main() -> None:
    keyword = "vir"
    generated_at = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = OUTPUT_DIR / f"booking_artist_cash_report_virsshy_{generated_at}.xlsx"

    historical = load_historical(keyword)
    live_shows, live_expenses, live_adjustments, live_movements = load_live(keyword)

    hist_rows = historical.select([
        "artista",
        "fecha",
        "venue_evento",
        "cachet_show",
        "gastos",
        "neto_show",
        "porcentaje_artista",
        "porcentaje_productora",
        "se_lleva_artista",
        "se_lleva_indyana",
        "control",
        "archivo_origen",
    ]).to_dicts() if historical.height else []

    caja_artista = [{
        "show_id": row["id"],
        "artista": row["artist"],
        "fecha": row["show_date"],
        "venue": row["venue"],
        "artist_share_amount": row["artist_share_amount"],
        "recupero_aplicado": row["recupero_aplicado"],
        "pago_artista_sugerido": row["pago_artista_sugerido"],
        "artist_paid_amount": row["artist_paid_amount"],
        "balance_artist_amount": row["balance_artista_post_recupero"],
        "status": row["status"],
    } for row in live_shows]

    caja_indyana = [{
        "show_id": row["id"],
        "artista": row["artist"],
        "fecha": row["show_date"],
        "venue": row["venue"],
        "producer_share_amount": row["producer_share_amount"],
        "recupero_aplicado": row["recupero_aplicado"],
        "caja_indyana_sugerida": row["caja_indyana_sugerida"],
        "producer_received_amount": row["producer_received_amount"],
        "balance_producer_amount": row["balance_indyana_post_recupero"],
        "status": row["status"],
    } for row in live_shows]

    live_rows = [{
        "show_id": row["id"],
        "artista": row["artist"],
        "fecha": row["show_date"],
        "venue": row["venue"],
        "tour_manager": row["tour_manager"],
        "cachet_show": row["cachet_amount"],
        "gastos_directos_show": row["expenses_amount"],
        "neto_show": row["net_amount"],
        "porcentaje_artista": row["artist_percent"] / 100,
        "porcentaje_productora": row["producer_percent"] / 100,
        "se_lleva_artista": row["artist_share_amount"],
        "se_lleva_indyana": row["producer_share_amount"],
        "recupero_aplicado": row["recupero_aplicado"],
        "pago_artista_sugerido": row["pago_artista_sugerido"],
        "caja_indyana_sugerida": row["caja_indyana_sugerida"],
        "artist_paid_amount": row["artist_paid_amount"],
        "producer_received_amount": row["producer_received_amount"],
        "status": row["status"],
        "notes": row["notes"],
    } for row in live_shows]

    live_movements_rows = [{
        "show_id": row["show_id"],
        "movement_type": row["movement_type"],
        "category": row["category"],
        "amount": row["amount"],
        "currency": row["currency"],
        "fx_rate": row["fx_rate"],
        "notes": row["notes"],
        "created_at": row["created_at"],
    } for row in live_movements]

    wb = Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet("Resumen")
    summary.sheet_view.showGridLines = False
    summary["A1"] = "Reporte booking Virsshy"
    summary["A1"].font = Font(size=18, bold=True, color="1F4E78")
    summary["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    summary["A3"] = "Historico = planillas anteriores modeladas. Live = shows cargados en el sistema nuevo."
    summary["A4"] = "Gastos directos del show bajan el neto. Pagos al artista y rendiciones a Indyana son movimientos de caja, no gastos."
    summary.append([])
    summary.append(["Indicador", "Valor"])
    style_header(summary, 6, 2)
    for row in build_summary_rows(historical, live_shows, live_adjustments):
        summary.append([row["Indicador"], row["Valor"]])
    for row_idx in range(6, summary.max_row + 1):
        if row_idx != 6 and "Shows" not in str(summary.cell(row=row_idx, column=1).value):
            summary.cell(row=row_idx, column=2).number_format = '$ #,##0'
    autosize(summary)

    write_sheet(wb, "Shows historicos", hist_rows, "ShowsHistoricosVirsshy")
    write_sheet(wb, "Shows live", live_rows, "ShowsLiveVirsshy")
    write_sheet(wb, "Caja artista", caja_artista, "CajaArtistaVirsshy")
    write_sheet(wb, "Caja Indyana", caja_indyana, "CajaIndyanaVirsshy")
    write_sheet(wb, "Gastos directos live", live_expenses, "GastosDirectosLiveVirsshy")
    write_sheet(wb, "Movimientos caja live", live_movements_rows, "MovimientosCajaLiveVirsshy")
    write_sheet(wb, "Ajustes live", live_adjustments, "AjustesLiveVirsshy")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(output)
    print("historical_rows", len(hist_rows))
    print("live_rows", len(live_rows))


if __name__ == "__main__":
    main()
