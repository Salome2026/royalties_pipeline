from __future__ import annotations

import re
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
import polars as pl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from lib.catalog_report_filter import apply_report_net_personalization, filter_reportable_generation
except ModuleNotFoundError:
    from scripts.lib.catalog_report_filter import apply_report_net_personalization, filter_reportable_generation


BASE = Path(r"C:\royalties_pipeline")
MARTS = BASE / "warehouse" / "marts"
REPORTS = BASE / "reports"

DEFAULT_RAW_PATH = MARTS / "standardized_raw_all_sources.parquet"
DEFAULT_SONG_PATH = MARTS / "song_level_all_sources.parquet"


SEARCH_COLUMNS_RAW = [
    "track_statement_style",
    "asset_title_statement",
    "artist_statement_style",
    "asset_artist_statement",
    "Product Title",
    "Asset Title",
    "Track Title",
    "Title",
    "PRODUCT",
    "TRACK",
    "Product Artist",
    "Asset Artist",
    "PRODUCT ARTIST",
    "TRACK ARTIST",
    "ISRC",
    "asset_isrc",
]

SEARCH_COLUMNS_SONG = [
    "asset_title_statement",
    "track_statement_style",
    "asset_artist_statement",
    "artist_statement_style",
    "asset_isrc",
]

TITLE_FALLBACK_COLUMNS = [
    "asset_title_statement",
    "Asset Title",
    "TRACK",
    "Track Title",
    "Title",
    "Product Title",
    "PRODUCT",
    "track_statement_style",
]

ARTIST_FALLBACK_COLUMNS = [
    "asset_artist_statement",
    "Asset Artist",
    "TRACK ARTIST",
    "Track Artist",
    "PRODUCT ARTIST",
    "Product Artist",
    "artist_statement_style",
]


DSP_FALLBACK_COLUMNS = [
    "dsp_normalized",
    "store_report_label",
    "dsp",
    "store",
    "store_raw",
    "store_name",
    "STORE",
    "Store",
    "Store Name",
    "Sale Store Name",
    "DSP",
]

USAGE_FALLBACK_COLUMNS = [
    "content_origin_normalized",
    "usage_type",
    "usage_raw",
    "TRANSACTION TYPE",
    "TRANSACTION SUBTYPE",
    "Sale Type",
    "Sale User Type",
    "Use Type",
    "Sales Type",
    "Sales Sub Type",
    "SERVICE DETAIL",
    "service_detail",
    "Royalty Type",
    "ROYALTY TYPE",
]

TERRITORY_FALLBACK_COLUMNS = [
    "pais",
    "report_territory",
    "territory",
    "Territory",
    "SALE COUNTRY",
    "Sales Region",
    "Sales Country",
]


DEFAULT_LOS_ANORMALES_TERMS = [
    "BAILE INOLVIDABLE RKT",
    "LLAM\u00c1NDOME RKT",
    "MOVIMIENTO RKT",
    "REGGAETON RKT",
    "COQUETA",
    "PAPASITO",
    "la fiesta empez\u00f3",
    "TU TA DEMASIADO LOCA RKT",
    "PIDE LO QUE T\u00da QUIERAS X TIKITIKI RKT",
    "PASO SOLITA X RELACI\u00d3N RKT",
    "LA LLEVO PA EL ESPACIO RKT",
    "INTRO TU JARD\u00cdN CON ENANITOS RKT",
    "INTRO AY VAMOS RKT",
    "BESOS MOJADOS RKT",
    "PIERDO LA CABEZA RKT",
    "NO PARE RKT",
    "M\u00c9TELO S\u00c1CALO RKT",
]


MANUAL_VARIANTS = {
    "LLAM\u00c1NDOME RKT": ["llamandome rkt", "llamandome"],
    "la fiesta empez\u00f3": ["la fiesta empezo", "la fiesta empez\u00f3"],
    "PIDE LO QUE T\u00da QUIERAS X TIKITIKI RKT": [
        "pide lo que tu quieras x tikitiki rkt",
        "pide lo que t\u00fa quieras x tikitiki rkt",
        "pide lo que tu quieras x tiki tiki rkt",
        "pide lo que t\u00fa quieras x tiki tiki rkt",
        "pide lo que quieras x tikitiki rkt",
        "pide lo que quieras x tiki tiki rkt",
        "pide lo que tu quieras",
        "pide lo que t\u00fa quieras",
        "pide lo que quieras",
    ],
    "PASO SOLITA X RELACI\u00d3N RKT": [
        "paso solita x relacion rkt",
        "paso solita x relaci\u00f3n rkt",
        "paso solita",
    ],
    "LA LLEVO PA EL ESPACIO RKT": [
        "la llevo pa el espacio rkt",
        "la llevo pa el espacio",
        "la llevo pal espacio rkt",
        "la llevo pal espacio",
    ],
    "INTRO TU JARD\u00cdN CON ENANITOS RKT": [
        "intro tu jardin con enanitos rkt",
        "intro tu jard\u00edn con enanitos rkt",
        "tu jardin con enanitos",
        "tu jard\u00edn con enanitos",
    ],
    "M\u00c9TELO S\u00c1CALO RKT": [
        "metelo sacalo rkt",
        "m\u00e9telo s\u00e1calo rkt",
        "metelo sacalo",
        "m\u00e9telo s\u00e1calo",
    ],
}


def strip_accents(value: str) -> str:
    return "".join(
        char for char in unicodedata.normalize("NFKD", value) if not unicodedata.combining(char)
    )


def normalize_text(value: str) -> str:
    value = strip_accents(str(value or "").lower())
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return " ".join(value.split())


def slugify(value: str) -> str:
    slug = normalize_text(value).replace(" ", "_")
    return slug[:80] or "reporte_personalizado"


def parse_search_line(line: str) -> dict[str, str]:
    raw = str(line or "").strip()
    if "|" in raw:
        title, artist = raw.split("|", 1)
    else:
        title, artist = raw, ""
    title = title.strip()
    artist = artist.strip()
    return {
        "raw": raw,
        "title": title,
        "artist": artist,
        "display": f"{title} | {artist}" if artist else title,
    }


def normalize_search_specs(raw_terms: list[str]) -> list[dict[str, str]]:
    specs: list[dict[str, str]] = []
    seen: set[str] = set()
    for item in raw_terms:
        spec = parse_search_line(item)
        key = f"{normalize_text(spec['title'])}|{normalize_text(spec['artist'])}"
        if spec["title"] and key not in seen:
            specs.append(spec)
            seen.add(key)
    return specs


def variants_for_text(value: str) -> list[str]:
    variants = [value, strip_accents(value)]
    variants.extend(MANUAL_VARIANTS.get(value, []))
    normalized_value = normalize_text(value)
    for key, manual_variants in MANUAL_VARIANTS.items():
        if normalize_text(key) == normalized_value and key != value:
            variants.extend(manual_variants)
    if normalized_value == "pide lo que t quieras x tikitiki rkt":
        variants.extend(MANUAL_VARIANTS["PIDE LO QUE T\u00da QUIERAS X TIKITIKI RKT"])
    normalized: list[str] = []
    for item in variants:
        clean = normalize_text(item)
        if clean and clean not in normalized:
            normalized.append(clean)
    return normalized


def normalized_column(col: str) -> pl.Expr:
    return (
        pl.col(col)
        .cast(pl.Utf8, strict=False)
        .str.to_lowercase()
        .str.replace_all("[\u00e1\u00e0\u00e4\u00e2\u00e3]", "a")
        .str.replace_all("[\u00e9\u00e8\u00eb\u00ea]", "e")
        .str.replace_all("[\u00ed\u00ec\u00ef\u00ee]", "i")
        .str.replace_all("[\u00f3\u00f2\u00f6\u00f4\u00f5]", "o")
        .str.replace_all("[\u00fa\u00f9\u00fc\u00fb]", "u")
        .str.replace_all("\u00f1", "n")
        .str.replace_all("[^a-z0-9]+", " ")
        .str.strip_chars()
    )


def text_match_expr(columns: set[str], search_columns: list[str], value: str) -> pl.Expr:
    usable_cols = [col for col in search_columns if col in columns]
    variants = variants_for_text(value)
    if not usable_cols or not variants:
        return pl.lit(False)

    exprs: list[pl.Expr] = []
    for col in usable_cols:
        norm = normalized_column(col)
        for variant in variants:
            exprs.append(norm.str.contains(variant, literal=True).fill_null(False))

    result = exprs[0]
    for expr in exprs[1:]:
        result = result | expr
    return result


def spec_expr(columns: set[str], title_columns: list[str], spec: dict[str, str]) -> pl.Expr:
    title_match = text_match_expr(columns, title_columns, spec["title"])
    if not spec["artist"]:
        return title_match
    artist_match = text_match_expr(columns, ARTIST_FALLBACK_COLUMNS, spec["artist"])
    return title_match & artist_match


def any_spec_expr(columns: set[str], search_columns: list[str], specs: list[dict[str, str]]) -> pl.Expr:
    result = spec_expr(columns, search_columns, specs[0])
    for spec in specs[1:]:
        result = result | spec_expr(columns, search_columns, spec)
    return result


def matched_spec_expr(columns: set[str], search_columns: list[str], specs: list[dict[str, str]]) -> list[pl.Expr]:
    title_expr = pl.lit("SIN_MATCH")
    artist_expr = pl.lit("")
    display_expr = pl.lit("SIN_MATCH")
    for spec in reversed(specs):
        match = spec_expr(columns, search_columns, spec)
        title_expr = pl.when(match).then(pl.lit(spec["title"])).otherwise(title_expr)
        artist_expr = pl.when(match).then(pl.lit(spec["artist"])).otherwise(artist_expr)
        display_expr = pl.when(match).then(pl.lit(spec["display"])).otherwise(display_expr)
    return [
        title_expr.alias("titulo_buscado"),
        artist_expr.alias("artista_buscado"),
        display_expr.alias("busqueda_buscada"),
    ]


def first_non_blank_expr(columns: set[str], candidates: list[str], alias: str) -> pl.Expr:
    exprs = []
    for col in candidates:
        if col in columns:
            value = pl.col(col).cast(pl.Utf8, strict=False).str.strip_chars()
            exprs.append(pl.when(value != "").then(value).otherwise(None))
    if not exprs:
        return pl.lit(None, dtype=pl.Utf8).alias(alias)
    return pl.coalesce(exprs).alias(alias)


def coalesce_text_expr(columns: set[str], candidates: list[str], alias: str) -> pl.Expr:
    exprs = []
    for col in candidates:
        if col in columns:
            value = pl.col(col).cast(pl.Utf8, strict=False).str.strip_chars()
            exprs.append(pl.when(value != "").then(value).otherwise(None))
    if not exprs:
        return pl.lit(None, dtype=pl.Utf8).alias(alias)
    return pl.coalesce(exprs).alias(alias)


def enrich_report_dimensions(df: pl.DataFrame) -> pl.DataFrame:
    columns = set(df.columns)
    return df.with_columns(
        [
            coalesce_text_expr(columns, TERRITORY_FALLBACK_COLUMNS, "pais"),
            coalesce_text_expr(columns, DSP_FALLBACK_COLUMNS, "dsp"),
            coalesce_text_expr(columns, USAGE_FALLBACK_COLUMNS, "usage_type"),
        ]
    )


def apply_statement_filter(
    lf: pl.LazyFrame,
    columns: set[str],
    start_month: str | None,
    end_month: str | None,
) -> pl.LazyFrame:
    period_col = "statement_period" if "statement_period" in columns else "transaction_month"
    period = pl.col(period_col).cast(pl.Utf8, strict=False)
    filters = [period.is_not_null()]
    if start_month:
        filters.append(period >= start_month)
    if end_month:
        filters.append(period <= end_month)
    result = filters[0]
    for expr in filters[1:]:
        result = result & expr
    return lf.filter(result)


def normalize_source_accounts(items: list[dict[str, Any]] | None) -> list[tuple[str, str]]:
    pairs: list[tuple[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for item in items or []:
        source = normalize_text(item.get("source", ""))
        account = normalize_text(item.get("account", ""))
        pair = (source, account)
        if source and account and pair not in seen:
            pairs.append(pair)
            seen.add(pair)
    return pairs


def apply_source_filter(
    lf: pl.LazyFrame,
    columns: set[str],
    sources: list[str] | None,
    source_accounts: list[dict[str, Any]] | None,
) -> pl.LazyFrame:
    pairs = normalize_source_accounts(source_accounts)
    if pairs and {"source", "account"}.issubset(columns):
        pair_values = [f"{source}||{account}" for source, account in pairs]
        pair_expr = pl.concat_str(
            [
                normalized_column("source"),
                pl.lit("||"),
                normalized_column("account"),
            ]
        )
        return lf.filter(pair_expr.is_in(pair_values))

    selected = [normalize_text(source) for source in sources or [] if normalize_text(source)]
    if not selected or "source" not in columns:
        return lf
    return lf.filter(normalized_column("source").is_in(selected))


def first_existing(columns: set[str], candidates: list[str]) -> str | None:
    return next((col for col in candidates if col in columns), None)


def group_sum(df: pl.DataFrame, by: list[str]) -> pd.DataFrame:
    if df.height == 0 or not by:
        return pd.DataFrame(columns=by + ["amount_usd", "units", "rows"])
    aggs = [pl.sum("amount_usd").alias("amount_usd"), pl.len().alias("rows")]
    if "units" in df.columns:
        aggs.insert(1, pl.sum("units").alias("units"))
    return df.group_by(by).agg(aggs).sort("amount_usd", descending=True).to_pandas()


def build_dataset(
    path: Path,
    search_columns: list[str],
    specs: list[dict[str, str]],
    start_month: str | None,
    end_month: str | None,
    sources: list[str] | None,
    source_accounts: list[dict[str, Any]] | None,
) -> pl.DataFrame:
    columns = set(pl.scan_parquet(path).collect_schema().names())
    lf = pl.scan_parquet(path)
    lf = apply_source_filter(
        apply_statement_filter(lf, columns, start_month, end_month),
        columns,
        sources,
        source_accounts,
    )
    return (
        lf.with_columns(matched_spec_expr(columns, search_columns, specs))
        .filter(any_spec_expr(columns, search_columns, specs))
        .with_columns(
            [
                first_non_blank_expr(columns, TITLE_FALLBACK_COLUMNS, "asset_title_statement"),
                first_non_blank_expr(columns, ARTIST_FALLBACK_COLUMNS, "asset_artist_statement"),
            ]
        )
        .collect()
    )


def select_columns(df: pl.DataFrame, wanted: list[str]) -> pl.DataFrame:
    return df.select([col for col in wanted if col in df.columns])


def title_summary(raw_df: pl.DataFrame) -> pd.DataFrame:
    columns = [
        "busqueda_buscada",
        "titulo_buscado",
        "artista_buscado",
        "asset_isrc",
        "asset_title_statement",
        "asset_artist_statement",
        "amount_usd",
        "units",
        "rows",
    ]
    if raw_df.height == 0:
        return pd.DataFrame(columns=columns)

    df = raw_df.with_columns(
        [
            (
                pl.col("asset_isrc").cast(pl.Utf8, strict=False).fill_null("")
                if "asset_isrc" in raw_df.columns
                else pl.lit("")
            ).alias("asset_isrc"),
            pl.col("asset_title_statement").cast(pl.Utf8, strict=False).fill_null("").alias("asset_title_statement"),
            pl.col("asset_artist_statement").cast(pl.Utf8, strict=False).fill_null("").alias("asset_artist_statement"),
        ]
    )
    return (
        df.group_by(["busqueda_buscada", "titulo_buscado", "artista_buscado", "asset_isrc"])
        .agg(
            [
                pl.col("asset_title_statement")
                .filter(pl.col("asset_title_statement") != "")
                .mode()
                .first()
                .alias("asset_title_statement"),
                pl.col("asset_artist_statement")
                .filter(pl.col("asset_artist_statement") != "")
                .mode()
                .first()
                .alias("asset_artist_statement"),
                pl.sum("amount_usd").alias("amount_usd"),
                pl.sum("units").alias("units") if "units" in df.columns else pl.lit(0).alias("units"),
                pl.len().alias("rows"),
            ]
        )
        .sort("amount_usd", descending=True)
        .to_pandas()
    )


def song_matches_from_raw(raw_df: pl.DataFrame) -> pd.DataFrame:
    columns = [
        "busqueda_buscada",
        "titulo_buscado",
        "artista_buscado",
        "source",
        "account",
        "asset_isrc",
        "asset_title_statement",
        "asset_artist_statement",
        "content_type",
        "statement_period",
        "transaction_month",
        "pais",
        "dsp",
        "usage_type",
        "amount_usd",
        "units",
        "rows",
    ]
    if raw_df.height == 0:
        return pd.DataFrame(columns=columns)

    raw_df = enrich_report_dimensions(raw_df)
    raw_cols = set(raw_df.columns)
    group_cols = [
        "busqueda_buscada",
        "titulo_buscado",
        "artista_buscado",
        "source",
        "account",
        "asset_isrc",
        "asset_title_statement",
        "asset_artist_statement",
        "content_type",
        "statement_period",
        "transaction_month",
        "pais",
        "dsp",
        "usage_type",
    ]

    df = raw_df.with_columns(
        [
            (
                pl.col("asset_isrc").cast(pl.Utf8, strict=False).fill_null("")
                if "asset_isrc" in raw_cols
                else pl.lit("")
            ).alias("asset_isrc"),
            (
                pl.col("content_type").cast(pl.Utf8, strict=False).str.strip_chars()
                if "content_type" in raw_cols
                else pl.lit(None, dtype=pl.Utf8)
            ).alias("content_type"),
        ]
    )
    return (
        df.group_by(group_cols)
        .agg(
            [
                pl.sum("amount_usd").alias("amount_usd"),
                pl.sum("units").alias("units") if "units" in df.columns else pl.lit(0).alias("units"),
                pl.len().alias("rows"),
            ]
        )
        .sort(["amount_usd", "rows"], descending=[True, True])
        .to_pandas()
    )


def style_workbook(wb) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                header = str(ws.cell(row=1, column=cell.column).value or "").lower()
                if "usd" in header or "ingresos" in header or header == "valor":
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '$#,##0.00'
                elif any(token in header for token in ["units", "unidades", "rows", "filas"]):
                    cell.number_format = '#,##0'
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = 0
            for cell in col[:500]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 46))


def source_account_label(source_accounts: list[dict[str, Any]] | None, sources: list[str] | None) -> str:
    pairs = normalize_source_accounts(source_accounts)
    if pairs:
        return ", ".join(f"{source}/{account}" for source, account in pairs)
    return ", ".join(sources or ["todas"])


def build_custom_title_report(
    report_title: str,
    terms: list[str],
    start_month: str | None = None,
    end_month: str | None = None,
    sources: list[str] | None = None,
    source_accounts: list[dict[str, Any]] | None = None,
    raw_path: Path = DEFAULT_RAW_PATH,
    song_path: Path = DEFAULT_SONG_PATH,
    output_dir: Path = REPORTS,
) -> Path:
    specs = normalize_search_specs(terms)
    if not specs:
        raise ValueError("La lista de busqueda no puede estar vacia.")

    raw_df = build_dataset(raw_path, SEARCH_COLUMNS_RAW, specs, start_month, end_month, sources, source_accounts)
    if raw_df.height:
        raw_df = (
            filter_reportable_generation(raw_df.lazy(), set(raw_df.columns))
            .pipe(lambda frame: apply_report_net_personalization(frame))
            .collect()
        )

    raw_df = select_columns(
        raw_df,
        [
            "busqueda_buscada",
            "titulo_buscado",
            "artista_buscado",
            "source",
            "account",
            "statement_period",
            "transaction_month",
            "asset_isrc",
            "ISRC",
            "track_statement_style",
            "asset_title_statement",
            "artist_statement_style",
            "asset_artist_statement",
            "content_type",
            "dsp_normalized",
            "monetization_normalized",
            "content_origin_normalized",
            "plan_normalized",
            "classification_status",
            "store_report_label",
            "store",
            "store_raw",
            "store_name",
            "STORE",
            "Store",
            "Store Name",
            "Sale Store Name",
            "DSP",
            "usage_type",
            "usage_raw",
            "TRANSACTION TYPE",
            "TRANSACTION SUBTYPE",
            "Sale Type",
            "Sale User Type",
            "Use Type",
            "Sales Type",
            "Sales Sub Type",
            "SERVICE DETAIL",
            "service_detail",
            "Royalty Type",
            "ROYALTY TYPE",
            "report_territory",
            "territory",
            "Territory",
            "SALE COUNTRY",
            "Sales Region",
            "amount_usd",
            "units",
            "source_file",
            "statement_file_name",
            "source_sheet",
            "catalog_key",
            "catalog_business_status",
            "catalog_status_notes",
        ],
    )
    raw_df = enrich_report_dimensions(raw_df)
    song_matches_table = song_matches_from_raw(raw_df)

    total_usd = float(raw_df["amount_usd"].sum()) if raw_df.height else 0.0
    total_units = float(raw_df["units"].sum()) if raw_df.height and "units" in raw_df.columns else 0.0

    period_label = f"{start_month or 'inicio'} a statement {end_month or 'ultimo disponible'}"
    overview = pd.DataFrame(
        [
            {"Indicador": "Reporte", "Valor": report_title},
            {"Indicador": "Generado", "Valor": datetime.now().strftime("%Y-%m-%d %H:%M")},
            {"Indicador": "Periodo", "Valor": period_label},
            {"Indicador": "Distribuidoras / cuentas", "Valor": source_account_label(source_accounts, sources)},
            {"Indicador": "Busquedas pedidas", "Valor": len(specs)},
            {
                "Indicador": "Busquedas con match raw",
                "Valor": raw_df.select("busqueda_buscada").unique().height if raw_df.height else 0,
            },
            {"Indicador": "Total ingresos USD", "Valor": total_usd},
            {"Indicador": "Total unidades", "Valor": total_units},
            {"Indicador": "Filas raw", "Valor": raw_df.height},
            {"Indicador": "Filas song_matches", "Valor": len(song_matches_table)},
        ]
    )

    control_rows = []
    for spec in specs:
        raw_term = raw_df.filter(pl.col("busqueda_buscada") == spec["display"])
        song_rows = int((song_matches_table["busqueda_buscada"] == spec["display"]).sum()) if not song_matches_table.empty else 0
        control_rows.append(
            {
                "busqueda_pedida": spec["display"],
                "titulo_pedido": spec["title"],
                "artista_pedido": spec["artist"],
                "variantes_titulo": ", ".join(variants_for_text(spec["title"])),
                "variantes_artista": ", ".join(variants_for_text(spec["artist"])) if spec["artist"] else "",
                "raw_rows": raw_term.height,
                "song_rows": song_rows,
                "raw_usd": float(raw_term["amount_usd"].sum()) if raw_term.height else 0.0,
                "raw_units": float(raw_term["units"].sum()) if raw_term.height and "units" in raw_term.columns else 0.0,
            }
        )
    match_control = pd.DataFrame(control_rows)

    tables = {
        "overview": overview,
        "match_control": match_control,
        "source_summary": group_sum(raw_df, ["source", "account"]),
        "title_summary": title_summary(raw_df),
        "monthly_statement": group_sum(raw_df, ["statement_period"]),
        "store_summary": group_sum(raw_df, ["dsp"]),
        "usage_summary": group_sum(raw_df, ["usage_type"]),
        "territory_summary": group_sum(raw_df, ["pais"]),
        "song_matches": song_matches_table,
    }

    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = output_dir / f"{slugify(report_title)}_{start_month or 'inicio'}_a_{end_month or 'ultimo'}_{timestamp}.xlsx"

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, table in tables.items():
            table.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        style_workbook(writer.book)

    return output
