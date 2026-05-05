from __future__ import annotations

from datetime import datetime
from pathlib import Path

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


BASE = Path(r"C:\royalties_pipeline")
INPUT_CSV = BASE / "reports" / "booking" / "booking_shows_report_base.csv"
OUTPUT_XLSX = BASE / "reports" / "booking" / "booking_shows_report.xlsx"


DETAIL_COLUMNS = [
    ("artista", "Artista"),
    ("fecha", "Fecha"),
    ("venue_evento", "Venue / Evento"),
    ("cachet_show", "Cachet show"),
    ("gastos", "Gastos"),
    ("neto_show", "Neto show"),
    ("porcentaje_artista", "% artista"),
    ("porcentaje_productora", "% productora"),
    ("se_lleva_artista", "Se lleva artista"),
    ("se_lleva_indyana", "Se lleva Indyana"),
    ("importe_artista_planilla", "Importe artista planilla"),
    ("importe_productora_planilla", "Importe productora planilla"),
    ("se_lleva_artista_movimientos", "Artista segun movimientos"),
    ("lineas_ingreso", "Lineas ingreso"),
    ("lineas_gasto", "Lineas gasto"),
    ("control", "Control"),
    ("archivo_origen", "Archivo origen"),
]


def autosize(ws, min_width: int = 10, max_width: int = 42) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min_width

        for cell in col[:500]:
            if cell.value is not None:
                width = max(width, min(len(str(cell.value)) + 2, max_width))

        ws.column_dimensions[letter].width = width


def style_header(ws, row: int, start_col: int, end_col: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="D9E2F3"))

    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border


def write_summary_sheet(wb: Workbook, df: pl.DataFrame) -> None:
    ws = wb.active
    ws.title = "Totales"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Reporte de shows booking"
    ws["A1"].font = Font(size=18, bold=True, color="1F4E78")
    ws["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(color="666666")

    total_cachet = df["cachet_show"].sum()
    total_gastos = df["gastos"].sum()
    total_neto = df["neto_show"].sum()
    total_artista = df["se_lleva_artista"].sum()
    total_indyana = df["se_lleva_indyana"].sum()

    kpis = [
        ("Shows", df.height),
        ("Cachet total", total_cachet),
        ("Gastos total", total_gastos),
        ("Neto total", total_neto),
        ("Se lleva artista", total_artista),
        ("Se lleva Indyana", total_indyana),
    ]

    ws.append([])
    ws.append(["Indicador", "Valor"])
    style_header(ws, 4, 1, 2)

    for label, value in kpis:
        ws.append([label, value])

    for row in range(6, 10):
        ws.cell(row=row, column=2).number_format = '$ #,##0'

    summary = (
        df.group_by("artista")
        .agg([
            pl.len().alias("shows"),
            pl.sum("cachet_show").alias("cachet_show"),
            pl.sum("gastos").alias("gastos"),
            pl.sum("neto_show").alias("neto_show"),
            pl.sum("se_lleva_artista").alias("se_lleva_artista"),
            pl.sum("se_lleva_indyana").alias("se_lleva_indyana"),
        ])
        .sort("se_lleva_indyana", descending=True)
    )

    start_row = 11
    headers = ["Artista", "Shows", "Cachet show", "Gastos", "Neto show", "Se lleva artista", "Se lleva Indyana"]
    ws.cell(row=start_row, column=1, value="Totales por artista")
    ws.cell(row=start_row, column=1).font = Font(size=13, bold=True, color="1F4E78")
    ws.append(headers)
    style_header(ws, start_row + 1, 1, len(headers))

    for row in summary.to_dicts():
        ws.append([
            row["artista"],
            row["shows"],
            row["cachet_show"],
            row["gastos"],
            row["neto_show"],
            row["se_lleva_artista"],
            row["se_lleva_indyana"],
        ])

    for row_idx in range(start_row + 2, ws.max_row + 1):
        for col_idx in [3, 4, 5, 6, 7]:
            ws.cell(row=row_idx, column=col_idx).number_format = '$ #,##0'

    table_ref = f"A{start_row + 1}:G{ws.max_row}"
    table = Table(displayName="TotalesPorArtista", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)

    autosize(ws)
    ws.freeze_panes = "A12"


def write_detail_sheet(wb: Workbook, df: pl.DataFrame) -> None:
    ws = wb.create_sheet("Shows")
    ws.sheet_view.showGridLines = False

    headers = [label for _, label in DETAIL_COLUMNS]
    ws.append(headers)
    style_header(ws, 1, 1, len(headers))

    for row in df.select([col for col, _ in DETAIL_COLUMNS]).to_dicts():
        ws.append([row[col] for col, _ in DETAIL_COLUMNS])

    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=2).number_format = "yyyy-mm-dd"

        for col_idx in [4, 5, 6, 9, 10, 11, 12, 13]:
            ws.cell(row=row_idx, column=col_idx).number_format = '$ #,##0'

        for col_idx in [7, 8]:
            ws.cell(row=row_idx, column=col_idx).number_format = "0%"

    table_ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    table = Table(displayName="ShowsBooking", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)

    ws.freeze_panes = "A2"
    autosize(ws, max_width=46)


def main() -> None:
    if not INPUT_CSV.exists():
        raise FileNotFoundError(INPUT_CSV)

    df = pl.read_csv(INPUT_CSV)

    wb = Workbook()
    write_summary_sheet(wb, df)
    write_detail_sheet(wb, df)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(OUTPUT_XLSX)
        output_path = OUTPUT_XLSX
    except PermissionError:
        output_path = OUTPUT_XLSX.with_name(
            f"{OUTPUT_XLSX.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{OUTPUT_XLSX.suffix}"
        )
        wb.save(output_path)

    print("Rows:", df.height)
    print("Output:", output_path)


if __name__ == "__main__":
    main()
