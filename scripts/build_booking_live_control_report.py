from __future__ import annotations

import argparse
import re
import sqlite3
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE = Path(r"C:\royalties_pipeline")
LIVE_DB = BASE / "warehouse" / "booking" / "live" / "booking_live.sqlite"
REPORT_DIR = BASE / "reports" / "booking"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
OK_FILL = PatternFill("solid", fgColor="E2F0D9")
BAD_FILL = PatternFill("solid", fgColor="FCE4D6")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="D9E2EC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY_FORMAT = '#,##0.00'
PERCENT_FORMAT = '0.00%'
INTEGER_FORMAT = '#,##0'


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")
    return slug or "artist"


def round_money(value: object) -> float:
    if value is None:
        return 0.0
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return 0.0


def fetch(sql: str, params: tuple[object, ...]) -> list[dict]:
    with sqlite3.connect(LIVE_DB) as conn:
        conn.row_factory = sqlite3.Row
        return [dict(row) for row in conn.execute(sql, params).fetchall()]


def safe_divide(amount: object, fx_rate: object) -> float:
    amount_value = round_money(amount)
    fx_value = round_money(fx_rate)
    if fx_value <= 0:
        return 0.0
    return round(amount_value / fx_value, 2)


def add_usd_fields(row: dict, pairs: list[tuple[str, str]]) -> dict:
    out = dict(row)
    fx_rate = out.get("TC")
    for ars_key, usd_key in pairs:
        out[usd_key] = safe_divide(out.get(ars_key), fx_rate)
    return out


def format_value(header: str, value: object) -> object:
    if value is None or value == "":
        return None
    if header in {"ID", "Show ID", "Filas"}:
        return int(value or 0)
    if header in {"% Artista", "% Indyana"}:
        return round_money(value)
    if header == "TC":
        return round_money(value)
    if header.endswith("ARS") or header.endswith("USD") or header in {"Importe", "Valor ARS", "Valor USD"}:
        if isinstance(value, str):
            return value
        return round_money(value)
    return value


def write_sheet(
    wb: Workbook,
    name: str,
    rows: list[dict],
    headers: list[str],
    *,
    money_headers: set[str],
    percent_headers: set[str] | None = None,
    integer_headers: set[str] | None = None,
    per_column_formats: dict[str, str] | None = None,
) -> None:
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    percent_headers = percent_headers or set()
    integer_headers = integer_headers or set()
    per_column_formats = per_column_formats or {}

    ws.append(headers)
    for row in rows:
        ws.append([format_value(header, row.get(header)) for header in headers])

    if not rows:
        ws.append(["Sin registros"] + [None] * (len(headers) - 1))

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            header = ws.cell(1, cell.column).value
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
            if header in money_headers:
                cell.number_format = MONEY_FORMAT
            elif header in percent_headers:
                cell.number_format = PERCENT_FORMAT
            elif header in integer_headers:
                cell.number_format = INTEGER_FORMAT
            elif header in per_column_formats:
                cell.number_format = per_column_formats[header]

            if isinstance(header, str) and header.startswith("Balance"):
                cell.fill = OK_FILL if abs(round_money(cell.value)) < 0.01 else BAD_FILL

    for col_idx in range(1, ws.max_column + 1):
        width = 10
        for (cell,) in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=min(ws.max_row, 300)):
            if cell.value is not None:
                width = max(width, min(len(str(cell.value)) + 2, 42))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_report(artist: str, keyword: str) -> Path:
    like = f"%{keyword.lower()}%"

    show_rows = fetch(
        """
        SELECT
            id AS ID,
            show_date AS Fecha,
            artist AS Artista,
            venue AS Venue,
            status AS Estado,
            fx_rate AS TC,
            cachet_amount AS "Cachet ARS",
            expenses_amount AS "Gastos ARS",
            net_amount AS "Neto ARS",
            pre_split_adjustments_amount AS "Ajustes pre split ARS",
            split_base_amount AS "Base split ARS",
            artist_percent / 100.0 AS "% Artista",
            producer_percent / 100.0 AS "% Indyana",
            artist_cash_target_amount AS "Objetivo artista ARS",
            artist_paid_amount AS "Pagado artista ARS",
            balance_artist_amount AS "Balance artista ARS",
            producer_cash_target_amount AS "Objetivo Indyana ARS",
            producer_received_amount AS "Rendido Indyana ARS",
            balance_producer_amount AS "Balance Indyana ARS",
            notes AS Notas
        FROM booking_shows
        WHERE lower(artist) LIKE ?
        ORDER BY show_date, id
        """,
        (like,),
    )

    show_rows = [
        add_usd_fields(row, [
            ("Cachet ARS", "Cachet USD"),
            ("Gastos ARS", "Gastos USD"),
            ("Neto ARS", "Neto USD"),
            ("Ajustes pre split ARS", "Ajustes pre split USD"),
            ("Base split ARS", "Base split USD"),
            ("Objetivo artista ARS", "Objetivo artista USD"),
            ("Pagado artista ARS", "Pagado artista USD"),
            ("Balance artista ARS", "Balance artista USD"),
            ("Objetivo Indyana ARS", "Objetivo Indyana USD"),
            ("Rendido Indyana ARS", "Rendido Indyana USD"),
            ("Balance Indyana ARS", "Balance Indyana USD"),
        ])
        for row in show_rows
    ]

    show_ids = [row["ID"] for row in show_rows]
    if show_ids:
        placeholders = ",".join("?" for _ in show_ids)
        expense_rows = fetch(
            f"""
            SELECT
                e.id AS ID,
                e.show_id AS "Show ID",
                s.show_date AS Fecha,
                s.venue AS Venue,
                e.category AS Categoria,
                e.concept AS Concepto,
                e.amount AS "Importe ARS",
                COALESCE(e.fx_rate, s.fx_rate) AS TC,
                e.notes AS Notas
            FROM booking_show_expenses e
            JOIN booking_shows s ON s.id = e.show_id
            WHERE e.show_id IN ({placeholders})
            ORDER BY s.show_date, e.show_id, e.id
            """,
            tuple(show_ids),
        )
        pre_split_rows = fetch(
            f"""
            SELECT
                a.id AS ID,
                a.show_id AS "Show ID",
                s.show_date AS Fecha,
                s.venue AS Venue,
                CASE WHEN a.destination = 'producer' THEN 'Indyana' ELSE 'Artista' END AS Destino,
                a.concept AS Concepto,
                a.amount AS "Importe ARS",
                COALESCE(a.fx_rate, s.fx_rate) AS TC,
                a.notes AS Notas
            FROM booking_pre_split_adjustments a
            JOIN booking_shows s ON s.id = a.show_id
            WHERE a.show_id IN ({placeholders})
            ORDER BY s.show_date, a.show_id, a.id
            """,
            tuple(show_ids),
        )
        adjustment_rows = fetch(
            f"""
            SELECT
                a.id AS ID,
                a.show_id AS "Show ID",
                s.show_date AS Fecha,
                s.venue AS Venue,
                a.concept AS Concepto,
                a.adjustment_type AS Tipo,
                a.area AS Area,
                a.impact AS Impacta,
                CASE WHEN a.recoverable = 1 THEN 'Si' ELSE 'No' END AS Recuperable,
                a.amount AS "Importe ARS",
                a.applied_amount AS "Aplicado ARS",
                a.artist_percent / 100.0 AS "% Artista",
                a.producer_percent / 100.0 AS "% Indyana",
                a.artist_amount AS "Costo artista ARS",
                a.producer_amount AS "Costo Indyana ARS",
                COALESCE(a.fx_rate, s.fx_rate) AS TC,
                a.notes AS Notas
            FROM booking_artist_adjustments a
            JOIN booking_shows s ON s.id = a.show_id
            WHERE a.show_id IN ({placeholders})
            ORDER BY s.show_date, a.show_id, a.id
            """,
            tuple(show_ids),
        )
        movement_rows = fetch(
            f"""
            SELECT
                m.id AS ID,
                m.show_id AS "Show ID",
                s.show_date AS Fecha,
                s.venue AS Venue,
                m.movement_type AS Tipo,
                m.category AS Categoria,
                m.amount AS "Importe ARS",
                COALESCE(m.fx_rate, s.fx_rate) AS TC,
                m.notes AS Notas
            FROM booking_movements m
            JOIN booking_shows s ON s.id = m.show_id
            WHERE m.show_id IN ({placeholders})
            ORDER BY s.show_date, m.show_id, m.id
            """,
            tuple(show_ids),
        )
    else:
        expense_rows = []
        pre_split_rows = []
        adjustment_rows = []
        movement_rows = []

    for rows in [expense_rows, pre_split_rows, movement_rows]:
        for row in rows:
            row["Importe USD"] = safe_divide(row.get("Importe ARS"), row.get("TC"))

    for row in adjustment_rows:
        for ars_key, usd_key in [
            ("Importe ARS", "Importe USD"),
            ("Aplicado ARS", "Aplicado USD"),
            ("Costo artista ARS", "Costo artista USD"),
            ("Costo Indyana ARS", "Costo Indyana USD"),
        ]:
            row[usd_key] = safe_divide(row.get(ars_key), row.get("TC"))

    total_cachet = sum(round_money(row.get("Cachet ARS")) for row in show_rows)
    total_gastos = sum(round_money(row.get("Gastos ARS")) for row in show_rows)
    total_artist = sum(round_money(row.get("Pagado artista ARS")) for row in show_rows)
    total_producer = sum(round_money(row.get("Rendido Indyana ARS")) for row in show_rows)
    total_cachet_usd = sum(round_money(row.get("Cachet USD")) for row in show_rows)
    total_gastos_usd = sum(round_money(row.get("Gastos USD")) for row in show_rows)
    total_artist_usd = sum(round_money(row.get("Pagado artista USD")) for row in show_rows)
    total_producer_usd = sum(round_money(row.get("Rendido Indyana USD")) for row in show_rows)

    summary_rows = [
        {"Indicador": "Artista", "Texto": artist, "Cantidad": None, "ARS": None, "USD": None},
        {"Indicador": "Generado", "Texto": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Cantidad": None, "ARS": None, "USD": None},
        {"Indicador": "Shows", "Texto": "", "Cantidad": len(show_rows), "ARS": None, "USD": None},
        {"Indicador": "Shows sin TC", "Texto": "", "Cantidad": sum(1 for row in show_rows if not row.get("TC")), "ARS": None, "USD": None},
        {"Indicador": "Cachet", "Texto": "", "Cantidad": None, "ARS": total_cachet, "USD": total_cachet_usd},
        {"Indicador": "Gastos", "Texto": "", "Cantidad": None, "ARS": total_gastos, "USD": total_gastos_usd},
        {"Indicador": "Artista", "Texto": "", "Cantidad": None, "ARS": total_artist, "USD": total_artist_usd},
        {"Indicador": "Indyana", "Texto": "", "Cantidad": None, "ARS": total_producer, "USD": total_producer_usd},
        {
            "Indicador": "Balance artista",
            "Texto": "",
            "Cantidad": None,
            "ARS": sum(round_money(row.get("Balance artista ARS")) for row in show_rows),
            "USD": sum(round_money(row.get("Balance artista USD")) for row in show_rows),
        },
        {
            "Indicador": "Balance Indyana",
            "Texto": "",
            "Cantidad": None,
            "ARS": sum(round_money(row.get("Balance Indyana ARS")) for row in show_rows),
            "USD": sum(round_money(row.get("Balance Indyana USD")) for row in show_rows),
        },
        {"Indicador": "Filas gastos", "Texto": "", "Cantidad": len(expense_rows), "ARS": None, "USD": None},
        {"Indicador": "Filas ajustes pre split", "Texto": "", "Cantidad": len(pre_split_rows), "ARS": None, "USD": None},
        {"Indicador": "Filas ajustes artista", "Texto": "", "Cantidad": len(adjustment_rows), "ARS": None, "USD": None},
    ]

    wb = Workbook()
    wb.remove(wb.active)

    write_sheet(
        wb,
        "Resumen",
        summary_rows,
        ["Indicador", "Texto", "Cantidad", "ARS", "USD"],
        money_headers={"ARS", "USD"},
        integer_headers={"Cantidad"},
    )
    wb["Resumen"]["C5"].fill = WARN_FILL if wb["Resumen"]["C5"].value else OK_FILL

    show_headers = [
        "ID", "Fecha", "Artista", "Venue", "Estado", "TC",
        "Cachet ARS", "Cachet USD", "Gastos ARS", "Gastos USD",
        "Neto ARS", "Neto USD", "Ajustes pre split ARS", "Ajustes pre split USD",
        "Base split ARS", "Base split USD", "% Artista", "% Indyana",
        "Objetivo artista ARS", "Objetivo artista USD", "Pagado artista ARS", "Pagado artista USD",
        "Balance artista ARS", "Balance artista USD", "Objetivo Indyana ARS", "Objetivo Indyana USD",
        "Rendido Indyana ARS", "Rendido Indyana USD", "Balance Indyana ARS", "Balance Indyana USD",
        "Notas",
    ]
    write_sheet(
        wb,
        "Shows",
        show_rows,
        show_headers,
        money_headers={header for header in show_headers if header.endswith("ARS") or header.endswith("USD")},
        percent_headers={"% Artista", "% Indyana"},
        integer_headers={"ID"},
        per_column_formats={"TC": "0.00"},
    )

    detail_headers = ["ID", "Show ID", "Fecha", "Venue", "Categoria", "Concepto", "Importe ARS", "Importe USD", "TC", "Notas"]
    write_sheet(
        wb,
        "Gastos",
        expense_rows,
        detail_headers,
        money_headers={"Importe ARS", "Importe USD"},
        integer_headers={"ID", "Show ID"},
        per_column_formats={"TC": "0.00"},
    )

    pre_headers = ["ID", "Show ID", "Fecha", "Venue", "Destino", "Concepto", "Importe ARS", "Importe USD", "TC", "Notas"]
    write_sheet(
        wb,
        "Ajustes pre split",
        pre_split_rows,
        pre_headers,
        money_headers={"Importe ARS", "Importe USD"},
        integer_headers={"ID", "Show ID"},
        per_column_formats={"TC": "0.00"},
    )

    adjustment_headers = [
        "ID", "Show ID", "Fecha", "Venue", "Concepto", "Tipo", "Area", "Impacta", "Recuperable",
        "Importe ARS", "Importe USD", "Aplicado ARS", "Aplicado USD", "% Artista", "% Indyana",
        "Costo artista ARS", "Costo artista USD", "Costo Indyana ARS", "Costo Indyana USD", "TC", "Notas",
    ]
    write_sheet(
        wb,
        "Ajustes artista",
        adjustment_rows,
        adjustment_headers,
        money_headers={header for header in adjustment_headers if header.endswith("ARS") or header.endswith("USD")},
        percent_headers={"% Artista", "% Indyana"},
        integer_headers={"ID", "Show ID"},
        per_column_formats={"TC": "0.00"},
    )

    movement_headers = ["ID", "Show ID", "Fecha", "Venue", "Tipo", "Categoria", "Importe ARS", "Importe USD", "TC", "Notas"]
    write_sheet(
        wb,
        "Movimientos caja",
        movement_rows,
        movement_headers,
        money_headers={"Importe ARS", "Importe USD"},
        integer_headers={"ID", "Show ID"},
        per_column_formats={"TC": "0.00"},
    )

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    output = REPORT_DIR / f"booking_live_control_report_{slugify(artist)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(output)

    loaded = load_workbook(output, data_only=True)
    assert set(loaded.sheetnames) == {"Resumen", "Shows", "Gastos", "Ajustes pre split", "Ajustes artista", "Movimientos caja"}
    return output


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--artist", required=True)
    parser.add_argument("--keyword", required=True)
    args = parser.parse_args()

    output = build_report(args.artist, args.keyword)
    print(output)


if __name__ == "__main__":
    main()
