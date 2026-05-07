from __future__ import annotations

import argparse
import re
import sqlite3
from datetime import datetime
from pathlib import Path

import polars as pl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE = Path(r"C:\royalties_pipeline")
HISTORICAL_CSV = BASE / "reports" / "booking" / "booking_shows_report_base.csv"
LIVE_DB = BASE / "warehouse" / "booking" / "live" / "booking_live.sqlite"
OUTPUT_DIR = BASE / "reports" / "booking"


MONEY_FMT = '$ #,##0;[Red]-$ #,##0;"-"'
PCT_FMT = "0%"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
OK_FILL = PatternFill("solid", fgColor="E2F0D9")
WHITE_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="D9E2EC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def normalize_text(value: object) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def slugify(value: str) -> str:
    text = normalize_text(value)
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return text or "artist"


def rebuild_historical_base() -> None:
    import subprocess
    import sys

    subprocess.run(
        [sys.executable, str(BASE / "scripts" / "build_booking_shows_report.py")],
        cwd=str(BASE),
        check=True,
    )


def live_show_keys() -> set[tuple[str, str, str]]:
    if not LIVE_DB.exists():
        return set()

    with sqlite3.connect(LIVE_DB) as conn:
        rows = conn.execute(
            """
            SELECT artist, show_date, venue
            FROM booking_shows
            """
        ).fetchall()

    return {
        (normalize_text(row[0]), str(row[1] or ""), normalize_text(row[2]))
        for row in rows
    }


def load_candidates(keyword: str) -> pl.DataFrame:
    if not HISTORICAL_CSV.exists():
        rebuild_historical_base()

    df = pl.read_csv(HISTORICAL_CSV)
    return (
        df
        .filter(pl.col("artista").cast(pl.Utf8).str.to_lowercase().str.contains(keyword))
        .with_columns([
            pl.col("artista").cast(pl.Utf8).str.strip_chars(),
            pl.col("fecha").cast(pl.Utf8),
            pl.col("venue_evento").cast(pl.Utf8).str.strip_chars(),
            pl.col("cachet_show").cast(pl.Float64, strict=False).fill_null(0),
            pl.col("gastos").cast(pl.Float64, strict=False).fill_null(0),
            pl.col("neto_show").cast(pl.Float64, strict=False).fill_null(0),
            pl.col("porcentaje_artista").cast(pl.Float64, strict=False),
            pl.col("porcentaje_productora").cast(pl.Float64, strict=False),
            pl.col("se_lleva_artista").cast(pl.Float64, strict=False).fill_null(0),
            pl.col("se_lleva_indyana").cast(pl.Float64, strict=False).fill_null(0),
        ])
        .sort(["fecha", "venue_evento", "archivo_origen"])
    )


def autosize(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        width = 10
        for (cell,) in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=min(ws.max_row, 300)):
            if cell.value is not None:
                width = max(width, min(len(str(cell.value)) + 2, 42))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def style_header(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def style_sheet(ws, money_headers: set[str], percent_headers: set[str]) -> None:
    style_header(ws)
    header_by_col = {idx: cell.value for idx, cell in enumerate(ws[1], start=1)}
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
            header = header_by_col.get(cell.column)
            if header in money_headers:
                cell.number_format = MONEY_FMT
            if header in percent_headers:
                cell.number_format = PCT_FMT
            if header == "estado_importacion":
                if cell.value == "pendiente":
                    cell.fill = WARN_FILL
                elif cell.value == "ya_cargado_posible":
                    cell.fill = OK_FILL
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.sheet_view.showGridLines = False
    autosize(ws)


def write_report(keyword: str, canonical_artist: str) -> Path:
    candidates = load_candidates(keyword)
    live_keys = live_show_keys()

    rows = []
    for item in candidates.to_dicts():
        key = (
            normalize_text(item["artista"]),
            str(item["fecha"] or ""),
            normalize_text(item["venue_evento"]),
        )
        status = "ya_cargado_posible" if key in live_keys else "pendiente"
        artist_percent = item["porcentaje_artista"]
        producer_percent = item["porcentaje_productora"]
        if artist_percent is None:
            artist_percent = 0.70
        if producer_percent is None:
            producer_percent = 1 - artist_percent

        rows.append({
            "estado_importacion": status,
            "cargar": "NO" if status != "pendiente" else "",
            "artista_sistema": canonical_artist,
            "fecha": item["fecha"],
            "venue": item["venue_evento"],
            "estado_show": "realizado" if item["control"] != "cachet_cero" else "no_cobrado",
            "cachet": item["cachet_show"],
            "gastos": item["gastos"],
            "neto": item["neto_show"],
            "porcentaje_artista": artist_percent,
            "porcentaje_indyana": producer_percent,
            "pagado_artista_sugerido": item["se_lleva_artista"],
            "rendido_indyana_sugerido": item["se_lleva_indyana"],
            "control_origen": item["control"],
            "archivo_origen": item["archivo_origen"],
            "lineas_ingreso": item["lineas_ingreso"],
            "lineas_gasto": item["lineas_gasto"],
            "nota_validacion": "",
        })

    output = OUTPUT_DIR / f"booking_import_candidates_{slugify(canonical_artist)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidatos"

    headers = list(rows[0].keys()) if rows else [
        "estado_importacion",
        "cargar",
        "artista_sistema",
        "fecha",
        "venue",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])

    style_sheet(
        ws,
        money_headers={"cachet", "gastos", "neto", "pagado_artista_sugerido", "rendido_indyana_sugerido"},
        percent_headers={"porcentaje_artista", "porcentaje_indyana"},
    )

    summary = wb.create_sheet("Resumen")
    summary.append(["Indicador", "Valor"])
    summary.append(["Artista", canonical_artist])
    summary.append(["Coincidencias historicas", len(rows)])
    summary.append(["Pendientes", sum(1 for row in rows if row["estado_importacion"] == "pendiente")])
    summary.append(["Ya cargados posibles", sum(1 for row in rows if row["estado_importacion"] == "ya_cargado_posible")])
    summary.append(["Cachet pendiente", sum(row["cachet"] for row in rows if row["estado_importacion"] == "pendiente")])
    summary.append(["Gastos pendiente", sum(row["gastos"] for row in rows if row["estado_importacion"] == "pendiente")])
    summary.append(["Indyana pendiente", sum(row["rendido_indyana_sugerido"] for row in rows if row["estado_importacion"] == "pendiente")])
    style_sheet(summary, money_headers={"Valor"}, percent_headers=set())

    wb.save(output)
    loaded = load_workbook(output, data_only=True)
    assert "Candidatos" in loaded.sheetnames
    return output


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--keyword", default="virr|virs|virsh")
    parser.add_argument("--artist", default="Virrshi Dj")
    parser.add_argument("--rebuild", action="store_true")
    args = parser.parse_args()

    if args.rebuild:
        rebuild_historical_base()

    output = write_report(args.keyword, args.artist)
    print(output)


if __name__ == "__main__":
    main()
