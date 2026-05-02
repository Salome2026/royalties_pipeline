from pathlib import Path

import duckdb
import pandas as pd

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


DETAIL_PATH = r"C:\royalties_pipeline\warehouse\detail\royalties_detail.parquet"
OUTPUT_PATH = Path(r"C:\royalties_pipeline\reports\reporte_ingresos_digitales_por_mes_de_statement.xlsx")

print("Generando reporte ingresos digitales por statement...")

con = duckdb.connect()

# =========================
# DETECTAR SI EXISTE FLAG GUSTY
# =========================
cols_df = con.execute(
    "DESCRIBE SELECT * FROM read_parquet(?)",
    [DETAIL_PATH]
).fetchdf()

detail_columns = set(cols_df["column_name"].tolist())

HAS_SHARE_FLAG_COL = "has_share_in_out" in detail_columns

if HAS_SHARE_FLAG_COL:
    share_flag_sql = """
    MAX(
        CASE
            WHEN COALESCE(TRY_CAST(has_share_in_out AS BOOLEAN), FALSE)
            THEN 1
            ELSE 0
        END
    ) AS has_share_in_out
    """
else:
    share_flag_sql = """
    0 AS has_share_in_out
    """


query = f"""
SELECT
    source,
    account,
    COALESCE(NULLIF(TRIM(artist_statement_style), ''), 'SIN ARTISTA') AS artist,
    statement_period,
    ROUND(SUM(
        CASE
            WHEN source = 'fuga'
            THEN COALESCE(net_amount_usd, net_amount) * 0.977832
            ELSE COALESCE(net_amount_usd, net_amount)
        END
    ), 2) AS total,
    {share_flag_sql}
FROM read_parquet(?)
WHERE statement_period IS NOT NULL
  AND TRIM(statement_period) <> ''
GROUP BY
    source,
    account,
    COALESCE(NULLIF(TRIM(artist_statement_style), ''), 'SIN ARTISTA'),
    statement_period
ORDER BY
    source,
    account,
    artist,
    statement_period
"""

df = con.execute(query, [DETAIL_PATH]).fetchdf()

if df.empty:
    print("ERROR: No hay datos para generar el reporte.")
    raise SystemExit(1)

OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)


# =========================
# ESTILOS
# =========================
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

ARTIST_FILL = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
TITLE_FILL = PatternFill(start_color="EAF2F8", end_color="EAF2F8", fill_type="solid")

# NUEVO: color suave para alerta Gusty Shares
SHARE_ALERT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

TITLE_FONT = Font(size=16, bold=True, color="1F4E78")
TOTAL_FONT = Font(bold=True)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")


def format_sheet(ws, pivot, company_name, share_flags=None):
    """
    Formatea una hoja ya escrita con pivot.to_excel(..., startrow=1).
    Layout:
      A1 = título
      A2 = ARTISTA
      B2.. = meses
      última columna = TOTAL

    share_flags:
      DataFrame opcional con mismo layout artista/statement_period.
      Solo se usa para pintar celdas; no modifica importes.
    """

    max_col = len(pivot.columns) + 1
    max_row = len(pivot) + 2

    # Limpiar fila 1 desde B hacia la derecha, por si Excel escribió algo
    for col in range(2, max_col + 1):
        ws.cell(row=1, column=col).value = None

    # Título
    ws["A1"] = company_name
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = LEFT_ALIGN
    ws.row_dimensions[1].height = 24

    # Encabezados
    ws["A2"] = "ARTISTA"
    ws["A2"].fill = HEADER_FILL
    ws["A2"].font = HEADER_FONT
    ws["A2"].alignment = CENTER_ALIGN

    for col_idx, col_name in enumerate(pivot.columns, start=2):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = col_name
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN

    # Formato de datos
    for row in range(3, max_row + 1):
        label = str(ws.cell(row=row, column=1).value or "")

        artist_cell = ws.cell(row=row, column=1)
        artist_cell.fill = ARTIST_FILL
        artist_cell.alignment = LEFT_ALIGN

        is_total_row = label.upper().startswith("TOTAL")

        if is_total_row:
            artist_cell.fill = TOTAL_FILL
            artist_cell.font = TOTAL_FONT

        for col in range(2, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.number_format = '#,##0.00'
            cell.alignment = CENTER_ALIGN

            if is_total_row:
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT

    # NUEVO: pintar alertas de Shares In/Out solo donde corresponda
    if share_flags is not None and not share_flags.empty:
        for artist_idx, artist_name in enumerate(pivot.index, start=3):
            artist_label = str(artist_name)

            if artist_label.upper().startswith("TOTAL"):
                continue

            if artist_name not in share_flags.index:
                continue

            for col_idx, period in enumerate(pivot.columns, start=2):
                if period == "TOTAL":
                    continue

                if period not in share_flags.columns:
                    continue

                try:
                    has_alert = int(share_flags.loc[artist_name, period]) == 1
                except Exception:
                    has_alert = False

                if has_alert:
                    cell = ws.cell(row=artist_idx, column=col_idx)
                    cell.fill = SHARE_ALERT_FILL

    # Resaltar columna TOTAL
    total_col = max_col
    for row in range(2, max_row + 1):
        cell = ws.cell(row=row, column=total_col)
        if row == 2:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        else:
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT

    # Anchos
    ws.column_dimensions["A"].width = 30

    for col in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # Congelar columna artista + encabezados
    ws.freeze_panes = "B3"

    # Filtro
    ws.auto_filter.ref = f"A2:{get_column_letter(max_col)}{max_row}"


def add_totals_and_clean(pivot):
    """
    Agrega TOTAL a la derecha y elimina artistas cuyo total sea 0.
    """
    pivot = pivot.reindex(sorted(pivot.columns), axis=1)
    pivot["TOTAL"] = pivot.sum(axis=1)

    before_rows = len(pivot)
    pivot = pivot[pivot["TOTAL"].round(2) != 0]
    removed_rows = before_rows - len(pivot)

    return pivot, removed_rows


with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:

    # =========================
    # HOJA 1: TOTAL CONSOLIDADO
    # =========================
    pivot_total = df.pivot_table(
        index="artist",
        columns="statement_period",
        values="total",
        aggfunc="sum",
        fill_value=0
    )

    pivot_total, _ = add_totals_and_clean(pivot_total)

    if not pivot_total.empty:
        total_general = pivot_total.sum(axis=0)
        total_general.name = "TOTAL USD"

        pivot_total = pd.concat([pivot_total, total_general.to_frame().T])

        pivot_total.to_excel(
            writer,
            sheet_name="TOTAL",
            startrow=1
        )

        ws_total = writer.sheets["TOTAL"]
        format_sheet(ws_total, pivot_total, "SALDO TOTAL ARTISTAS")

    # =========================
    # HOJAS POR SOURCE / ACCOUNT
    # =========================
    for (source, account), group in df.groupby(["source", "account"], dropna=False):

        pivot = group.pivot_table(
            index="artist",
            columns="statement_period",
            values="total",
            aggfunc="sum",
            fill_value=0
        )

        pivot, removed_rows = add_totals_and_clean(pivot)

        if removed_rows > 0:
            print(f"  - {source}/{account}: artistas eliminados por TOTAL = 0: {removed_rows}")

        if pivot.empty:
            print(f"  - {source}/{account}: sin artistas con total distinto de 0. Se omite hoja.")
            continue

        total_usd = pivot.sum(axis=0)
        total_usd.name = "TOTAL USD"

        rows_to_add = [total_usd.to_frame().T]

        # =========================
        # FUGA: agregar TOTAL EUR
        # =========================
        if source == "fuga":
            query_eur = """
            SELECT
                statement_period AS period,
                ROUND(SUM(net_amount), 2) AS total_eur
            FROM read_parquet(?)
            WHERE source = ?
              AND account = ?
              AND statement_period IS NOT NULL
              AND TRIM(statement_period) <> ''
            GROUP BY statement_period
            """

            eur_df = con.execute(query_eur, [DETAIL_PATH, source, account]).fetchdf()

            total_eur = pd.Series(0.0, index=pivot.columns)
            total_eur.name = "TOTAL EUR"

            for col in pivot.columns:
                if col == "TOTAL":
                    continue

                matches = eur_df.loc[eur_df["period"] == col, "total_eur"]

                if not matches.empty:
                    total_eur[col] = float(matches.iloc[0])

            total_eur["TOTAL"] = total_eur.drop(labels=["TOTAL"], errors="ignore").sum()

            rows_to_add.append(total_eur.to_frame().T)

        # =========================
        # NUEVO: flags Gusty DJ para pintar celdas
        # =========================
        share_flags = None

        if source == "onerpm" and account == "gusty_dj":
            flag_group = group.copy()

            if "has_share_in_out" in flag_group.columns:
                share_flags = flag_group.pivot_table(
                    index="artist",
                    columns="statement_period",
                    values="has_share_in_out",
                    aggfunc="max",
                    fill_value=0
                )

                share_flags = share_flags.reindex(
                    index=pivot.index,
                    columns=[c for c in pivot.columns if c != "TOTAL"],
                    fill_value=0
                )

                flagged_cells = int(share_flags.sum().sum()) if not share_flags.empty else 0
                print(f"  - onerpm/gusty_dj: celdas con alerta Shares In/Out: {flagged_cells}")

        pivot = pd.concat([pivot] + rows_to_add)

        sheet_name = f"{source}_{account}"[:31]

        pivot.to_excel(
            writer,
            sheet_name=sheet_name,
            startrow=1
        )

        ws = writer.sheets[sheet_name]
        format_sheet(ws, pivot, f"{str(source).upper()} - {account}", share_flags=share_flags)


print("Reporte generado en:")
print(OUTPUT_PATH)