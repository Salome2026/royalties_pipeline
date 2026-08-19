from __future__ import annotations

import argparse
import re
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
import polars as pl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from lib.catalog_report_filter import apply_report_net_personalization, current_catalog_status_path, with_catalog_report_status
    from lib.store_taxonomy import ensure_store_dimensions
except ModuleNotFoundError:
    from scripts.lib.catalog_report_filter import apply_report_net_personalization, current_catalog_status_path, with_catalog_report_status
    from scripts.lib.store_taxonomy import ensure_store_dimensions

BASE = Path(r"C:\royalties_pipeline")
MARTS_DIR = BASE / "warehouse" / "marts"
REPORTS_DIR = BASE / "reports" / "api"

STANDARDIZED_ONERPM_PATH = MARTS_DIR / "standardized_raw_onerpm.parquet"
STANDARDIZED_FUGA_PATH = MARTS_DIR / "standardized_raw_fuga.parquet"
CATALOG_RELEASE_METADATA_PATH = MARTS_DIR / "catalog_release_metadata.parquet"
CATALOG_MASTER_PATH = MARTS_DIR / "catalog_master.parquet"

SOURCE = "onerpm"
ACCOUNT = "la_nueva_sangre"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOTAL_FILL = PatternFill("solid", fgColor="D9EAF7")
BOX_FILL = PatternFill("solid", fgColor="FCE4D6")
PROPIO_FILL = PatternFill("solid", fgColor="FFF2CC")
OK_FILL = PatternFill("solid", fgColor="E2F0D9")
FUGA_FILL = PatternFill("solid", fgColor="EAF4FF")
TITLE_FONT = Font(size=14, bold=True, color="1F4E78")
TOTAL_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
ZERO_EPSILON = 0.005
REPORT_PRESENTATION_DROP_COLUMNS = {
    "Hoja origen",
    "ISRC original",
    "Resolucion",
    "Metadata",
}


def has_col(schema: dict[str, pl.DataType], col: str) -> bool:
    return col in schema


def col_or_null(schema: dict[str, pl.DataType], col: str, dtype=pl.Utf8) -> pl.Expr:
    if has_col(schema, col):
        return pl.col(col).cast(dtype, strict=False)
    return pl.lit(None).cast(dtype)


def amount_expr(schema: dict[str, pl.DataType]) -> pl.Expr:
    candidates = [
        pl.col(col).cast(pl.Float64, strict=False)
        for col in ["amount_usd", "net_amount_usd", "net_amount"]
        if has_col(schema, col)
    ]
    return pl.coalesce(candidates) if candidates else pl.lit(0.0)


def first_text(schema: dict[str, pl.DataType], columns: list[str]) -> pl.Expr:
    exprs = [
        pl.col(col).cast(pl.Utf8, strict=False).str.strip_chars()
        for col in columns
        if has_col(schema, col)
    ]
    if not exprs:
        return pl.lit(None).cast(pl.Utf8)
    return pl.coalesce(exprs)


def release_metadata_provider_expr(schema: dict[str, pl.DataType]) -> pl.Expr:
    source_sheet = col_or_null(schema, "source_sheet")
    return (
        pl.when(source_sheet == "Masters")
        .then(pl.lit("spotify"))
        .when(source_sheet == "Youtube Channels")
        .then(pl.lit("youtube"))
        .otherwise(pl.lit(None).cast(pl.Utf8))
    )


def release_metadata_lookup_key_expr(schema: dict[str, pl.DataType]) -> pl.Expr:
    source_sheet = col_or_null(schema, "source_sheet")
    isrc = first_text(schema, ["ISRC", "asset_isrc"])
    video_id = first_text(schema, ["Video ID", "ID", "Parent ID", "Video ID"])
    return (
        pl.when((source_sheet == "Masters") & isrc.is_not_null() & (isrc != ""))
        .then(pl.concat_str([pl.lit("isrc:"), isrc.str.to_uppercase()]))
        .when((source_sheet == "Youtube Channels") & video_id.is_not_null() & (video_id != ""))
        .then(pl.concat_str([pl.lit("video:"), video_id]))
        .otherwise(pl.lit(None).cast(pl.Utf8))
    )


def metadata_ready() -> pl.LazyFrame:
    if not CATALOG_RELEASE_METADATA_PATH.exists():
        return pl.DataFrame({
            "_metadata_provider": pl.Series([], dtype=pl.Utf8),
            "_metadata_lookup_key": pl.Series([], dtype=pl.Utf8),
            "_include_after_release_cutoff": pl.Series([], dtype=pl.Boolean),
            "release_date": pl.Series([], dtype=pl.Utf8),
            "release_year_month": pl.Series([], dtype=pl.Utf8),
            "external_label": pl.Series([], dtype=pl.Utf8),
            "metadata_status": pl.Series([], dtype=pl.Utf8),
            "match_url": pl.Series([], dtype=pl.Utf8),
        }).lazy()

    return (
        pl.scan_parquet(CATALOG_RELEASE_METADATA_PATH)
        .filter(
            (pl.col("source") == SOURCE)
            & (pl.col("account") == ACCOUNT)
        )
        .select([
            pl.col("preferred_provider").alias("_metadata_provider"),
            pl.col("preferred_lookup_key").alias("_metadata_lookup_key"),
            pl.col("include_after_release_cutoff")
            .cast(pl.Boolean, strict=False)
            .alias("_include_after_release_cutoff"),
            pl.col("release_date").cast(pl.Utf8, strict=False),
            pl.col("release_year_month").cast(pl.Utf8, strict=False),
            pl.col("external_label").cast(pl.Utf8, strict=False),
            pl.col("metadata_status").cast(pl.Utf8, strict=False),
            pl.col("match_url").cast(pl.Utf8, strict=False),
        ])
        .unique(["_metadata_provider", "_metadata_lookup_key"])
    )


def catalog_label_overrides() -> pl.DataFrame:
    status_path = current_catalog_status_path()
    if not status_path.exists():
        return pl.DataFrame({
            "catalog_key": pl.Series([], dtype=pl.Utf8),
            "_label_normalized_override": pl.Series([], dtype=pl.Utf8),
        })

    status = pl.read_parquet(status_path)
    if status.is_empty() or "catalog_key" not in status.columns or "label_normalized_override" not in status.columns:
        return pl.DataFrame({
            "catalog_key": pl.Series([], dtype=pl.Utf8),
            "_label_normalized_override": pl.Series([], dtype=pl.Utf8),
        })

    override = pl.col("label_normalized_override").cast(pl.Utf8, strict=False).str.strip_chars()
    return (
        status
        .select([
            pl.col("catalog_key").cast(pl.Utf8, strict=False),
            pl.when(override == "")
            .then(pl.lit(None).cast(pl.Utf8))
            .otherwise(override)
            .alias("_label_normalized_override"),
        ])
        .filter(pl.col("catalog_key").is_not_null())
        .unique(["catalog_key"], keep="last")
    )


def catalog_label_lookup() -> pl.DataFrame:
    empty = pl.DataFrame({
        "catalog_key": pl.Series([], dtype=pl.Utf8),
        "label_normalized_report": pl.Series([], dtype=pl.Utf8),
    })
    if not CATALOG_MASTER_PATH.exists():
        return empty

    catalog = pl.read_parquet(CATALOG_MASTER_PATH)
    if "catalog_key" not in catalog.columns:
        return empty

    columns = set(catalog.columns)
    label_candidates = [
        pl.col(col).cast(pl.Utf8, strict=False).str.strip_chars()
        for col in ["label_normalized", "label_normalized_auto", "external_label"]
        if col in columns
    ]
    if not label_candidates:
        label_expr = pl.lit(None).cast(pl.Utf8)
    else:
        label_expr = pl.coalesce(label_candidates)

    labels = catalog.select([
        pl.col("catalog_key").cast(pl.Utf8, strict=False),
        label_expr.alias("_label_normalized_catalog"),
    ])
    overrides = catalog_label_overrides()
    if not overrides.is_empty():
        labels = labels.join(overrides, on="catalog_key", how="left")
    else:
        labels = labels.with_columns(pl.lit(None).cast(pl.Utf8).alias("_label_normalized_override"))

    return (
        labels
        .with_columns(
            pl.coalesce(["_label_normalized_override", "_label_normalized_catalog"]).alias("label_normalized_report")
        )
        .select(["catalog_key", "label_normalized_report"])
        .unique(["catalog_key"])
    )


def catalog_label_by_video() -> pl.DataFrame:
    empty = pl.DataFrame({
        "_video_id": pl.Series([], dtype=pl.Utf8),
        "Video label normalizado": pl.Series([], dtype=pl.Utf8),
    })
    if not CATALOG_MASTER_PATH.exists():
        return empty

    catalog = pl.read_parquet(CATALOG_MASTER_PATH)
    if "catalog_key" not in catalog.columns:
        return empty

    rows: list[pl.DataFrame] = []
    if "video_ids" in catalog.columns:
        rows.append(
            catalog
            .select(["catalog_key", "video_ids"])
            .with_columns(pl.col("video_ids").fill_null("").str.split(" | ").alias("_video_id"))
            .explode("_video_id")
            .with_columns(pl.col("_video_id").cast(pl.Utf8, strict=False).str.strip_chars())
            .filter(pl.col("_video_id").is_not_null() & (pl.col("_video_id") != ""))
            .select(["catalog_key", "_video_id"])
        )
    if "identity_video_id" in catalog.columns:
        rows.append(
            catalog
            .select([
                "catalog_key",
                pl.col("identity_video_id").cast(pl.Utf8, strict=False).str.strip_chars().alias("_video_id"),
            ])
            .filter(pl.col("_video_id").is_not_null() & (pl.col("_video_id") != ""))
        )
    if rows:
        aliases = pl.concat(rows, how="diagonal_relaxed").unique(["catalog_key", "_video_id"])
    else:
        aliases = pl.DataFrame({
            "catalog_key": pl.Series([], dtype=pl.Utf8),
            "_video_id": pl.Series([], dtype=pl.Utf8),
        })

    labels = catalog_label_lookup()
    if aliases.is_empty() or labels.is_empty():
        return empty

    return (
        aliases
        .join(labels, on="catalog_key", how="left")
        .filter(pl.col("label_normalized_report").is_not_null())
        .group_by("_video_id")
        .agg(pl.first("label_normalized_report").alias("Video label normalizado"))
    )


def metadata_by_isrc() -> pl.DataFrame:
    meta = catalog_metadata_by_key()
    if meta.is_empty() or "_resolved_isrc" not in meta.columns:
        return pl.DataFrame({
            "_isrc": pl.Series([], dtype=pl.Utf8),
            "Release date": pl.Series([], dtype=pl.Utf8),
            "Label normalizado": pl.Series([], dtype=pl.Utf8),
            "Metadata": pl.Series([], dtype=pl.Utf8),
            "_release_year_month": pl.Series([], dtype=pl.Utf8),
        })

    return (
        meta
        .filter(pl.col("_resolved_isrc").is_not_null())
        .group_by("_resolved_isrc")
        .agg([
            pl.first("Release date").cast(pl.Utf8, strict=False).alias("Release date"),
            pl.first("Label normalizado").cast(pl.Utf8, strict=False).alias("Label normalizado"),
            pl.first("Metadata").cast(pl.Utf8, strict=False).alias("Metadata"),
            pl.first("_release_year_month").cast(pl.Utf8, strict=False).alias("_release_year_month"),
        ])
        .rename({"_resolved_isrc": "_isrc"})
    )


def catalog_metadata_by_key() -> pl.DataFrame:
    empty = pl.DataFrame({
        "catalog_key": pl.Series([], dtype=pl.Utf8),
        "_resolved_isrc": pl.Series([], dtype=pl.Utf8),
        "Release date": pl.Series([], dtype=pl.Utf8),
        "Label normalizado": pl.Series([], dtype=pl.Utf8),
        "Metadata": pl.Series([], dtype=pl.Utf8),
        "_release_year_month": pl.Series([], dtype=pl.Utf8),
    })
    if not CATALOG_MASTER_PATH.exists():
        return empty

    catalog = pl.read_parquet(CATALOG_MASTER_PATH)
    if "catalog_key" not in catalog.columns:
        return empty

    columns = set(catalog.columns)
    resolved_isrc = (
        pl.coalesce([
            pl.col(col).cast(pl.Utf8, strict=False).str.strip_chars()
            for col in ["asset_isrc", "identity_asset_isrc"]
            if col in columns
        ])
        if {"asset_isrc", "identity_asset_isrc"} & columns
        else pl.lit(None).cast(pl.Utf8)
    )
    resolved_isrc = (
        pl.when(resolved_isrc.is_not_null() & (resolved_isrc != ""))
        .then(resolved_isrc)
        .when(pl.col("catalog_key").cast(pl.Utf8).str.starts_with("ISRC:"))
        .then(pl.col("catalog_key").cast(pl.Utf8).str.strip_prefix("ISRC:"))
        .otherwise(pl.lit(None).cast(pl.Utf8))
    )

    release_date = (
        pl.col("external_release_date").cast(pl.Utf8, strict=False)
        if "external_release_date" in columns
        else pl.lit(None).cast(pl.Utf8)
    )
    release_year_month = (
        pl.col("external_release_year_month").cast(pl.Utf8, strict=False)
        if "external_release_year_month" in columns
        else release_date.str.slice(0, 7)
    )
    metadata_status = (
        pl.col("external_metadata_status").cast(pl.Utf8, strict=False)
        if "external_metadata_status" in columns
        else pl.lit(None).cast(pl.Utf8)
    )

    out = (
        catalog
        .select([
            pl.col("catalog_key").cast(pl.Utf8, strict=False),
            resolved_isrc.alias("_resolved_isrc"),
            release_date.alias("Release date"),
            metadata_status.alias("Metadata"),
            release_year_month.alias("_release_year_month"),
        ])
        .unique(["catalog_key"])
    )
    labels = catalog_label_lookup()
    if not labels.is_empty():
        out = out.join(labels, on="catalog_key", how="left")
    else:
        out = out.with_columns(pl.lit(None).cast(pl.Utf8).alias("label_normalized_report"))

    return out.rename({"label_normalized_report": "Label normalizado"})


def metadata_by_video() -> pl.DataFrame:
    if not CATALOG_RELEASE_METADATA_PATH.exists():
        return pl.DataFrame({
            "_video_id": pl.Series([], dtype=pl.Utf8),
            "Video release date": pl.Series([], dtype=pl.Utf8),
            "Video label normalizado": pl.Series([], dtype=pl.Utf8),
            "Video metadata": pl.Series([], dtype=pl.Utf8),
            "Video match url": pl.Series([], dtype=pl.Utf8),
            "_video_release_year_month": pl.Series([], dtype=pl.Utf8),
        })

    meta = pl.read_parquet(CATALOG_RELEASE_METADATA_PATH)
    if "lookup_video_id" not in meta.columns:
        return pl.DataFrame({
            "_video_id": pl.Series([], dtype=pl.Utf8),
            "Video release date": pl.Series([], dtype=pl.Utf8),
            "Video label normalizado": pl.Series([], dtype=pl.Utf8),
            "Video metadata": pl.Series([], dtype=pl.Utf8),
            "Video match url": pl.Series([], dtype=pl.Utf8),
            "_video_release_year_month": pl.Series([], dtype=pl.Utf8),
        })

    grouped = (
        meta
        .filter(pl.col("lookup_video_id").is_not_null())
        .sort("amount_usd", descending=True)
        .group_by("lookup_video_id")
        .agg([
            pl.first("release_date").cast(pl.Utf8, strict=False).alias("Video release date"),
            pl.first("external_label").cast(pl.Utf8, strict=False).alias("_video_label_metadata"),
            pl.first("metadata_status").cast(pl.Utf8, strict=False).alias("Video metadata"),
            pl.first("match_url").cast(pl.Utf8, strict=False).alias("Video match url"),
            pl.first("release_year_month").cast(pl.Utf8, strict=False).alias("_video_release_year_month"),
        ])
        .rename({"lookup_video_id": "_video_id"})
    )
    labels = catalog_label_by_video()
    if not labels.is_empty():
        grouped = grouped.join(labels, on="_video_id", how="left")
    else:
        grouped = grouped.with_columns(pl.lit(None).cast(pl.Utf8).alias("Video label normalizado"))
    return (
        grouped
        .with_columns(
            pl.coalesce(["Video label normalizado", "_video_label_metadata"]).alias("Video label normalizado")
        )
        .drop("_video_label_metadata")
    )


def onerpm_lns_isrcs() -> pl.DataFrame:
    if not STANDARDIZED_ONERPM_PATH.exists():
        return pl.DataFrame({
            "_isrc": pl.Series([], dtype=pl.Utf8),
            "_onerpm_usd": pl.Series([], dtype=pl.Float64),
            "_onerpm_from_tx": pl.Series([], dtype=pl.Utf8),
            "_onerpm_to_tx": pl.Series([], dtype=pl.Utf8),
        })

    return (
        pl.scan_parquet(STANDARDIZED_ONERPM_PATH)
        .filter((pl.col("source") == SOURCE) & (pl.col("account") == ACCOUNT))
        .with_columns(first_text({"ISRC": pl.Utf8}, ["ISRC"]).alias("_isrc"))
        .filter(pl.col("_isrc").is_not_null() & (pl.col("_isrc") != ""))
        .group_by("_isrc")
        .agg([
            pl.sum("amount_usd").alias("_onerpm_usd"),
            pl.min("transaction_month").alias("_onerpm_from_tx"),
            pl.max("transaction_month").alias("_onerpm_to_tx"),
        ])
        .collect()
    )


def normalize_match_text(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^a-z0-9 ]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def presentation_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    return df.drop(columns=[col for col in REPORT_PRESENTATION_DROP_COLUMNS if col in df.columns])


def combined_income_sheet(onerpm: pd.DataFrame, fuga: pd.DataFrame) -> pd.DataFrame:
    frames = []
    for source, df in [("ONErpm", onerpm), ("FUGA", fuga)]:
        if df.empty:
            continue
        frame = df.copy()
        frame.insert(0, "Fuente", source)
        frames.append(frame)

    if not frames:
        return pd.DataFrame()

    column_order = ["Fuente"]
    for frame in frames:
        for col in frame.columns:
            if col not in column_order:
                column_order.append(col)

    aligned = []
    for frame in frames:
        work = frame.copy()
        for col in column_order:
            if col not in work.columns:
                work[col] = pd.NA
        aligned.append(work[column_order])

    combined = pd.concat(aligned, ignore_index=True)
    if "Ingresos USD" in combined.columns:
        combined = combined.sort_values("Ingresos USD", ascending=False, kind="mergesort")
    return combined


def filter_detail_to_parent(detail: pd.DataFrame, parent: pd.DataFrame) -> pd.DataFrame:
    if detail.empty or parent.empty:
        return detail.iloc[0:0].copy()

    parent_key_cols = [
        "Tema / video",
        "Artista statement",
        "ISRC",
        "UPC",
        "Video ID",
        "Release date",
        "Label normalizado",
        "Origen",
    ]
    key_cols = [
        col
        for col in parent_key_cols
        if col in detail.columns and col in parent.columns
    ]
    if not key_cols:
        return detail

    sentinel = "__VPO_NULL__"
    detail_work = detail.copy()
    parent_keys = parent[key_cols].copy()
    for col in key_cols:
        detail_work[col] = detail_work[col].fillna(sentinel).astype(str)
        parent_keys[col] = parent_keys[col].fillna(sentinel).astype(str)

    allowed = parent_keys.drop_duplicates()
    filtered = detail_work.merge(allowed, on=key_cols, how="inner")
    for col in key_cols:
        filtered[col] = filtered[col].replace(sentinel, pd.NA)
    return filtered


def apply_related_video_exclusions(rows: pl.DataFrame) -> pl.DataFrame:
    if rows.is_empty():
        return rows

    reportable_master_titles = (
        rows
        .filter(
            (pl.col("_source_sheet") == "Masters")
            & (pl.col("_business_ok") == True)
            & pl.col("_title").is_not_null()
        )
        .select("_title")
        .unique()
        .to_series()
        .to_list()
    )
    reportable_master_terms = {
        normalize_match_text(title)
        for title in reportable_master_titles
        if len(normalize_match_text(title)) >= 8
    }

    excluded_masters = (
        rows
        .filter(
            (pl.col("_source_sheet") == "Masters")
            & (pl.col("_catalog_ok") == True)
            & (pl.col("_business_ok") == False)
            & pl.col("_title").is_not_null()
        )
        .select("_title")
        .unique()
        .to_series()
        .to_list()
    )
    master_terms = [
        (title, normalize_match_text(title))
        for title in excluded_masters
        if (
            len(normalize_match_text(title)) >= 8
            and normalize_match_text(title) not in reportable_master_terms
        )
    ]
    if not master_terms:
        return rows

    video_titles = (
        rows
        .filter(
            (pl.col("_source_sheet") == "Youtube Channels")
            & (pl.col("_business_ok") == True)
            & pl.col("_title").is_not_null()
        )
        .select("_title")
        .unique()
        .to_series()
        .to_list()
    )

    related = []
    for video_title in video_titles:
        normalized_video = normalize_match_text(video_title)
        for master_title, normalized_master in master_terms:
            if normalized_master and normalized_master in normalized_video:
                related.append({
                    "_title": video_title,
                    "_related_excluded_master": master_title,
                })
                break

    if not related:
        return rows.with_columns(pl.lit(None).cast(pl.Utf8).alias("_related_excluded_master"))

    related_df = pl.DataFrame(related)
    return (
        rows
        .join(related_df, on="_title", how="left")
        .with_columns([
            (
                (pl.col("_source_sheet") == "Youtube Channels")
                & (pl.col("_business_ok") == True)
                & pl.col("_related_excluded_master").is_not_null()
            ).alias("_is_related_excluded_video")
        ])
        .with_columns([
            pl.when(pl.col("_is_related_excluded_video"))
            .then(pl.lit(False))
            .otherwise(pl.col("_business_ok"))
            .alias("_business_ok"),
            pl.when(pl.col("_is_related_excluded_video"))
            .then(pl.lit("excluido_video_relacionado"))
            .otherwise(pl.col("_reason"))
            .alias("_reason"),
        ])
        .drop("_is_related_excluded_video")
    )


def classified_rows(end_month: str) -> pl.DataFrame:
    if not STANDARDIZED_ONERPM_PATH.exists():
        raise FileNotFoundError(f"No existe {STANDARDIZED_ONERPM_PATH}")

    lf = pl.scan_parquet(STANDARDIZED_ONERPM_PATH)
    lf = ensure_store_dimensions(lf, set(lf.collect_schema().names()))
    schema = lf.collect_schema()
    schema_names = set(schema.names())

    base = (
        lf
        .filter((pl.col("source") == SOURCE) & (pl.col("account") == ACCOUNT))
        .filter(col_or_null(schema, "statement_period") <= end_month)
        .with_columns([
            col_or_null(schema, "source_sheet").alias("_source_sheet"),
            col_or_null(schema, "statement_period").alias("_statement_period"),
            col_or_null(schema, "transaction_month").alias("_transaction_month"),
            first_text(schema, ["Track Title", "Title", "Video Title", "Album Title"]).alias("_title"),
            first_text(schema, ["artist_statement_style", "artists_raw", "Channel Name"]).alias("_artist"),
            first_text(schema, ["ISRC", "asset_isrc"]).alias("_isrc"),
            first_text(schema, ["UPC", "asset_upc"]).alias("_upc"),
            first_text(schema, ["Video ID", "ID", "Parent ID"]).alias("_video_id"),
            first_text(schema, ["store_report_label"]).alias("_dsp_store"),
            first_text(schema, ["monetization_normalized"]).alias("_monetization"),
            first_text(schema, ["content_origin_normalized"]).alias("_content_origin"),
            first_text(schema, ["plan_normalized"]).alias("_plan"),
            first_text(schema, ["Territory"]).alias("_territory"),
            amount_expr(schema).alias("_amount_usd"),
            col_or_null(schema, "Quantity", pl.Float64).alias("_units"),
            release_metadata_provider_expr(schema).alias("_metadata_provider"),
            release_metadata_lookup_key_expr(schema).alias("_metadata_lookup_key"),
        ])
    )
    base = apply_report_net_personalization(base, set(base.collect_schema().names()), amount_col="_amount_usd")

    with_catalog = with_catalog_report_status(base, schema_names)
    labels = catalog_label_lookup()
    joined = (
        with_catalog
        .join(
            metadata_ready(),
            on=["_metadata_provider", "_metadata_lookup_key"],
            how="left",
        )
    )
    if not labels.is_empty():
        joined = joined.join(labels.lazy(), on="catalog_key", how="left")
    else:
        joined = joined.with_columns(pl.lit(None).cast(pl.Utf8).alias("label_normalized_report"))
    joined = (
        joined
        .with_columns([
            pl.coalesce(["label_normalized_report", "external_label"]).alias("label_normalized_report"),
            (pl.col("include_in_reports") == True).alias("_catalog_ok"),
            (pl.col("_include_after_release_cutoff") == True).fill_null(False).alias("_cutoff_ok"),
        ])
        .with_columns([
            (
                (pl.col("_catalog_ok") == True)
                & (pl.col("_cutoff_ok") == True)
            ).fill_null(False).alias("_business_ok"),
            pl.when(pl.col("_catalog_ok") == False)
            .then(pl.lit("excluido_catalogo"))
            .when(pl.col("_include_after_release_cutoff") == False)
            .then(pl.lit("excluido_fecha_contractual"))
            .when(pl.col("_include_after_release_cutoff").is_null())
            .then(pl.lit("excluido_sin_metadata"))
            .otherwise(pl.lit("reportable"))
            .alias("_reason"),
        ])
    )
    return joined.collect()


def nonzero_summary(summary: pl.DataFrame) -> pl.DataFrame:
    if summary.is_empty() or "Ingresos USD" not in summary.columns:
        return summary
    return summary.filter(pl.col("Ingresos USD").abs() >= ZERO_EPSILON)


def summarize(rows: pl.DataFrame, category_filter: pl.Expr, *, hide_zero_amounts: bool = False) -> pd.DataFrame:
    selected = rows.filter(category_filter)
    if selected.is_empty():
        return pd.DataFrame()

    grouped = (
        selected
        .group_by([
            "_source_sheet",
            "_title",
            "_artist",
            "_isrc",
            "_upc",
            "_video_id",
            "release_date",
            "label_normalized_report",
            "metadata_status",
            "_reason",
            "_statement_period",
        ])
        .agg([
            pl.sum("_amount_usd").alias("amount_usd"),
            pl.sum("_units").alias("units"),
            pl.len().alias("rows"),
            pl.min("_transaction_month").alias("first_transaction_month"),
            pl.max("_transaction_month").alias("last_transaction_month"),
        ])
    )

    base_cols = [
        "_source_sheet",
        "_title",
        "_artist",
        "_isrc",
        "_upc",
        "_video_id",
        "release_date",
        "label_normalized_report",
        "metadata_status",
    ]
    totals = (
        grouped
        .group_by(base_cols)
        .agg([
            pl.sum("amount_usd").alias("Ingresos USD"),
            pl.sum("units").alias("Unidades"),
            pl.sum("rows").alias("Filas"),
            pl.min("first_transaction_month").alias("Desde transaction"),
            pl.max("last_transaction_month").alias("Hasta transaction"),
            pl.min("_statement_period").alias("Desde statement"),
            pl.max("_statement_period").alias("Hasta statement"),
        ])
    )
    if hide_zero_amounts:
        totals = nonzero_summary(totals)
        if totals.is_empty():
            return pd.DataFrame()

    out = totals
    rename = {
        "_source_sheet": "Hoja origen",
        "_title": "Tema / video",
        "_artist": "Artista statement",
        "_isrc": "ISRC",
        "_upc": "UPC",
        "_video_id": "Video ID",
        "release_date": "Release date",
        "label_normalized_report": "Label normalizado",
        "metadata_status": "Metadata",
    }
    out = out.rename(rename).sort("Ingresos USD", descending=True)
    return out.to_pandas()


def granular_onerpm_rows(rows: pl.DataFrame, *, hide_zero_amounts: bool = False) -> pd.DataFrame:
    selected = rows.filter(pl.col("_business_ok") == True)
    if hide_zero_amounts:
        parent_cols = [
            "_source_sheet",
            "_title",
            "_artist",
            "_isrc",
            "_upc",
            "_video_id",
            "release_date",
            "label_normalized_report",
            "metadata_status",
        ]
        selected = (
            selected
            .with_columns(pl.sum("_amount_usd").over(parent_cols).alias("_parent_amount_usd"))
            .filter(pl.col("_parent_amount_usd").abs() >= ZERO_EPSILON)
            .drop("_parent_amount_usd")
        )
    if selected.is_empty():
        return pd.DataFrame()

    grouped = (
        selected
        .group_by([
            "_source_sheet",
            "_title",
            "_artist",
            "_isrc",
            "_upc",
            "_video_id",
            "release_date",
            "label_normalized_report",
            "metadata_status",
            "_territory",
            "_dsp_store",
            "_monetization",
            "_content_origin",
            "_plan",
        ])
        .agg([
            pl.sum("_amount_usd").alias("Ingresos USD"),
            pl.sum("_units").alias("Unidades"),
            pl.len().alias("Filas"),
            pl.min("_transaction_month").alias("Desde transaction"),
            pl.max("_transaction_month").alias("Hasta transaction"),
            pl.min("_statement_period").alias("Desde statement"),
            pl.max("_statement_period").alias("Hasta statement"),
        ])
        .rename({
            "_source_sheet": "Hoja origen",
            "_title": "Tema / video",
            "_artist": "Artista statement",
            "_isrc": "ISRC",
            "_upc": "UPC",
            "_video_id": "Video ID",
            "release_date": "Release date",
            "label_normalized_report": "Label normalizado",
            "metadata_status": "Metadata",
            "_territory": "Pais",
            "_dsp_store": "DSP / Store",
            "_monetization": "Monetizacion",
            "_content_origin": "Origen contenido",
            "_plan": "Plan",
        })
        .sort("Ingresos USD", descending=True)
    )
    return grouped.to_pandas()


def report_rows(rows: pl.DataFrame, hide_zero_amounts: bool) -> pl.DataFrame:
    if not hide_zero_amounts:
        return rows
    return rows.filter(pl.col("_amount_usd").abs() >= ZERO_EPSILON)


def summary_table(rows: pl.DataFrame, *, hide_zero_amounts: bool = False) -> pd.DataFrame:
    visible_rows = report_rows(rows, hide_zero_amounts)
    reportable = visible_rows.filter(pl.col("_business_ok") == True)
    excluded_catalog = visible_rows.filter(pl.col("_catalog_ok") == False)
    excluded_own = visible_rows.filter((pl.col("_catalog_ok") == True) & (pl.col("_business_ok") == False))
    raw_total = visible_rows["_amount_usd"].sum() or 0.0

    items = [
        ("Ingresos ONErpm", reportable),
        ("Ingresos Excluidos Boxindanga", excluded_catalog),
        ("Ingresos Excluidos Propios", excluded_own),
        ("Total raw control", visible_rows),
    ]
    summary_rows = []
    for label, frame in items:
        summary_rows.append({
            "Bloque": label,
            "Ingresos USD": float(frame["_amount_usd"].sum() or 0.0),
            "Unidades": float(frame["_units"].sum() or 0.0),
            "Filas": int(frame.height),
            "% raw": (float(frame["_amount_usd"].sum() or 0.0) / raw_total) if raw_total else 0,
        })

    by_sheet = (
        visible_rows
        .with_columns(
            pl.when(pl.col("_business_ok") == True)
            .then(pl.lit("Ingresos ONErpm"))
            .when(pl.col("_catalog_ok") == False)
            .then(pl.lit("Ingresos Excluidos Boxindanga"))
            .otherwise(pl.lit("Ingresos Excluidos Propios"))
            .alias("Bloque")
        )
        .group_by(["Bloque", "_source_sheet"])
        .agg([
            pl.sum("_amount_usd").alias("Ingresos USD"),
            pl.sum("_units").alias("Unidades"),
            pl.len().alias("Filas"),
        ])
        .rename({"_source_sheet": "Hoja origen"})
        .sort(["Bloque", "Hoja origen"])
        .to_pandas()
    )
    return pd.DataFrame(summary_rows), by_sheet


def summary_from_outputs(
    reportable: pd.DataFrame,
    excluded_box: pd.DataFrame,
    excluded_own: pd.DataFrame,
    fuga: pd.DataFrame | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    blocks = [
        ("Ingresos ONErpm", reportable),
        ("Ingresos Excluidos Boxindanga", excluded_box),
        ("Ingresos Excluidos Propios", excluded_own),
    ]
    if fuga is not None:
        blocks.append(("Ingresos FUGA", fuga))
    raw_total = sum(float(df["Ingresos USD"].sum()) for _, df in blocks if "Ingresos USD" in df.columns)
    summary_rows = []
    by_sheet_frames = []

    for label, df in blocks:
        amount = float(df["Ingresos USD"].sum()) if "Ingresos USD" in df.columns else 0.0
        units = float(df["Unidades"].sum()) if "Unidades" in df.columns else 0.0
        rows_count = int(df["Filas"].sum()) if "Filas" in df.columns else 0
        summary_rows.append({
            "Bloque": label,
            "Ingresos USD": amount,
            "Unidades": units,
            "Filas": rows_count,
            "% raw": (amount / raw_total) if raw_total else 0,
        })

        if not df.empty and "Hoja origen" in df.columns:
            by_sheet = (
                df
                .groupby("Hoja origen", dropna=False, as_index=False)
                .agg({
                    "Ingresos USD": "sum",
                    "Unidades": "sum",
                    "Filas": "sum",
                })
            )
            by_sheet.insert(0, "Bloque", label)
            by_sheet_frames.append(by_sheet)

    summary_rows.append({
        "Bloque": "Total raw control",
        "Ingresos USD": raw_total,
        "Unidades": sum(row["Unidades"] for row in summary_rows),
        "Filas": sum(row["Filas"] for row in summary_rows),
        "% raw": 1 if raw_total else 0,
    })
    summary = pd.DataFrame(summary_rows)
    by_sheet_out = pd.concat(by_sheet_frames, ignore_index=True) if by_sheet_frames else pd.DataFrame()
    return summary, by_sheet_out


def fuga_candidate_rows(end_month: str, *, hide_zero_amounts: bool = False) -> pd.DataFrame:
    if not STANDARDIZED_FUGA_PATH.exists():
        return pd.DataFrame()

    lf = pl.scan_parquet(STANDARDIZED_FUGA_PATH)
    lf = ensure_store_dimensions(lf, set(lf.collect_schema().names()))
    schema = lf.collect_schema()
    schema_names = set(schema.names())
    string_cols = [
        col
        for col, dtype in zip(schema.names(), schema.dtypes())
        if dtype == pl.String
    ]
    search_expr = pl.any_horizontal([
        pl.col(col)
        .cast(pl.Utf8, strict=False)
        .str.to_lowercase()
        .str.contains("dj plaga|la nueva sangre")
        .fill_null(False)
        for col in string_cols
    ])

    base = (
        lf
        .filter(col_or_null(schema, "statement_period") <= end_month)
        .filter(search_expr)
        .with_columns([
            pl.lit("FUGA").alias("Hoja origen"),
            first_text(schema, ["track_statement_style", "Asset Title", "Product Title"]).alias("Tema / video"),
            first_text(schema, ["artist_statement_style", "Asset Artist", "Product Artist"]).alias("Artista statement"),
            first_text(schema, ["asset_isrc", "Asset ISRC"]).alias("_isrc"),
            first_text(schema, ["product_upc", "Product UPC"]).alias("UPC"),
            pl.lit(None).cast(pl.Utf8).alias("Video ID"),
            amount_expr(schema).alias("_amount_usd"),
            pl.coalesce([
                col_or_null(schema, "asset_quantity_num", pl.Float64),
                col_or_null(schema, "product_quantity_num", pl.Float64),
            ]).alias("_units"),
            col_or_null(schema, "transaction_month").alias("_transaction_month"),
            col_or_null(schema, "statement_period").alias("_statement_period"),
        ])
    )
    base = apply_report_net_personalization(base, set(base.collect_schema().names()), amount_col="_amount_usd")
    enriched = with_catalog_report_status(base, schema_names)
    if hide_zero_amounts:
        parent_cols = [
            "Hoja origen",
            "Tema / video",
            "Artista statement",
            "catalog_key",
            "UPC",
            "Video ID",
        ]
        enriched = (
            enriched
            .with_columns(pl.sum("_amount_usd").over(parent_cols).alias("_parent_amount_usd"))
            .filter(pl.col("_parent_amount_usd").abs() >= ZERO_EPSILON)
            .drop("_parent_amount_usd")
        )

    grouped = (
        enriched
        .group_by([
            "Hoja origen",
            "Tema / video",
            "Artista statement",
            "catalog_key",
            "UPC",
            "Video ID",
        ])
        .agg([
            pl.col("_isrc").drop_nulls().unique().sort().str.join(" | ").alias("ISRC original"),
            pl.col("_isrc").is_null().any().alias("_had_blank_isrc"),
            pl.sum("_amount_usd").alias("Ingresos USD"),
            pl.sum("_units").alias("Unidades"),
            pl.len().alias("Filas"),
            pl.min("_transaction_month").alias("Desde transaction"),
            pl.max("_transaction_month").alias("Hasta transaction"),
            pl.min("_statement_period").alias("Desde statement"),
            pl.max("_statement_period").alias("Hasta statement"),
        ])
        .collect()
    )
    if grouped.is_empty():
        return pd.DataFrame()

    metadata = catalog_metadata_by_key()
    onerpm_isrcs = onerpm_lns_isrcs()
    out = (
        grouped
        .join(metadata, on="catalog_key", how="left")
        .with_columns([
            pl.coalesce(["_resolved_isrc", "ISRC original"]).alias("ISRC"),
            pl.when(pl.col("_resolved_isrc").is_not_null() & pl.col("_had_blank_isrc") & pl.col("UPC").is_not_null())
            .then(pl.lit("upc_unico"))
            .when(pl.col("_resolved_isrc").is_not_null())
            .then(pl.lit("isrc"))
            .otherwise(pl.lit("pendiente"))
            .alias("Resolucion"),
            pl.when(pl.col("_had_blank_isrc") & pl.col("ISRC original").is_not_null() & (pl.col("ISRC original") != ""))
            .then(pl.concat_str([pl.col("ISRC original"), pl.lit(" | (vacio)")]))
            .when(pl.col("_had_blank_isrc"))
            .then(pl.lit("(vacio)"))
            .otherwise(pl.col("ISRC original"))
            .alias("ISRC original"),
        ])
        .join(onerpm_isrcs, left_on="ISRC", right_on="_isrc", how="left")
        .with_columns([
            pl.when(pl.col("_onerpm_usd").is_not_null())
            .then(pl.lit("onerpm"))
            .when(pl.col("_release_year_month").is_not_null() & (pl.col("_release_year_month") >= "2023-06"))
            .then(pl.lit("nuevo"))
            .when(pl.col("_release_year_month").is_not_null() & (pl.col("_release_year_month") < "2023-06"))
            .then(pl.lit("viejo"))
            .otherwise(pl.lit("revisar"))
            .alias("Origen"),
        ])
        .select([
            "Hoja origen",
            "Tema / video",
            "Artista statement",
            "ISRC",
            "ISRC original",
            "Resolucion",
            "UPC",
            "Video ID",
            "Release date",
            "Label normalizado",
            "Metadata",
            "Origen",
            "Ingresos USD",
            "Unidades",
            "Filas",
            "Desde transaction",
            "Hasta transaction",
            "Desde statement",
            "Hasta statement",
        ])
        .sort("Ingresos USD", descending=True)
    )
    return out.to_pandas()


def granular_fuga_rows(end_month: str, *, hide_zero_amounts: bool = False) -> pd.DataFrame:
    if not STANDARDIZED_FUGA_PATH.exists():
        return pd.DataFrame()

    lf = pl.scan_parquet(STANDARDIZED_FUGA_PATH)
    lf = ensure_store_dimensions(lf, set(lf.collect_schema().names()))
    schema = lf.collect_schema()
    schema_names = set(schema.names())
    string_cols = [
        col
        for col, dtype in zip(schema.names(), schema.dtypes())
        if dtype == pl.String
    ]
    search_expr = pl.any_horizontal([
        pl.col(col)
        .cast(pl.Utf8, strict=False)
        .str.to_lowercase()
        .str.contains("dj plaga|la nueva sangre")
        .fill_null(False)
        for col in string_cols
    ])

    base = (
        lf
        .filter(col_or_null(schema, "statement_period") <= end_month)
        .filter(search_expr)
        .with_columns([
            pl.lit("FUGA").alias("Hoja origen"),
            first_text(schema, ["track_statement_style", "Asset Title", "Product Title"]).alias("Tema / video"),
            first_text(schema, ["artist_statement_style", "Asset Artist", "Product Artist"]).alias("Artista statement"),
            first_text(schema, ["asset_isrc", "Asset ISRC"]).alias("_isrc"),
            first_text(schema, ["product_upc", "Product UPC"]).alias("UPC"),
            pl.lit(None).cast(pl.Utf8).alias("Video ID"),
            first_text(schema, ["store_report_label"]).alias("DSP / Store"),
            first_text(schema, ["monetization_normalized"]).alias("Monetizacion"),
            first_text(schema, ["content_origin_normalized"]).alias("Origen contenido"),
            first_text(schema, ["plan_normalized"]).alias("Plan"),
            first_text(schema, ["territory", "Territory"]).alias("Pais"),
            amount_expr(schema).alias("_amount_usd"),
            pl.coalesce([
                col_or_null(schema, "asset_quantity_num", pl.Float64),
                col_or_null(schema, "product_quantity_num", pl.Float64),
            ]).alias("_units"),
            col_or_null(schema, "transaction_month").alias("_transaction_month"),
            col_or_null(schema, "statement_period").alias("_statement_period"),
        ])
    )
    base = apply_report_net_personalization(base, set(base.collect_schema().names()), amount_col="_amount_usd")
    grouped = (
        with_catalog_report_status(base, schema_names)
        .group_by([
            "Hoja origen",
            "Tema / video",
            "Artista statement",
            "catalog_key",
            "UPC",
            "Video ID",
            "Pais",
            "DSP / Store",
            "Monetizacion",
            "Origen contenido",
            "Plan",
        ])
        .agg([
            pl.col("_isrc").drop_nulls().unique().sort().str.join(" | ").alias("ISRC original"),
            pl.col("_isrc").is_null().any().alias("_had_blank_isrc"),
            pl.sum("_amount_usd").alias("Ingresos USD"),
            pl.sum("_units").alias("Unidades"),
            pl.len().alias("Filas"),
            pl.min("_transaction_month").alias("Desde transaction"),
            pl.max("_transaction_month").alias("Hasta transaction"),
            pl.min("_statement_period").alias("Desde statement"),
            pl.max("_statement_period").alias("Hasta statement"),
        ])
        .collect()
    )
    if hide_zero_amounts and not grouped.is_empty():
        grouped = grouped.filter(pl.col("Ingresos USD").abs() >= ZERO_EPSILON)
    if grouped.is_empty():
        return pd.DataFrame()

    metadata = catalog_metadata_by_key()
    onerpm_isrcs = onerpm_lns_isrcs()
    out = (
        grouped
        .join(metadata, on="catalog_key", how="left")
        .with_columns([
            pl.coalesce(["_resolved_isrc", "ISRC original"]).alias("ISRC"),
            pl.when(pl.col("_resolved_isrc").is_not_null() & pl.col("_had_blank_isrc") & pl.col("UPC").is_not_null())
            .then(pl.lit("upc_unico"))
            .when(pl.col("_resolved_isrc").is_not_null())
            .then(pl.lit("isrc"))
            .otherwise(pl.lit("pendiente"))
            .alias("Resolucion"),
            pl.when(pl.col("_had_blank_isrc") & pl.col("ISRC original").is_not_null() & (pl.col("ISRC original") != ""))
            .then(pl.concat_str([pl.col("ISRC original"), pl.lit(" | (vacio)")]))
            .when(pl.col("_had_blank_isrc"))
            .then(pl.lit("(vacio)"))
            .otherwise(pl.col("ISRC original"))
            .alias("ISRC original"),
        ])
        .join(onerpm_isrcs, left_on="ISRC", right_on="_isrc", how="left")
        .with_columns([
            pl.when(pl.col("_onerpm_usd").is_not_null())
            .then(pl.lit("onerpm"))
            .when(pl.col("_release_year_month").is_not_null() & (pl.col("_release_year_month") >= "2023-06"))
            .then(pl.lit("nuevo"))
            .when(pl.col("_release_year_month").is_not_null() & (pl.col("_release_year_month") < "2023-06"))
            .then(pl.lit("viejo"))
            .otherwise(pl.lit("revisar"))
            .alias("Origen"),
        ])
        .select([
            "Hoja origen",
            "Tema / video",
            "Artista statement",
            "ISRC",
            "ISRC original",
            "Resolucion",
            "UPC",
            "Video ID",
            "Release date",
            "Label normalizado",
            "Metadata",
            "Origen",
            "Pais",
            "DSP / Store",
            "Monetizacion",
            "Origen contenido",
            "Plan",
            "Ingresos USD",
            "Unidades",
            "Filas",
            "Desde transaction",
            "Hasta transaction",
            "Desde statement",
            "Hasta statement",
        ])
        .sort("Ingresos USD", descending=True)
    )
    return out.to_pandas()


def add_youtube_delay_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    def month_diff(start: object, end: object) -> float | None:
        if pd.isna(start) or pd.isna(end):
            return None
        start_text = str(start)[:7]
        end_text = str(end)[:7]
        if not re.match(r"^\d{4}-\d{2}$", start_text) or not re.match(r"^\d{4}-\d{2}$", end_text):
            return None
        start_year, start_month = map(int, start_text.split("-"))
        end_year, end_month = map(int, end_text.split("-"))
        return float((end_year - start_year) * 12 + (end_month - start_month))

    release_month = df["Release date"].fillna("").astype(str).str.slice(0, 7)
    df.insert(
        df.columns.get_loc("Primer mes monetizado") + 1,
        "Meses demora",
        [
            month_diff(start, end)
            for start, end in zip(release_month, df["Primer mes monetizado"], strict=False)
        ],
    )
    return df


def youtube_summary_rows(detail: pd.DataFrame) -> pd.DataFrame:
    if detail.empty:
        return detail

    def clean_text(value: object) -> str:
        if pd.isna(value):
            return ""
        return str(value).strip()

    def first_non_empty(values: pd.Series) -> object:
        for value in values:
            text = clean_text(value)
            if text:
                return value
        return None

    def unique_join(values: pd.Series) -> str:
        seen = []
        for value in values:
            text = clean_text(value)
            if text and text not in seen:
                seen.append(text)
        return ", ".join(seen)

    work = detail.copy()
    work["_group_type"] = work.apply(
        lambda row: "video_id"
        if clean_text(row.get("Video ID"))
        else ("isrc_titulo" if clean_text(row.get("ISRC")) else "titulo"),
        axis=1,
    )
    work["_group_key"] = work.apply(
        lambda row: "|".join([
            clean_text(row.get("Fuente")),
            clean_text(row.get("Video ID"))
            if clean_text(row.get("Video ID"))
            else (
                clean_text(row.get("ISRC")) + "|" + normalize_match_text(row.get("Tema / video"))
                if clean_text(row.get("ISRC"))
                else normalize_match_text(row.get("Tema / video"))
            ),
        ]),
        axis=1,
    )

    grouped = (
        work
        .groupby(["Fuente", "_group_type", "_group_key"], dropna=False)
        .agg({
            "DSP / Store": unique_join,
            "Tema / video": first_non_empty,
            "Artista statement": first_non_empty,
            "Video ID": first_non_empty,
            "ISRC": first_non_empty,
            "UPC": first_non_empty,
            "Release date": first_non_empty,
            "Label normalizado": first_non_empty,
            "Metadata": unique_join,
            "URL metadata": first_non_empty,
            "Primer mes monetizado": "min",
            "Primer statement": "min",
            "Ultimo mes monetizado": "max",
            "Ultimo statement": "max",
            "Ingresos USD": "sum",
            "Unidades": "sum",
            "Filas": "sum",
            "Estado negocio": unique_join,
        })
        .reset_index()
        .rename(columns={
            "_group_type": "Agrupado por",
            "_group_key": "Clave analisis",
            "Filas": "Filas detalle",
        })
    )
    grouped = add_youtube_delay_columns(grouped)
    return grouped.sort_values(
        by=["Meses demora", "Ingresos USD"],
        ascending=[False, False],
        na_position="last",
    )


def youtube_analysis_rows(
    rows: pl.DataFrame,
    end_month: str,
    *,
    hide_zero_amounts: bool = False,
) -> pd.DataFrame:
    blocks: list[pd.DataFrame] = []

    onerpm_video = rows.filter(pl.col("_source_sheet") == "Youtube Channels")
    if hide_zero_amounts:
        onerpm_video = report_rows(onerpm_video, hide_zero_amounts)
    if not onerpm_video.is_empty():
        one = (
            onerpm_video
            .group_by([
                "_title",
                "_artist",
                "_video_id",
                "_isrc",
                "_upc",
                "release_date",
                "label_normalized_report",
                "metadata_status",
                "match_url",
                "_business_ok",
                "_reason",
            ])
            .agg([
                pl.sum("_amount_usd").alias("Ingresos USD"),
                pl.sum("_units").alias("Unidades"),
                pl.len().alias("Filas"),
                pl.min("_transaction_month").alias("Primer mes monetizado"),
                pl.min("_statement_period").alias("Primer statement"),
                pl.max("_transaction_month").alias("Ultimo mes monetizado"),
                pl.max("_statement_period").alias("Ultimo statement"),
            ])
            .with_columns([
                pl.lit("ONErpm").alias("Fuente"),
                pl.lit("Youtube Channels").alias("DSP / Store"),
                pl.when(pl.col("_business_ok"))
                .then(pl.lit("incluido"))
                .otherwise(pl.col("_reason"))
                .alias("Estado negocio"),
            ])
            .rename({
                "_title": "Tema / video",
                "_artist": "Artista statement",
                "_video_id": "Video ID",
                "_isrc": "ISRC",
                "_upc": "UPC",
                "release_date": "Release date",
                "label_normalized_report": "Label normalizado",
                "metadata_status": "Metadata",
                "match_url": "URL metadata",
            })
            .select([
                "Fuente",
                "DSP / Store",
                "Tema / video",
                "Artista statement",
                "Video ID",
                "ISRC",
                "UPC",
                "Release date",
                "Label normalizado",
                "Metadata",
                "URL metadata",
                "Primer mes monetizado",
                "Primer statement",
                "Ultimo mes monetizado",
                "Ultimo statement",
                "Ingresos USD",
                "Unidades",
                "Filas",
                "Estado negocio",
            ])
            .to_pandas()
        )
        blocks.append(one)

    if STANDARDIZED_FUGA_PATH.exists():
        lf = pl.scan_parquet(STANDARDIZED_FUGA_PATH)
        schema = lf.collect_schema()
        string_cols = [
            col
            for col, dtype in zip(schema.names(), schema.dtypes())
            if dtype == pl.String
        ]
        search_expr = pl.any_horizontal([
            pl.col(col)
            .cast(pl.Utf8, strict=False)
            .str.to_lowercase()
            .str.contains("dj plaga|la nueva sangre")
            .fill_null(False)
            for col in string_cols
        ])
        youtube_expr = (
            col_or_null(schema, "dsp").str.to_lowercase().str.contains("youtube").fill_null(False)
            | col_or_null(schema, "DSP").str.to_lowercase().str.contains("youtube").fill_null(False)
            | col_or_null(schema, "store_name").str.to_lowercase().str.contains("youtube").fill_null(False)
            | col_or_null(schema, "Sale Store Name").str.to_lowercase().str.contains("youtube").fill_null(False)
            | col_or_null(schema, "sale_type").str.to_lowercase().str.contains("video").fill_null(False)
            | col_or_null(schema, "Sale Type").str.to_lowercase().str.contains("video").fill_null(False)
        )
        fuga = (
            lf
            .filter(col_or_null(schema, "statement_period") <= end_month)
            .filter(search_expr & youtube_expr)
            .with_columns([
                first_text(schema, ["track_statement_style", "Asset Title", "Product Title"]).alias("Tema / video"),
                first_text(schema, ["artist_statement_style", "Asset Artist", "Product Artist"]).alias("Artista statement"),
                first_text(schema, ["DSP Container ID", "Video ID", "DSP Unit ID"]).alias("_video_id"),
                first_text(schema, ["asset_isrc", "Asset ISRC"]).alias("_isrc"),
                first_text(schema, ["product_upc", "Product UPC"]).alias("UPC"),
                amount_expr(schema).alias("_amount_usd"),
                col_or_null(schema, "asset_quantity_num", pl.Float64).alias("_units"),
                col_or_null(schema, "transaction_month").alias("_transaction_month"),
                col_or_null(schema, "statement_period").alias("_statement_period"),
                pl.concat_str([
                    col_or_null(schema, "DSP"),
                    pl.lit(" / "),
                    col_or_null(schema, "Sale Store Name"),
                    pl.lit(" / "),
                    col_or_null(schema, "Sale Type"),
                ], ignore_nulls=True).alias("DSP / Store"),
            ])
            .pipe(lambda frame: apply_report_net_personalization(frame, set(frame.collect_schema().names()), amount_col="_amount_usd"))
            .group_by([
                "Tema / video",
                "Artista statement",
                "_video_id",
                "_isrc",
                "UPC",
                "DSP / Store",
            ])
            .agg([
                pl.sum("_amount_usd").alias("Ingresos USD"),
                pl.sum("_units").alias("Unidades"),
                pl.len().alias("Filas"),
                pl.min("_transaction_month").alias("Primer mes monetizado"),
                pl.min("_statement_period").alias("Primer statement"),
                pl.max("_transaction_month").alias("Ultimo mes monetizado"),
                pl.max("_statement_period").alias("Ultimo statement"),
            ])
            .collect()
        )
        if hide_zero_amounts and not fuga.is_empty():
            fuga = fuga.filter(pl.col("Ingresos USD").abs() >= ZERO_EPSILON)
        if not fuga.is_empty():
            by_video = metadata_by_video()
            by_isrc = metadata_by_isrc().rename({
                "Release date": "ISRC release date",
                "Label normalizado": "ISRC label normalizado",
                "Metadata": "ISRC metadata",
                "_release_year_month": "_isrc_release_year_month",
            })
            fuga = (
                fuga
                .join(by_video, on="_video_id", how="left")
                .join(by_isrc, on="_isrc", how="left")
                .with_columns([
                    pl.coalesce(["Video release date", "ISRC release date"]).alias("Release date"),
                    pl.coalesce(["Video label normalizado", "ISRC label normalizado"]).alias("Label normalizado"),
                    pl.coalesce(["Video metadata", "ISRC metadata"]).alias("Metadata"),
                    pl.lit(None).cast(pl.Utf8).alias("Estado negocio"),
                    pl.lit("FUGA").alias("Fuente"),
                ])
                .rename({
                    "_video_id": "Video ID",
                    "_isrc": "ISRC",
                    "Video match url": "URL metadata",
                })
                .select([
                    "Fuente",
                    "DSP / Store",
                    "Tema / video",
                    "Artista statement",
                    "Video ID",
                    "ISRC",
                    "UPC",
                    "Release date",
                    "Label normalizado",
                    "Metadata",
                    "URL metadata",
                    "Primer mes monetizado",
                    "Primer statement",
                    "Ultimo mes monetizado",
                    "Ultimo statement",
                    "Ingresos USD",
                    "Unidades",
                    "Filas",
                    "Estado negocio",
                ])
                .to_pandas()
            )
            blocks.append(fuga)

    if not blocks:
        return pd.DataFrame()

    out = pd.concat(blocks, ignore_index=True)
    out = add_youtube_delay_columns(out)
    out = out.sort_values(
        by=["Meses demora", "Ingresos USD"],
        ascending=[False, False],
        na_position="last",
    )
    return out


def style_sheet(ws, *, fill: PatternFill | None = None) -> None:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    if fill:
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 500), max_col=ws.max_column):
            for cell in row:
                cell.fill = fill
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        max_len = 0
        for cell in column[:400]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 11), 48)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = LEFT
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                header = str(ws.cell(row=1, column=cell.column).value or "")
                if header in {"Filas", "Filas detalle", "Unidades"}:
                    cell.number_format = '#,##0'
                elif "USD" in header:
                    cell.number_format = '$#,##0.00'
                elif "%" in header:
                    cell.number_format = '0.00%'
                else:
                    cell.number_format = '#,##0.00'


def style_header_row(ws, row_number: int) -> None:
    for cell in ws[row_number]:
        if cell.value is None:
            continue
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def format_numeric_columns(ws, header_row: int, data_start_row: int) -> None:
    for row in ws.iter_rows(min_row=data_start_row):
        for cell in row:
            if not isinstance(cell.value, (int, float)):
                continue
            header = str(ws.cell(row=header_row, column=cell.column).value or "")
            if header in {"Filas", "Filas detalle", "Unidades"}:
                cell.number_format = '#,##0'
            elif "USD" in header:
                cell.number_format = '$#,##0.00'
            elif "%" in header:
                cell.number_format = '0.00%'
            else:
                cell.number_format = '#,##0.00'


def write_report(rows: pl.DataFrame, end_month: str, *, hide_zero_amounts: bool = False) -> Path:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = REPORTS_DIR / f"la_nueva_sangre_inicio_a_{end_month}_{stamp}.xlsx"

    reportable = summarize(rows, pl.col("_business_ok") == True, hide_zero_amounts=hide_zero_amounts)
    excluded_box = summarize(rows, pl.col("_catalog_ok") == False, hide_zero_amounts=hide_zero_amounts)
    excluded_own = summarize(
        rows,
        (pl.col("_catalog_ok") == True) & (pl.col("_business_ok") == False),
        hide_zero_amounts=hide_zero_amounts,
    )
    fuga = fuga_candidate_rows(end_month, hide_zero_amounts=hide_zero_amounts)
    reportable_granular = filter_detail_to_parent(
        granular_onerpm_rows(rows, hide_zero_amounts=False),
        reportable,
    )
    fuga_granular = filter_detail_to_parent(
        granular_fuga_rows(end_month, hide_zero_amounts=False),
        fuga,
    )
    youtube_analysis = youtube_analysis_rows(rows, end_month, hide_zero_amounts=hide_zero_amounts)
    youtube_summary = youtube_summary_rows(youtube_analysis)
    summary, by_sheet = summary_from_outputs(reportable, excluded_box, excluded_own, fuga)

    reportable_out = presentation_columns(reportable)
    reportable_granular_out = presentation_columns(reportable_granular)
    fuga_out = presentation_columns(fuga)
    fuga_granular_out = presentation_columns(fuga_granular)
    combined_income_out = combined_income_sheet(reportable_out, fuga_out)
    youtube_summary_out = presentation_columns(youtube_summary)
    youtube_analysis_out = presentation_columns(youtube_analysis)
    excluded_box_out = presentation_columns(excluded_box)
    excluded_own_out = presentation_columns(excluded_own)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="Resumen", startrow=1)
        detail_title_startrow = len(summary) + 4
        if not by_sheet.empty:
            by_sheet.to_excel(writer, index=False, sheet_name="Resumen", startrow=detail_title_startrow + 1)
        reportable_out.to_excel(writer, index=False, sheet_name="Ingresos OneRPM")
        fuga_out.to_excel(writer, index=False, sheet_name="Ingresos FUGA")
        combined_income_out.to_excel(writer, index=False, sheet_name="Ingresos OneRPM + FUGA")
        reportable_granular_out.to_excel(writer, index=False, sheet_name="OneRPM Pais DSP")
        fuga_granular_out.to_excel(writer, index=False, sheet_name="FUGA Pais DSP")
        youtube_summary_out.to_excel(writer, index=False, sheet_name="YouTube Resumen")
        youtube_analysis_out.to_excel(writer, index=False, sheet_name="Analisis YouTube")
        excluded_box_out.to_excel(writer, index=False, sheet_name="Excluidos Boxindanga")
        excluded_own_out.to_excel(writer, index=False, sheet_name="Excluidos Propios")

        ws = writer.book["Resumen"]
        ws["A1"] = f"La Nueva Sangre - inicio a {end_month}"
        ws["A1"].font = TITLE_FONT
        ws["A1"].fill = TOTAL_FILL
        ws["A1"].alignment = LEFT
        if not by_sheet.empty:
            detail_title_cell = ws.cell(row=detail_title_startrow + 1, column=1)
            detail_title_cell.value = "Detalle por bloque / hoja origen"
            detail_title_cell.font = TITLE_FONT
            detail_title_cell.fill = TOTAL_FILL
            detail_title_cell.alignment = LEFT

        style_sheet(writer.book["Resumen"], fill=TOTAL_FILL)
        style_header_row(writer.book["Resumen"], 2)
        format_numeric_columns(writer.book["Resumen"], 2, 3)
        if not by_sheet.empty:
            style_header_row(writer.book["Resumen"], detail_title_startrow + 2)
            format_numeric_columns(
                writer.book["Resumen"],
                detail_title_startrow + 2,
                detail_title_startrow + 3,
            )
        style_sheet(writer.book["Ingresos OneRPM"], fill=OK_FILL)
        style_sheet(writer.book["OneRPM Pais DSP"], fill=OK_FILL)
        style_sheet(writer.book["Ingresos FUGA"], fill=FUGA_FILL)
        style_sheet(writer.book["Ingresos OneRPM + FUGA"], fill=TOTAL_FILL)
        style_sheet(writer.book["FUGA Pais DSP"], fill=FUGA_FILL)
        style_sheet(writer.book["YouTube Resumen"], fill=FUGA_FILL)
        style_sheet(writer.book["Analisis YouTube"], fill=FUGA_FILL)
        style_sheet(writer.book["Excluidos Boxindanga"], fill=BOX_FILL)
        style_sheet(writer.book["Excluidos Propios"], fill=PROPIO_FILL)

    return output


def main() -> None:
    parser = argparse.ArgumentParser(description="Reporte personalizado La Nueva Sangre.")
    parser.add_argument("--end-month", default="2026-03")
    parser.add_argument("--hide-zero-amounts", action="store_true")
    parser.add_argument("--exclude-related-videos", action="store_true")
    args = parser.parse_args()

    rows = classified_rows(args.end_month)
    if args.exclude_related_videos:
        rows = apply_related_video_exclusions(rows)
    output = write_report(rows, args.end_month, hide_zero_amounts=args.hide_zero_amounts)
    print("Reporte La Nueva Sangre generado")
    print(output)


if __name__ == "__main__":
    main()
