from __future__ import annotations

import argparse
import re
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
import polars as pl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from lib.catalog_report_filter import current_catalog_status_path, with_catalog_report_status
    from lib.distributor_policy_store import load_distributor_policy_document
    from lib.store_taxonomy import ensure_store_dimensions
except ModuleNotFoundError:
    from scripts.lib.catalog_report_filter import current_catalog_status_path, with_catalog_report_status
    from scripts.lib.distributor_policy_store import load_distributor_policy_document
    from scripts.lib.store_taxonomy import ensure_store_dimensions


BASE = Path(r"C:\royalties_pipeline")
MARTS_DIR = BASE / "warehouse" / "marts"
REGISTRY_DIR = BASE / "warehouse" / "registry"
REPORTS_DIR = BASE / "reports" / "api"

SONG_LEVEL_PATH = MARTS_DIR / "song_level_all_sources.parquet"
RAW_ALL_PATH = MARTS_DIR / "standardized_raw_all_sources.parquet"
CATALOG_MASTER_PATH = MARTS_DIR / "catalog_master.parquet"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOTAL_FILL = PatternFill("solid", fgColor="D9EAF7")
OK_FILL = PatternFill("solid", fgColor="E2F0D9")
FUGA_FILL = PatternFill("solid", fgColor="EAF4FF")
EXCLUDED_FILL = PatternFill("solid", fgColor="FFF2CC")
TITLE_FONT = Font(size=14, bold=True, color="1F4E78")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
ZERO_EPSILON = 0.005

LAALO_PATTERN = r"juli\\s*savioli|julieta\\s*savioli|savioli"


def has_col(schema: dict[str, pl.DataType], col: str) -> bool:
    return col in schema


def col_or_null(schema: dict[str, pl.DataType], col: str, dtype=pl.Utf8) -> pl.Expr:
    if has_col(schema, col):
        return pl.col(col).cast(dtype, strict=False)
    return pl.lit(None).cast(dtype)


def non_blank(expr: pl.Expr) -> pl.Expr:
    text = expr.cast(pl.Utf8, strict=False).str.strip_chars()
    return pl.when(text.is_not_null() & (text != "")).then(text).otherwise(None)


def first_text(schema: dict[str, pl.DataType], columns: list[str]) -> pl.Expr:
    exprs = [non_blank(pl.col(col)) for col in columns if has_col(schema, col)]
    if not exprs:
        return pl.lit(None).cast(pl.Utf8)
    return pl.coalesce(exprs)


def amount_expr(schema: dict[str, pl.DataType]) -> pl.Expr:
    candidates = [
        pl.col(col).cast(pl.Float64, strict=False)
        for col in ["amount_usd", "net_amount_usd", "net_amount"]
        if has_col(schema, col)
    ]
    return pl.coalesce(candidates) if candidates else pl.lit(0.0)


def units_expr(schema: dict[str, pl.DataType]) -> pl.Expr:
    candidates = [
        pl.col(col).cast(pl.Float64, strict=False)
        for col in [
            "Quantity",
            "units",
            "Units",
            "streams",
            "Product Quantity",
            "Asset Quantity",
            "product_quantity_num",
            "asset_quantity_num",
        ]
        if has_col(schema, col)
    ]
    return pl.coalesce(candidates) if candidates else pl.lit(0.0)


def source_sheet_policy_expr(schema: dict[str, pl.DataType]) -> pl.Expr:
    source = col_or_null(schema, "source")
    statement_type = col_or_null(schema, "statement_type")
    source_sheet = col_or_null(schema, "source_sheet")
    return (
        pl.when((source == "fuga") & (statement_type == "correction"))
        .then(pl.lit("correction_csv"))
        .when((source == "fuga") & (statement_type == "regular"))
        .then(pl.lit("standard_statement_csv"))
        .when((source == "orchard") & (statement_type == "altafonte_legacy"))
        .then(pl.lit("legacy_altafonte"))
        .when((source == "orchard") & (statement_type == "orchard_statement"))
        .then(pl.lit("revenue_detail"))
        .otherwise(source_sheet)
    )


def load_statement_policy() -> pl.DataFrame:
    rows: list[dict[str, object]] = []
    config = load_distributor_policy_document()
    for entry in config.get("entries", []):
        source = entry.get("source")
        account = entry.get("account")
        display_name = entry.get("display_name") or f"{source} / {account}"
        for sheet, rule in (entry.get("sheet_rules") or {}).items():
            statement_view = rule.get("statement_view")
            rows.append({
                "source": source,
                "account": account,
                "_policy_sheet": sheet,
                "_policy_display": display_name,
                "_policy_statement_ok": statement_view is not False,
                "_policy_revenue_basis": rule.get("revenue_basis"),
            })
    return pl.DataFrame(rows)


def catalog_label_overrides() -> pl.DataFrame:
    status_path = current_catalog_status_path()
    if not status_path.exists():
        return pl.DataFrame({
            "catalog_key": pl.Series([], dtype=pl.Utf8),
            "_label_normalized_override": pl.Series([], dtype=pl.Utf8),
        })
    status = pl.read_parquet(status_path)
    if status.is_empty() or "catalog_key" not in status.columns or "label_normalized_override" not in status.columns:
        return pl.DataFrame({
            "catalog_key": pl.Series([], dtype=pl.Utf8),
            "_label_normalized_override": pl.Series([], dtype=pl.Utf8),
        })
    override = pl.col("label_normalized_override").cast(pl.Utf8, strict=False).str.strip_chars()
    return (
        status
        .select([
            pl.col("catalog_key").cast(pl.Utf8, strict=False),
            pl.when(override == "")
            .then(pl.lit(None).cast(pl.Utf8))
            .otherwise(override)
            .alias("_label_normalized_override"),
        ])
        .filter(pl.col("catalog_key").is_not_null())
        .unique(["catalog_key"], keep="last")
    )


def catalog_metadata_lookup() -> pl.DataFrame:
    empty = pl.DataFrame({
        "catalog_key": pl.Series([], dtype=pl.Utf8),
        "Release date": pl.Series([], dtype=pl.Utf8),
        "Label normalizado": pl.Series([], dtype=pl.Utf8),
        "Metadata": pl.Series([], dtype=pl.Utf8),
        "URL metadata": pl.Series([], dtype=pl.Utf8),
    })
    if not CATALOG_MASTER_PATH.exists():
        return empty
    catalog = pl.read_parquet(CATALOG_MASTER_PATH)
    if catalog.is_empty() or "catalog_key" not in catalog.columns:
        return empty

    columns = set(catalog.columns)
    label_candidates = [
        pl.col(col).cast(pl.Utf8, strict=False).str.strip_chars()
        for col in ["label_normalized", "label_normalized_auto", "external_label"]
        if col in columns
    ]
    label_expr = pl.coalesce(label_candidates) if label_candidates else pl.lit(None).cast(pl.Utf8)
    out = catalog.select([
        pl.col("catalog_key").cast(pl.Utf8, strict=False),
        (
            pl.col("external_release_date").cast(pl.Utf8, strict=False)
            if "external_release_date" in columns
            else pl.lit(None).cast(pl.Utf8)
        ).alias("Release date"),
        label_expr.alias("_label_normalized_catalog"),
        (
            pl.col("external_metadata_status").cast(pl.Utf8, strict=False)
            if "external_metadata_status" in columns
            else pl.lit(None).cast(pl.Utf8)
        ).alias("Metadata"),
        (
            pl.col("external_match_url").cast(pl.Utf8, strict=False)
            if "external_match_url" in columns
            else pl.lit(None).cast(pl.Utf8)
        ).alias("URL metadata"),
    ])
    overrides = catalog_label_overrides()
    if not overrides.is_empty():
        out = out.join(overrides, on="catalog_key", how="left")
    else:
        out = out.with_columns(pl.lit(None).cast(pl.Utf8).alias("_label_normalized_override"))
    return (
        out
        .with_columns(pl.coalesce(["_label_normalized_override", "_label_normalized_catalog"]).alias("Label normalizado"))
        .select(["catalog_key", "Release date", "Label normalizado", "Metadata", "URL metadata"])
        .unique(["catalog_key"])
    )


def catalog_metadata_by_isrc() -> pl.DataFrame:
    empty = pl.DataFrame({
        "ISRC": pl.Series([], dtype=pl.Utf8),
        "Release date": pl.Series([], dtype=pl.Utf8),
        "Label normalizado": pl.Series([], dtype=pl.Utf8),
        "Metadata": pl.Series([], dtype=pl.Utf8),
        "URL metadata": pl.Series([], dtype=pl.Utf8),
    })
    if not CATALOG_MASTER_PATH.exists():
        return empty
    catalog = pl.read_parquet(CATALOG_MASTER_PATH)
    if catalog.is_empty() or "isrcs" not in catalog.columns:
        return empty
    metadata = catalog_metadata_lookup()
    return (
        catalog
        .select(["catalog_key", "isrcs"])
        .with_columns(pl.col("isrcs").fill_null("").str.split(" | ").alias("ISRC"))
        .explode("ISRC")
        .with_columns(pl.col("ISRC").cast(pl.Utf8, strict=False).str.strip_chars().str.to_uppercase())
        .filter(pl.col("ISRC").is_not_null() & (pl.col("ISRC") != ""))
        .join(metadata, on="catalog_key", how="left")
        .select(["ISRC", "Release date", "Label normalizado", "Metadata", "URL metadata"])
        .unique(["ISRC"])
    )


def catalog_metadata_by_video_id() -> pl.DataFrame:
    empty = pl.DataFrame({
        "Video ID": pl.Series([], dtype=pl.Utf8),
        "Release date": pl.Series([], dtype=pl.Utf8),
        "Label normalizado": pl.Series([], dtype=pl.Utf8),
        "Metadata": pl.Series([], dtype=pl.Utf8),
        "URL metadata": pl.Series([], dtype=pl.Utf8),
    })
    if not CATALOG_MASTER_PATH.exists():
        return empty
    catalog = pl.read_parquet(CATALOG_MASTER_PATH)
    if catalog.is_empty() or "video_ids" not in catalog.columns:
        return empty
    metadata = catalog_metadata_lookup()
    return (
        catalog
        .select(["catalog_key", "video_ids"])
        .with_columns(pl.col("video_ids").fill_null("").str.split(" | ").alias("Video ID"))
        .explode("Video ID")
        .with_columns(pl.col("Video ID").cast(pl.Utf8, strict=False).str.strip_chars())
        .filter(pl.col("Video ID").is_not_null() & (pl.col("Video ID") != ""))
        .join(metadata, on="catalog_key", how="left")
        .select(["Video ID", "Release date", "Label normalizado", "Metadata", "URL metadata"])
        .unique(["Video ID"])
    )


def artist_catalog_identifiers() -> dict[str, set[str]]:
    empty = {"catalog_keys": set(), "isrcs": set(), "upcs": set(), "videos": set()}
    if not CATALOG_MASTER_PATH.exists():
        return empty
    catalog = pl.scan_parquet(CATALOG_MASTER_PATH)
    schema = catalog.collect_schema()
    text_cols = [
        col
        for col in [
            "identity_track_title",
            "identity_artist_statement",
            "track_title",
            "artist_statement",
            "title_variants",
            "artist_variants",
            "identity_artist_variants",
            "external_label",
            "label_normalized",
        ]
        if col in schema
    ]
    if not text_cols:
        return empty
    search = pl.any_horizontal([
        pl.col(col).cast(pl.Utf8, strict=False).str.to_lowercase().str.contains(LAALO_PATTERN).fill_null(False)
        for col in text_cols
    ])
    selected = (
        catalog
        .filter(search)
        .select([
            pl.col("catalog_key").cast(pl.Utf8, strict=False),
            pl.col("isrcs").cast(pl.Utf8, strict=False) if "isrcs" in schema else pl.lit(None).cast(pl.Utf8).alias("isrcs"),
            pl.col("upcs").cast(pl.Utf8, strict=False) if "upcs" in schema else pl.lit(None).cast(pl.Utf8).alias("upcs"),
            pl.col("video_ids").cast(pl.Utf8, strict=False) if "video_ids" in schema else pl.lit(None).cast(pl.Utf8).alias("video_ids"),
        ])
        .collect()
    )
    if selected.is_empty():
        return empty

    def split_values(column: str) -> set[str]:
        values: set[str] = set()
        for value in selected.get_column(column).drop_nulls().to_list():
            for item in str(value).split(" | "):
                clean = item.strip()
                if clean:
                    values.add(clean)
        return values

    return {
        "catalog_keys": set(selected.get_column("catalog_key").drop_nulls().to_list()),
        "isrcs": split_values("isrcs"),
        "upcs": split_values("upcs"),
        "videos": split_values("video_ids"),
    }


def classified_rows(end_month: str) -> pl.DataFrame:
    if not SONG_LEVEL_PATH.exists():
        raise FileNotFoundError(f"No existe {SONG_LEVEL_PATH}")

    lf = ensure_store_dimensions(pl.scan_parquet(SONG_LEVEL_PATH))
    schema = lf.collect_schema()
    schema_dict = {name: dtype for name, dtype in zip(schema.names(), schema.dtypes())}
    schema_names = set(schema.names())

    policy = load_statement_policy()
    catalog_ids = artist_catalog_identifiers()

    tx_month = col_or_null(schema_dict, "transaction_month")
    base = lf.filter(tx_month <= end_month).with_columns([
        col_or_null(schema_dict, "source_sheet").alias("_policy_sheet"),
        col_or_null(schema_dict, "source").alias("_source"),
        col_or_null(schema_dict, "account").alias("_account"),
        col_or_null(schema_dict, "source_sheet").alias("_source_sheet_raw"),
        col_or_null(schema_dict, "statement_type").alias("_statement_type"),
        tx_month.alias("_statement_period"),
        tx_month.alias("_transaction_month"),
        first_text(schema_dict, ["track_statement_style", "asset_title_statement"]).alias("_title"),
        first_text(schema_dict, ["artist_statement_style", "asset_artist_statement"]).alias("_artist"),
        first_text(schema_dict, ["asset_isrc"]).alias("_isrc"),
        pl.lit(None).cast(pl.Utf8).alias("_upc"),
        first_text(schema_dict, ["video_id", "Video ID", "VideoId"]).alias("_video_id"),
        col_or_null(schema_dict, "store_report_label").alias("_dsp_store"),
        col_or_null(schema_dict, "monetization_normalized").alias("_monetization"),
        col_or_null(schema_dict, "content_origin_normalized").alias("_content_origin"),
        col_or_null(schema_dict, "plan_normalized").alias("_plan"),
        pl.lit(None).cast(pl.Utf8).alias("_territory"),
        amount_expr(schema_dict).alias("_amount_usd"),
        units_expr(schema_dict).alias("_units"),
    ])

    id_match = pl.lit(False)
    if catalog_ids["isrcs"]:
        id_match = id_match | pl.col("_isrc").is_in(list(catalog_ids["isrcs"]))

    search_columns = [
        col
        for col in [
            "artist_statement_style",
            "asset_artist_statement",
            "track_statement_style",
            "asset_title_statement",
        ]
        if col in schema_names
    ]
    text_search = (
        pl.any_horizontal([
            pl.col(col).cast(pl.Utf8, strict=False).str.to_lowercase().str.contains(LAALO_PATTERN).fill_null(False)
            for col in search_columns
        ])
        if search_columns
        else pl.lit(False)
    )
    base = base.filter(id_match | text_search)

    with_status = with_catalog_report_status(base, schema_names)
    joined = (
        with_status
        .join(policy.lazy(), left_on=["_source", "_account", "_policy_sheet"], right_on=["source", "account", "_policy_sheet"], how="left")
        .with_columns([
            pl.col("_policy_statement_ok").fill_null(False),
            pl.col("_policy_revenue_basis").fill_null("sin_policy"),
            pl.col("_policy_display").fill_null(pl.concat_str([pl.col("_source"), pl.lit(" / "), pl.col("_account")])),
        ])
    )

    metadata = catalog_metadata_lookup()
    if not metadata.is_empty():
        joined = joined.join(metadata.lazy(), on="catalog_key", how="left")
    else:
        joined = joined.with_columns([
            pl.lit(None).cast(pl.Utf8).alias("Release date"),
            pl.lit(None).cast(pl.Utf8).alias("Label normalizado"),
            pl.lit(None).cast(pl.Utf8).alias("Metadata"),
            pl.lit(None).cast(pl.Utf8).alias("URL metadata"),
        ])

    return (
        joined
        .with_columns([
            (pl.col("include_in_reports") == True).alias("_catalog_ok"),
            ((pl.col("include_in_reports") == True) & (pl.col("_policy_statement_ok") == True)).alias("_business_ok"),
            pl.when(pl.col("_policy_statement_ok") == False)
            .then(pl.lit("excluido_policy"))
            .when(pl.col("include_in_reports") == False)
            .then(pl.lit("excluido_catalogo"))
            .otherwise(pl.lit("reportable"))
            .alias("_reason"),
        ])
        .collect()
    )


def maybe_nonzero(df: pl.DataFrame, hide_zero_amounts: bool) -> pl.DataFrame:
    if hide_zero_amounts and not df.is_empty() and "Ingresos USD" in df.columns:
        return df.filter(pl.col("Ingresos USD").abs() >= ZERO_EPSILON)
    return df


def summarize_assets(rows: pl.DataFrame, filter_expr: pl.Expr, *, hide_zero_amounts: bool = False) -> pd.DataFrame:
    selected = rows.filter(filter_expr)
    if selected.is_empty():
        return pd.DataFrame()
    grouped = (
        selected
        .group_by([
            "_source",
            "_account",
            "_policy_sheet",
            "_title",
            "_artist",
            "_isrc",
            "_upc",
            "_video_id",
            "Release date",
            "Label normalizado",
            "Metadata",
        ])
        .agg([
            pl.sum("_amount_usd").alias("Ingresos USD"),
            pl.sum("_units").alias("Unidades"),
            pl.len().alias("Filas"),
            pl.min("_transaction_month").alias("Desde transaction"),
            pl.max("_transaction_month").alias("Hasta transaction"),
            pl.min("_statement_period").alias("Desde statement"),
            pl.max("_statement_period").alias("Hasta statement"),
        ])
    )
    grouped = maybe_nonzero(grouped, hide_zero_amounts)
    if grouped.is_empty():
        return pd.DataFrame()
    return (
        grouped
        .rename({
            "_source": "Distribuidora",
            "_account": "Cuenta",
            "_policy_sheet": "Hoja origen",
            "_title": "Tema / video",
            "_artist": "Artista statement",
            "_isrc": "ISRC",
            "_upc": "UPC",
            "_video_id": "Video ID",
        })
        .sort("Ingresos USD", descending=True)
        .to_pandas()
    )


def granular_rows(rows: pl.DataFrame, *, hide_zero_amounts: bool = False) -> pd.DataFrame:
    selected = rows.filter(pl.col("_business_ok") == True)
    if selected.is_empty():
        return pd.DataFrame()
    grouped = (
        selected
        .group_by([
            "_source",
            "_account",
            "_policy_sheet",
            "_title",
            "_artist",
            "_isrc",
            "_upc",
            "_video_id",
            "Release date",
            "Label normalizado",
            "Metadata",
            "_territory",
            "_dsp_store",
            "_monetization",
            "_content_origin",
            "_plan",
        ])
        .agg([
            pl.sum("_amount_usd").alias("Ingresos USD"),
            pl.sum("_units").alias("Unidades"),
            pl.len().alias("Filas"),
            pl.min("_transaction_month").alias("Desde transaction"),
            pl.max("_transaction_month").alias("Hasta transaction"),
            pl.min("_statement_period").alias("Desde statement"),
            pl.max("_statement_period").alias("Hasta statement"),
        ])
    )
    grouped = maybe_nonzero(grouped, hide_zero_amounts)
    if grouped.is_empty():
        return pd.DataFrame()
    return (
        grouped
        .rename({
            "_source": "Distribuidora",
            "_account": "Cuenta",
            "_policy_sheet": "Hoja origen",
            "_title": "Tema / video",
            "_artist": "Artista statement",
            "_isrc": "ISRC",
            "_upc": "UPC",
            "_video_id": "Video ID",
            "_territory": "Pais",
            "_dsp_store": "DSP / Store",
            "_monetization": "Monetizacion",
            "_content_origin": "Origen contenido",
            "_plan": "Plan",
        })
        .sort("Ingresos USD", descending=True)
        .to_pandas()
    )


def raw_country_dsp_rows(
    end_month: str,
    reportable_assets: pd.DataFrame,
    *,
    hide_zero_amounts: bool = False,
) -> pd.DataFrame:
    if reportable_assets.empty or not RAW_ALL_PATH.exists():
        return pd.DataFrame()

    allowed_isrcs = {
        str(value).strip().upper()
        for value in reportable_assets.get("ISRC", pd.Series(dtype=object)).dropna().to_list()
        if str(value).strip()
    }
    lf = ensure_store_dimensions(pl.scan_parquet(RAW_ALL_PATH))
    schema = lf.collect_schema()
    schema_dict = {name: dtype for name, dtype in zip(schema.names(), schema.dtypes())}
    schema_names = set(schema.names())

    search_columns = [
        col
        for col in [
            "artist_statement_style",
            "artist_name_statement",
            "track_artist_statement",
            "asset_artist_statement",
            "Product Artist",
            "Asset Artist",
            "Track Artist",
            "TRACK ARTIST",
            "PRODUCT ARTIST",
            "track_statement_style",
            "asset_title_statement",
            "Track Title",
            "Title",
            "Asset Title",
            "Product Title",
            "Album Name",
        ]
        if col in schema_names
    ]
    text_search = (
        pl.any_horizontal([
            pl.col(col).cast(pl.Utf8, strict=False).str.to_lowercase().str.contains(LAALO_PATTERN).fill_null(False)
            for col in search_columns
        ])
        if search_columns
        else pl.lit(False)
    )

    raw_isrc = first_text(schema_dict, ["asset_isrc", "ISRC", "Asset ISRC"]).str.to_uppercase()
    id_search = raw_isrc.is_in(list(allowed_isrcs)) if allowed_isrcs else pl.lit(False)
    base = (
        lf
        .filter(col_or_null(schema_dict, "statement_period") <= end_month)
        .with_columns([
            source_sheet_policy_expr(schema_dict).alias("_policy_sheet"),
            col_or_null(schema_dict, "source").alias("_source"),
            col_or_null(schema_dict, "account").alias("_account"),
            first_text(schema_dict, ["track_statement_style", "asset_title_statement", "Track Title", "Title", "Asset Title", "Product Title", "Album Name"]).alias("_title"),
            first_text(schema_dict, ["asset_artist_statement", "artist_statement_style", "Track Artist", "Asset Artist", "Product Artist", "artist_name_statement", "track_artist_statement"]).alias("_artist"),
            raw_isrc.alias("ISRC"),
            first_text(schema_dict, ["product_upc", "UPC", "Product UPC"]).alias("UPC"),
            first_text(schema_dict, ["video_id", "Video ID", "VideoId", "ID", "Parent ID"]).alias("Video ID"),
            col_or_null(schema_dict, "store_report_label").alias("DSP / Store"),
            col_or_null(schema_dict, "dsp_normalized").alias("DSP normalizado"),
            col_or_null(schema_dict, "monetization_normalized").alias("Monetizacion"),
            col_or_null(schema_dict, "content_origin_normalized").alias("Origen contenido"),
            col_or_null(schema_dict, "plan_normalized").alias("Plan"),
            first_text(schema_dict, ["DSP", "Sale Store Name", "Store", "store_name", "Store Name"]).alias("Store original"),
            first_text(schema_dict, ["Territory", "territory", "Region", "SALE COUNTRY", "Sales Region", "Sales Country"]).alias("Pais"),
            amount_expr(schema_dict).alias("_amount_usd"),
            units_expr(schema_dict).alias("_units"),
            col_or_null(schema_dict, "transaction_month").alias("Desde transaction"),
            col_or_null(schema_dict, "statement_period").alias("Desde statement"),
        ])
        .filter(id_search | text_search)
    )
    policy = load_statement_policy()
    filtered = (
        base
        .join(policy.lazy(), left_on=["_source", "_account", "_policy_sheet"], right_on=["source", "account", "_policy_sheet"], how="left")
        .filter(pl.col("_policy_statement_ok") == True)
    )
    grouped = (
        filtered
        .group_by([
            "_source",
            "_account",
            "_policy_sheet",
            "_title",
            "_artist",
            "ISRC",
            "UPC",
            "Video ID",
            "Pais",
            "DSP / Store",
            "DSP normalizado",
            "Monetizacion",
            "Origen contenido",
            "Plan",
            "Store original",
        ])
        .agg([
            pl.sum("_amount_usd").alias("Ingresos USD"),
            pl.sum("_units").alias("Unidades"),
            pl.len().alias("Filas"),
            pl.min("Desde transaction").alias("Desde transaction"),
            pl.max("Desde transaction").alias("Hasta transaction"),
            pl.min("Desde statement").alias("Desde statement"),
            pl.max("Desde statement").alias("Hasta statement"),
        ])
        .collect()
    )
    grouped = maybe_nonzero(grouped, hide_zero_amounts)
    if grouped.is_empty():
        return pd.DataFrame()

    metadata_isrc = catalog_metadata_by_isrc()
    if not metadata_isrc.is_empty():
        grouped = grouped.join(metadata_isrc, on="ISRC", how="left")
    else:
        grouped = grouped.with_columns([
            pl.lit(None).cast(pl.Utf8).alias("Release date"),
            pl.lit(None).cast(pl.Utf8).alias("Label normalizado"),
            pl.lit(None).cast(pl.Utf8).alias("Metadata"),
            pl.lit(None).cast(pl.Utf8).alias("URL metadata"),
        ])

    metadata_video = catalog_metadata_by_video_id()
    if not metadata_video.is_empty():
        grouped = (
            grouped
            .join(
                metadata_video.rename({
                    "Release date": "_video_release_date",
                    "Label normalizado": "_video_label_normalizado",
                    "Metadata": "_video_metadata",
                    "URL metadata": "_video_url_metadata",
                }),
                on="Video ID",
                how="left",
            )
            .with_columns([
                pl.coalesce(["Release date", "_video_release_date"]).alias("Release date"),
                pl.coalesce(["Label normalizado", "_video_label_normalizado"]).alias("Label normalizado"),
                pl.coalesce(["Metadata", "_video_metadata"]).alias("Metadata"),
                pl.coalesce(["URL metadata", "_video_url_metadata"]).alias("URL metadata"),
            ])
            .drop(["_video_release_date", "_video_label_normalizado", "_video_metadata", "_video_url_metadata"], strict=False)
        )

    return (
        grouped
        .rename({
            "_source": "Distribuidora",
            "_account": "Cuenta",
            "_policy_sheet": "Hoja origen",
            "_title": "Tema / video",
            "_artist": "Artista statement",
        })
        .select([
            "Distribuidora",
            "Cuenta",
            "Hoja origen",
            "Tema / video",
            "Artista statement",
            "ISRC",
            "UPC",
            "Video ID",
            "Release date",
            "Label normalizado",
            "Pais",
            "DSP / Store",
            "DSP normalizado",
            "Monetizacion",
            "Origen contenido",
            "Plan",
            "Store original",
            "Ingresos USD",
            "Unidades",
            "Filas",
            "Desde transaction",
            "Hasta transaction",
            "Desde statement",
            "Hasta statement",
        ])
        .sort("Ingresos USD", descending=True)
        .to_pandas()
    )


def youtube_rows(rows: pl.DataFrame, *, hide_zero_amounts: bool = False) -> pd.DataFrame:
    selected = rows.filter(
        (pl.col("_business_ok") == True)
        & (
            pl.col("_dsp_store").cast(pl.Utf8, strict=False).str.to_lowercase().str.contains("youtube").fill_null(False)
            | pl.col("_policy_sheet").cast(pl.Utf8, strict=False).str.to_lowercase().str.contains("youtube").fill_null(False)
        )
    )
    if selected.is_empty():
        return pd.DataFrame()
    grouped = (
        selected
        .group_by([
            "_source",
            "_account",
            "_policy_sheet",
            "_title",
            "_artist",
            "_isrc",
            "_upc",
            "_video_id",
            "Release date",
            "Label normalizado",
            "URL metadata",
            "_dsp_store",
        ])
        .agg([
            pl.sum("_amount_usd").alias("Ingresos USD"),
            pl.sum("_units").alias("Unidades"),
            pl.len().alias("Filas"),
            pl.min("_transaction_month").alias("Primer mes monetizado"),
            pl.min("_statement_period").alias("Primer statement"),
            pl.max("_transaction_month").alias("Ultimo mes monetizado"),
            pl.max("_statement_period").alias("Ultimo statement"),
        ])
    )
    grouped = maybe_nonzero(grouped, hide_zero_amounts)
    if grouped.is_empty():
        return pd.DataFrame()
    out = (
        grouped
        .rename({
            "_source": "Distribuidora",
            "_account": "Cuenta",
            "_policy_sheet": "Hoja origen",
            "_title": "Tema / video",
            "_artist": "Artista statement",
            "_isrc": "ISRC",
            "_upc": "UPC",
            "_video_id": "Video ID",
            "_dsp_store": "DSP / Store",
        })
        .sort("Ingresos USD", descending=True)
        .to_pandas()
    )
    return out


def summary_tables(rows: pl.DataFrame, reportable_assets: pd.DataFrame, excluded_assets: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    blocks = [
        ("Ingresos reportables", reportable_assets),
        ("Excluidos catalogo / policy", excluded_assets),
    ]
    total_amount = sum(float(df["Ingresos USD"].sum()) for _, df in blocks if not df.empty and "Ingresos USD" in df.columns)
    summary = []
    by_source_frames = []
    for label, df in blocks:
        amount = float(df["Ingresos USD"].sum()) if not df.empty and "Ingresos USD" in df.columns else 0.0
        units = float(df["Unidades"].sum()) if not df.empty and "Unidades" in df.columns else 0.0
        count = int(df["Filas"].sum()) if not df.empty and "Filas" in df.columns else 0
        summary.append({
            "Bloque": label,
            "Ingresos USD": amount,
            "Unidades": units,
            "Filas": count,
            "% total": amount / total_amount if total_amount else 0.0,
        })
        if not df.empty:
            src = (
                df.groupby(["Distribuidora", "Cuenta", "Hoja origen"], dropna=False, as_index=False)
                .agg({"Ingresos USD": "sum", "Unidades": "sum", "Filas": "sum"})
            )
            src.insert(0, "Bloque", label)
            by_source_frames.append(src)
    summary.append({
        "Bloque": "Total control",
        "Ingresos USD": total_amount,
        "Unidades": sum(row["Unidades"] for row in summary),
        "Filas": sum(row["Filas"] for row in summary),
        "% total": 1.0 if total_amount else 0.0,
    })
    by_source = pd.concat(by_source_frames, ignore_index=True) if by_source_frames else pd.DataFrame()
    return pd.DataFrame(summary), by_source


def style_sheet(ws, *, fill: PatternFill | None = None) -> None:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    if fill:
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 500), max_col=ws.max_column):
            for cell in row:
                cell.fill = fill
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        max_len = 0
        for cell in column[:400]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 11), 48)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = LEFT
            if isinstance(cell.value, (int, float)):
                header = str(ws.cell(row=1, column=cell.column).value or "")
                if header in {"Filas", "Unidades"}:
                    cell.number_format = '#,##0'
                elif "USD" in header:
                    cell.number_format = '$#,##0.00'
                elif "%" in header:
                    cell.number_format = '0.00%'
                else:
                    cell.number_format = '#,##0.00'


def style_header_row(ws, row_number: int) -> None:
    for cell in ws[row_number]:
        if cell.value is None:
            continue
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def format_numeric_columns(ws, header_row: int, data_start_row: int) -> None:
    for row in ws.iter_rows(min_row=data_start_row):
        for cell in row:
            if not isinstance(cell.value, (int, float)):
                continue
            header = str(ws.cell(row=header_row, column=cell.column).value or "")
            if header in {"Filas", "Unidades"}:
                cell.number_format = '#,##0'
            elif "USD" in header:
                cell.number_format = '$#,##0.00'
            elif "%" in header:
                cell.number_format = '0.00%'
            else:
                cell.number_format = '#,##0.00'


def normalize_topic_key(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def join_unique(values: pd.Series) -> str:
    items: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = str(value or "").strip()
        if not text or text.lower() in {"nan", "none", "-"}:
            continue
        key = text.lower()
        if key in seen:
            continue
        seen.add(key)
        items.append(text)
    return " | ".join(items)


def first_non_empty_value(values: pd.Series) -> str:
    for value in values:
        text = str(value or "").strip()
        if text and text.lower() not in {"nan", "none", "-"}:
            return text
    return ""


def consolidated_by_title(reportable: pd.DataFrame) -> pd.DataFrame:
    if reportable.empty:
        return pd.DataFrame()

    rows = reportable.copy()

    def row_key(row: pd.Series) -> str:
        isrc = str(row.get("ISRC", "") or "").strip().upper()
        if isrc and isrc not in {"-", "NAN", "NONE"}:
            return f"ISRC:{isrc}"
        video_id = str(row.get("Video ID", "") or "").strip()
        if video_id and video_id.lower() not in {"-", "nan", "none"}:
            return f"VIDEO:{video_id}"
        return f"TITLE:{normalize_topic_key(row.get('Tema / video', ''))}"

    rows["_topic_key"] = rows.apply(row_key, axis=1)
    rows["_abs_usd"] = pd.to_numeric(rows.get("Ingresos USD", 0), errors="coerce").fillna(0).abs()

    output_rows: list[dict[str, object]] = []
    for _key, group in rows.groupby("_topic_key", dropna=False):
        ordered = group.sort_values("_abs_usd", ascending=False)
        main = ordered.iloc[0]
        amount = pd.to_numeric(group.get("Ingresos USD", 0), errors="coerce").fillna(0).sum()
        units = pd.to_numeric(group.get("Unidades", 0), errors="coerce").fillna(0).sum()
        row_count = pd.to_numeric(group.get("Filas", 0), errors="coerce").fillna(0).sum()
        output_rows.append({
            "Tema": main.get("Tema / video", ""),
            "Artista": first_non_empty_value(group.get("Artista statement", pd.Series(dtype=object))),
            "ISRCs": join_unique(group.get("ISRC", pd.Series(dtype=object))),
            "UPCs": join_unique(group.get("UPC", pd.Series(dtype=object))),
            "Video IDs": join_unique(group.get("Video ID", pd.Series(dtype=object))),
            "Release date": first_non_empty_value(group.get("Release date", pd.Series(dtype=object))),
            "Label normalizado": first_non_empty_value(group.get("Label normalizado", pd.Series(dtype=object))),
            "Distribuidoras": join_unique(group.get("Distribuidora", pd.Series(dtype=object))),
            "Cuentas": join_unique(group.get("Cuenta", pd.Series(dtype=object))),
            "Hojas origen": join_unique(group.get("Hoja origen", pd.Series(dtype=object))),
            "Ingresos USD": amount,
            "Unidades": units,
            "Filas": int(row_count),
            "Desde transaction": first_non_empty_value(group.get("Desde transaction", pd.Series(dtype=object)).sort_values()),
            "Hasta transaction": first_non_empty_value(group.get("Hasta transaction", pd.Series(dtype=object)).sort_values(ascending=False)),
            "Desde statement": first_non_empty_value(group.get("Desde statement", pd.Series(dtype=object)).sort_values()),
            "Hasta statement": first_non_empty_value(group.get("Hasta statement", pd.Series(dtype=object)).sort_values(ascending=False)),
        })

    out = pd.DataFrame(output_rows)
    if not out.empty:
        out = out.sort_values("Ingresos USD", ascending=False).reset_index(drop=True)
    return out


def write_report(rows: pl.DataFrame, end_month: str, *, hide_zero_amounts: bool = False) -> Path:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = REPORTS_DIR / f"juli_savioli_inicio_a_{end_month}_{stamp}.xlsx"

    reportable = summarize_assets(rows, pl.col("_business_ok") == True, hide_zero_amounts=hide_zero_amounts)
    excluded = summarize_assets(rows, pl.col("_business_ok") == False, hide_zero_amounts=hide_zero_amounts)
    title_summary = consolidated_by_title(reportable)
    granular = raw_country_dsp_rows(end_month, reportable, hide_zero_amounts=hide_zero_amounts)
    if granular.empty:
        granular = granular_rows(rows, hide_zero_amounts=hide_zero_amounts)
    if not granular.empty and "DSP / Store" in granular.columns:
        youtube = granular[
            granular["DSP / Store"]
            .fillna("")
            .astype(str)
            .str.contains("youtube", case=False, regex=False)
        ].copy()
        if not youtube.empty and "Desde transaction" in youtube.columns:
            key_cols = [
                col
                for col in ["Distribuidora", "Cuenta", "Hoja origen", "Tema / video", "Artista statement", "ISRC", "UPC", "Video ID"]
                if col in youtube.columns
            ]
            if key_cols:
                youtube["Primer mes monetizado"] = (
                    youtube
                    .groupby(key_cols, dropna=False)["Desde transaction"]
                    .transform("min")
                )
                insert_at = youtube.columns.get_loc("Desde transaction")
                col = youtube.pop("Primer mes monetizado")
                youtube.insert(insert_at, "Primer mes monetizado", col)
    else:
        youtube = youtube_rows(rows, hide_zero_amounts=hide_zero_amounts)
    summary, by_source = summary_tables(rows, reportable, excluded)

    source_sheets: list[tuple[str, pd.DataFrame, PatternFill]] = []
    if not reportable.empty:
        for source, label, fill in [
            ("fuga", "Ingresos FUGA", FUGA_FILL),
            ("onerpm", "Ingresos OneRPM", OK_FILL),
            ("orchard", "Ingresos Orchard", OK_FILL),
            ("dashgo", "Ingresos DashGo", OK_FILL),
            ("soundon", "Ingresos SoundOn", OK_FILL),
        ]:
            frame = reportable[reportable["Distribuidora"].fillna("").astype(str).str.lower() == source].copy()
            if not frame.empty:
                source_sheets.append((label, frame, fill))

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="Resumen", startrow=1)
        detail_title_startrow = len(summary) + 4
        if not by_source.empty:
            by_source.to_excel(writer, index=False, sheet_name="Resumen", startrow=detail_title_startrow + 1)
        reportable.to_excel(writer, index=False, sheet_name="Ingresos Todos")
        title_summary.to_excel(writer, index=False, sheet_name="Consolidado Temas")
        for sheet_name, frame, _fill in source_sheets:
            frame.to_excel(writer, index=False, sheet_name=sheet_name[:31])
        granular.to_excel(writer, index=False, sheet_name="Pais DSP")
        youtube.to_excel(writer, index=False, sheet_name="Analisis YouTube")
        excluded.to_excel(writer, index=False, sheet_name="Excluidos")

        ws = writer.book["Resumen"]
        ws["A1"] = f"Juli Savioli - inicio a {end_month}"
        ws["A1"].font = TITLE_FONT
        ws["A1"].fill = TOTAL_FILL
        if not by_source.empty:
            cell = ws.cell(row=detail_title_startrow + 1, column=1)
            cell.value = "Detalle por bloque / distribuidora / cuenta"
            cell.font = TITLE_FONT
            cell.fill = TOTAL_FILL
        style_sheet(writer.book["Resumen"], fill=TOTAL_FILL)
        style_header_row(writer.book["Resumen"], 2)
        format_numeric_columns(writer.book["Resumen"], 2, 3)
        if not by_source.empty:
            style_header_row(writer.book["Resumen"], detail_title_startrow + 2)
            format_numeric_columns(writer.book["Resumen"], detail_title_startrow + 2, detail_title_startrow + 3)

        style_sheet(writer.book["Ingresos Todos"], fill=TOTAL_FILL)
        style_sheet(writer.book["Consolidado Temas"], fill=TOTAL_FILL)
        for sheet_name, _frame, fill in source_sheets:
            style_sheet(writer.book[sheet_name[:31]], fill=fill)
        style_sheet(writer.book["Pais DSP"], fill=FUGA_FILL)
        style_sheet(writer.book["Analisis YouTube"], fill=FUGA_FILL)
        style_sheet(writer.book["Excluidos"], fill=EXCLUDED_FILL)

    return output


def main() -> None:
    parser = argparse.ArgumentParser(description="Reporte personalizado Juli Savioli.")
    parser.add_argument("--end-month", default="2026-03")
    parser.add_argument("--hide-zero-amounts", action="store_true")
    args = parser.parse_args()

    rows = classified_rows(args.end_month)
    output = write_report(rows, args.end_month, hide_zero_amounts=args.hide_zero_amounts)
    reportable = rows.filter(pl.col("_business_ok") == True)
    print("Reporte Juli Savioli generado")
    print(output)
    print(f"Filas reportables: {reportable.height}")
    print(f"USD reportable: {float(reportable['_amount_usd'].sum() or 0):.2f}")


if __name__ == "__main__":
    main()
