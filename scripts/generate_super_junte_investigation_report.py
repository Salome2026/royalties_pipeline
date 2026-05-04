from datetime import datetime
from pathlib import Path
import sys

import polars as pl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

BASE = Path(r"C:\royalties_pipeline")
SCRIPTS = BASE / "scripts"
MARTS = BASE / "warehouse" / "marts"
REPORTS = BASE / "reports"

if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from build_keyword_royalty_report import (  # noqa: E402
    SEARCH_COLUMNS_STANDARDIZED,
    build_filter,
    normalize_report_store,
    normalize_report_units,
    normalize_report_usage,
)


KEYWORDS = ["super junte"]
END_STATEMENT_PERIOD = "2025-03"
YOUTUBE_PUBLIC_VIEWS = 88_478_836
OUTPUT = REPORTS / "super_junte_deep_youtube_audit_hasta_2025-03.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(size=16, bold=True, color="1F4E78")
CENTER = Alignment(horizontal="center", vertical="center")


def amount_expr():
    return pl.col("amount_usd").cast(pl.Float64, strict=False)


def units_expr():
    return pl.col("units").cast(pl.Float64, strict=False)


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER


def set_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def append_frame(ws, frame, headers):
    ws.append(headers)
    for row in frame.itertuples(index=False):
        ws.append(list(row))
    style_header(ws, 1)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def load_super_junte():
    standardized = MARTS / "standardized_raw_all_sources.parquet"
    lf = pl.scan_parquet(standardized)
    cols = set(lf.collect_schema().names())
    keyword_filter = build_filter(cols, SEARCH_COLUMNS_STANDARDIZED, KEYWORDS, "any")

    enriched = normalize_report_usage(
        normalize_report_store(
            normalize_report_units(lf, cols),
            cols,
        ),
        cols,
    )

    return (
        enriched
        .filter(keyword_filter & pl.col("statement_period").is_not_null())
        .with_columns([
            pl.when(pl.col("usage_type").is_in(["video_stream", "video_ugc", "short_video"]))
            .then(pl.lit("video"))
            .when(pl.col("usage_type").is_in(["audio_stream", "download"]))
            .then(pl.lit("audio"))
            .otherwise(pl.lit("otros"))
            .alias("media_group"),
            pl.col("store_raw")
            .cast(pl.Utf8, strict=False)
            .str.to_lowercase()
            .str.contains("youtube")
            .fill_null(False)
            .alias("is_youtube_store"),
            pl.coalesce([
                pl.col("ISRC").cast(pl.Utf8, strict=False).str.strip_chars(),
                pl.col("asset_isrc").cast(pl.Utf8, strict=False).str.strip_chars(),
            ]).alias("report_isrc"),
            pl.coalesce([
                pl.col("Track Title").cast(pl.Utf8, strict=False).str.strip_chars(),
                pl.col("track_statement_style").cast(pl.Utf8, strict=False).str.strip_chars(),
                pl.col("Asset Title").cast(pl.Utf8, strict=False).str.strip_chars(),
                pl.col("Product Title").cast(pl.Utf8, strict=False).str.strip_chars(),
                pl.col("asset_title_statement").cast(pl.Utf8, strict=False).str.strip_chars(),
            ]).alias("report_title"),
            pl.coalesce([
                pl.col("artist_statement_style").cast(pl.Utf8, strict=False).str.strip_chars(),
                pl.col("Artist Name").cast(pl.Utf8, strict=False).str.strip_chars(),
                pl.col("Product Artist").cast(pl.Utf8, strict=False).str.strip_chars(),
                pl.col("artists_raw").cast(pl.Utf8, strict=False).str.strip_chars(),
            ]).alias("report_artist"),
        ])
    )


def aggregate_totals(data):
    return data.select([
        pl.len().alias("Filas"),
        amount_expr().sum().alias("Ingresos USD"),
        units_expr().sum().alias("Unidades"),
    ]).collect().to_dicts()[0]


def main() -> Path:
    all_data = load_super_junte()
    cutoff_data = all_data.filter(pl.col("statement_period") <= END_STATEMENT_PERIOD)
    video_youtube = cutoff_data.filter((pl.col("media_group") == "video") & pl.col("is_youtube_store"))

    wb = Workbook()
    ws = wb.active
    ws.title = "dashboard"

    ws["A1"] = "Super Junte - auditoria YouTube por statement"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = "Selector grafico"
    ws["B3"] = "video_youtube"
    ws["A4"] = "Rango principal"
    ws["B4"] = f"Inicio a {END_STATEMENT_PERIOD} inclusive"
    ws["A5"] = "Comparador YouTube"
    ws["B5"] = YOUTUBE_PUBLIC_VIEWS
    ws["A6"] = "Nota"
    ws["B6"] = "El comparador es el contador publico actual del video; no es mensual."

    validation = DataValidation(type="list", formula1='"video_youtube,video_total,audio_total,todos"', allow_blank=False)
    ws.add_data_validation(validation)
    validation.add(ws["B3"])

    cutoff_totals = aggregate_totals(cutoff_data)
    vy_totals = aggregate_totals(video_youtube)
    all_vy_totals = aggregate_totals(
        all_data.filter((pl.col("media_group") == "video") & pl.col("is_youtube_store"))
    )
    zero_unit_revenue_totals = aggregate_totals(
        video_youtube.filter(
            (units_expr().fill_null(0) == 0)
            & (amount_expr().fill_null(0) > 0)
        )
    )

    summary_rows = [
        ["Todos hasta 2025-03", cutoff_totals["Ingresos USD"], cutoff_totals["Unidades"], cutoff_totals["Filas"], ""],
        ["Video YouTube hasta 2025-03", vy_totals["Ingresos USD"], vy_totals["Unidades"], vy_totals["Filas"], YOUTUBE_PUBLIC_VIEWS - vy_totals["Unidades"]],
        ["Video YouTube todo reportado", all_vy_totals["Ingresos USD"], all_vy_totals["Unidades"], all_vy_totals["Filas"], YOUTUBE_PUBLIC_VIEWS - all_vy_totals["Unidades"]],
    ]
    ws.append([])
    ws.append(["Corte", "Ingresos USD", "Unidades", "Filas", "Brecha vs 88.478.836"])
    for row in summary_rows:
        ws.append(row)
    style_header(ws, 8)
    ws["A12"] = "Video YouTube con ingresos y unidades 0"
    ws["B12"] = zero_unit_revenue_totals["Ingresos USD"]
    ws["C12"] = zero_unit_revenue_totals["Filas"]
    ws["D12"] = "Revisar hoja zero_unit_revenue"

    monthly = (
        cutoff_data
        .group_by(["statement_period", "media_group", "is_youtube_store"])
        .agg([
            amount_expr().sum().alias("amount_usd"),
            units_expr().sum().alias("units"),
            pl.len().alias("rows"),
        ])
        .collect()
        .to_pandas()
    )
    periods = sorted(monthly["statement_period"].dropna().unique().tolist())

    def metric(period, group=None, youtube=None, column="units"):
        frame = monthly[monthly["statement_period"] == period]
        if group is not None:
            frame = frame[frame["media_group"] == group]
        if youtube is not None:
            frame = frame[frame["is_youtube_store"] == youtube]
        if frame.empty:
            return 0
        return float(frame[column].sum())

    header_row = 14
    headers = [
        "Statement",
        "Video YouTube USD",
        "Video YouTube unidades",
        "Video total USD",
        "Video total unidades",
        "Audio total USD",
        "Audio total unidades",
        "Todos USD",
        "Todos unidades",
        "USD seleccionado",
        "Unidades seleccionadas",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    for row_idx, period in enumerate(periods, start=header_row + 1):
        ws.cell(row_idx, 1).value = period
        ws.cell(row_idx, 2).value = metric(period, "video", True, "amount_usd")
        ws.cell(row_idx, 3).value = metric(period, "video", True, "units")
        ws.cell(row_idx, 4).value = metric(period, "video", None, "amount_usd")
        ws.cell(row_idx, 5).value = metric(period, "video", None, "units")
        ws.cell(row_idx, 6).value = metric(period, "audio", None, "amount_usd")
        ws.cell(row_idx, 7).value = metric(period, "audio", None, "units")
        ws.cell(row_idx, 8).value = metric(period, None, None, "amount_usd")
        ws.cell(row_idx, 9).value = metric(period, None, None, "units")
        ws.cell(row_idx, 10).value = (
            f'=IF($B$3="video_youtube",B{row_idx},IF($B$3="video_total",D{row_idx},'
            f'IF($B$3="audio_total",F{row_idx},H{row_idx})))'
        )
        ws.cell(row_idx, 11).value = (
            f'=IF($B$3="video_youtube",C{row_idx},IF($B$3="video_total",E{row_idx},'
            f'IF($B$3="audio_total",G{row_idx},I{row_idx})))'
        )

    last_row = header_row + len(periods)
    for row in range(header_row + 1, last_row + 1):
        for col in [2, 4, 6, 8, 10]:
            ws.cell(row, col).number_format = '#,##0.00'
        for col in [3, 5, 7, 9, 11]:
            ws.cell(row, col).number_format = '#,##0'
    for row in range(9, 12):
        ws.cell(row, 2).number_format = '#,##0.00'
        ws.cell(row, 3).number_format = '#,##0'
        ws.cell(row, 5).number_format = '#,##0'
    ws["B12"].number_format = '#,##0.00'

    set_widths(ws, [16, 16, 20, 16, 20, 16, 20, 14, 18, 18, 22, 2])
    ws.freeze_panes = "A15"

    units_chart = BarChart()
    units_chart.title = "Unidades por statement"
    units_chart.y_axis.title = "Unidades"
    units_chart.x_axis.title = "Statement"
    units_chart.add_data(Reference(ws, min_col=11, min_row=header_row, max_row=last_row), titles_from_data=True)
    units_chart.set_categories(Reference(ws, min_col=1, min_row=header_row + 1, max_row=last_row))
    units_chart.height = 8
    units_chart.width = 21
    ws.add_chart(units_chart, "M3")

    usd_chart = BarChart()
    usd_chart.title = "Ingresos USD por statement"
    usd_chart.y_axis.title = "USD"
    usd_chart.x_axis.title = "Statement"
    usd_chart.add_data(Reference(ws, min_col=10, min_row=header_row, max_row=last_row), titles_from_data=True)
    usd_chart.set_categories(Reference(ws, min_col=1, min_row=header_row + 1, max_row=last_row))
    usd_chart.height = 8
    usd_chart.width = 21
    ws.add_chart(usd_chart, "M20")

    # Sheets for investigation
    store_type = (
        cutoff_data
        .group_by(["source", "account", "store_raw", "usage_type", "media_group"])
        .agg([amount_expr().sum().alias("Ingresos USD"), units_expr().sum().alias("Unidades"), pl.len().alias("Filas")])
        .sort("Unidades", descending=True)
        .collect()
        .to_pandas()
    )
    ws_store = wb.create_sheet("store_tipo")
    append_frame(ws_store, store_type, ["Fuente", "Cuenta", "Store original", "Tipo", "Grupo", "Ingresos USD", "Unidades", "Filas"])
    set_widths(ws_store, [12, 18, 28, 16, 10, 14, 16, 10])

    assets = (
        cutoff_data
        .group_by(["source", "account", "report_isrc", "report_title", "report_artist", "store_raw", "usage_type", "media_group"])
        .agg([amount_expr().sum().alias("Ingresos USD"), units_expr().sum().alias("Unidades"), pl.len().alias("Filas")])
        .sort("Unidades", descending=True)
        .collect()
        .to_pandas()
    )
    ws_assets = wb.create_sheet("assets")
    append_frame(ws_assets, assets, ["Fuente", "Cuenta", "ISRC", "Titulo", "Artista", "Store original", "Tipo", "Grupo", "Ingresos USD", "Unidades", "Filas"])
    set_widths(ws_assets, [12, 18, 14, 38, 28, 24, 16, 10, 14, 16, 10])

    month_store = (
        cutoff_data
        .group_by(["statement_period", "source", "account", "store_raw", "usage_type", "media_group"])
        .agg([amount_expr().sum().alias("Ingresos USD"), units_expr().sum().alias("Unidades"), pl.len().alias("Filas")])
        .sort(["statement_period", "Unidades"], descending=[False, True])
        .collect()
        .to_pandas()
    )
    ws_month_store = wb.create_sheet("statement_store")
    append_frame(ws_month_store, month_store, ["Statement", "Fuente", "Cuenta", "Store original", "Tipo", "Grupo", "Ingresos USD", "Unidades", "Filas"])
    set_widths(ws_month_store, [12, 12, 18, 28, 16, 10, 14, 16, 10])

    transaction_store = (
        cutoff_data
        .group_by(["transaction_month", "source", "account", "store_raw", "usage_type", "media_group"])
        .agg([
            amount_expr().sum().alias("Ingresos USD"),
            units_expr().sum().alias("Unidades"),
            pl.len().alias("Filas"),
            pl.min("statement_period").alias("Primer statement"),
            pl.max("statement_period").alias("Ultimo statement"),
        ])
        .sort(["transaction_month", "Unidades"], descending=[False, True])
        .collect()
        .to_pandas()
    )
    ws_transaction_store = wb.create_sheet("transaction_store")
    append_frame(
        ws_transaction_store,
        transaction_store,
        [
            "Mes consumo",
            "Fuente",
            "Cuenta",
            "Store original",
            "Tipo",
            "Grupo",
            "Ingresos USD",
            "Unidades",
            "Filas",
            "Primer statement",
            "Ultimo statement",
        ],
    )
    set_widths(ws_transaction_store, [12, 12, 18, 28, 16, 10, 14, 16, 10, 16, 16])

    zero_unit_revenue = (
        video_youtube
        .filter((units_expr().fill_null(0) == 0) & (amount_expr().fill_null(0) > 0))
        .group_by([
            "statement_period",
            "transaction_month",
            "source",
            "account",
            "store_raw",
            "usage_type",
            "report_isrc",
            "report_title",
        ])
        .agg([
            amount_expr().sum().alias("Ingresos USD"),
            pl.len().alias("Filas"),
            pl.min("statement_file_name").alias("Archivo ejemplo"),
        ])
        .sort(["statement_period", "transaction_month", "Ingresos USD"], descending=[False, False, True])
        .collect()
        .to_pandas()
    )
    ws_zero_units = wb.create_sheet("zero_unit_revenue")
    append_frame(
        ws_zero_units,
        zero_unit_revenue,
        [
            "Statement",
            "Mes consumo",
            "Fuente",
            "Cuenta",
            "Store original",
            "Tipo",
            "ISRC",
            "Titulo",
            "Ingresos USD",
            "Filas",
            "Archivo ejemplo",
        ],
    )
    set_widths(ws_zero_units, [12, 12, 12, 18, 28, 16, 14, 36, 14, 10, 34])

    month_assets = (
        cutoff_data
        .filter((pl.col("media_group") == "video") & pl.col("is_youtube_store"))
        .group_by(["statement_period", "report_isrc", "report_title", "store_raw", "usage_type"])
        .agg([amount_expr().sum().alias("Ingresos USD"), units_expr().sum().alias("Unidades"), pl.len().alias("Filas")])
        .sort(["statement_period", "Unidades"], descending=[False, True])
        .collect()
        .to_pandas()
    )
    ws_month_assets = wb.create_sheet("video_youtube_assets")
    append_frame(ws_month_assets, month_assets, ["Statement", "ISRC", "Titulo", "Store original", "Tipo", "Ingresos USD", "Unidades", "Filas"])
    set_widths(ws_month_assets, [12, 14, 38, 24, 16, 14, 16, 10])

    REPORTS.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(OUTPUT)
        return OUTPUT
    except PermissionError:
        fallback = OUTPUT.with_name(
            f"{OUTPUT.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{OUTPUT.suffix}"
        )
        wb.save(fallback)
        return fallback


if __name__ == "__main__":
    output = main()
    print(output)
