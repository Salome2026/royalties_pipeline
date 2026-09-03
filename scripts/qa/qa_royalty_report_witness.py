from __future__ import annotations

import argparse
import os
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook


BASE = Path(__file__).resolve().parents[2]
SCRIPTS = BASE / "scripts"
for path in (BASE, SCRIPTS):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

from app.operational_db import operational_connect  # noqa: E402
from app.royalty_reports.engine import ReportEngine, ReportRuntime  # noqa: E402
from app.royalty_reports.contracts import StoredArtifact  # noqa: E402
from build_keyword_royalty_report import normalize_keywords  # noqa: E402
from lib.distributor_policy_store import load_distributor_policy_document  # noqa: E402


SONG_FILE = "song_level_all_sources.parquet"
STANDARDIZED_FILE = "standardized_raw_all_sources.parquet"
CATALOG_MASTER_FILE = "catalog_master.parquet"


def read_job(job_id: int) -> dict:
    with operational_connect() as conn:
        row = conn.execute(
            "SELECT id, report_key, output_format, params_json FROM report_runs WHERE id = %s",
            (job_id,),
        ).fetchone()
    if row is None:
        raise RuntimeError(f"No existe report_run {job_id}.")
    return {
        "id": int(row["id"]),
        "report_key": row["report_key"],
        "output_format": row["output_format"],
        "params": dict(row["params_json"] or {}),
    }


def workbook_signature(
    path: Path,
    *,
    ignore_generated_at: bool = False,
    float_decimals: int | None = None,
) -> dict:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        sheets = []
        for sheet in workbook.worksheets:
            generated_at_columns = {
                cell.column
                for cell in sheet[1]
                if ignore_generated_at and cell.value == "Generado el"
            }
            cells = []
            for row in sheet.iter_rows():
                cells.append(
                    [
                        (
                            "<generated-at>"
                            if cell.column in generated_at_columns and cell.row > 1
                            else (
                                round(cell.value, float_decimals)
                                if float_decimals is not None
                                and isinstance(cell.value, float)
                                else cell.value
                            ),
                            cell.number_format,
                            cell.style_id,
                        )
                        for cell in row
                    ]
                )
            widths = {
                key: value.width
                for key, value in sheet.column_dimensions.items()
                if value.width is not None
            }
            sheets.append(
                {
                    "title": sheet.title,
                    "state": sheet.sheet_state,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "freeze_panes": str(sheet.freeze_panes or ""),
                    "auto_filter": str(sheet.auto_filter.ref or ""),
                    "merged": sorted(str(value) for value in sheet.merged_cells.ranges),
                    "widths": widths,
                    "cells": cells,
                }
            )
        return {"sheets": sheets}
    finally:
        workbook.close()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--job-id", type=int, required=True)
    parser.add_argument("--witness", type=Path, required=True)
    args = parser.parse_args()

    witness = args.witness.resolve()
    if not witness.exists():
        raise RuntimeError(f"No existe el testigo: {witness}")
    mart_dir = Path(os.environ.get("VPO_LOCAL_MARTS_DIR") or BASE / "warehouse" / "marts")
    catalog_status = BASE / "warehouse" / "registry" / "catalog_status.parquet"
    job = read_job(args.job_id)
    job["policy_snapshot"] = load_distributor_policy_document()

    with tempfile.TemporaryDirectory(prefix="vpo-report-witness-") as raw_dir:
        output_dir = Path(raw_dir)
        comparison: dict[str, bool] = {}

        def resolve_marts(manifest: dict, filenames: list[str]) -> dict[str, Path]:
            paths = {name: mart_dir / name for name in filenames}
            missing = [str(path) for path in paths.values() if not path.exists()]
            if missing:
                raise RuntimeError(f"Faltan marts para QA: {', '.join(missing)}")
            return paths

        def configure_catalog(marts: dict[str, Path]) -> None:
            os.environ["VPO_CATALOG_MASTER_PATH"] = str(marts[CATALOG_MASTER_FILE])
            os.environ["VPO_CATALOG_STATUS_PATH"] = str(catalog_status)

        def compare_artifact(
            job_id: int,
            generated: Path,
            content_type: str,
        ) -> StoredArtifact:
            expected = workbook_signature(witness)
            actual = workbook_signature(generated)
            comparison["equal"] = actual == expected
            if not comparison["equal"]:
                raise AssertionError(
                    f"El reporte regenerado para job {job_id} no coincide con el testigo."
                )
            return StoredArtifact(
                uri=f"qa://report/{job_id}/{generated.name}",
                size_bytes=generated.stat().st_size,
                sha256="0" * 64,
            )

        runtime = ReportRuntime(
            song_filename=SONG_FILE,
            standardized_filename=STANDARDIZED_FILE,
            catalog_master_filename=CATALOG_MASTER_FILE,
            output_dir=output_dir,
            resolve_marts=resolve_marts,
            configure_catalog_environment=configure_catalog,
            normalize_keywords=normalize_keywords,
            create_google_sheet=lambda *values: (_ for _ in ()).throw(
                RuntimeError("Este QA valida artefactos Excel.")
            ),
            upload_artifact=compare_artifact,
            report_stage=lambda job_id, stage: print(f"job={job_id} stage={stage}", flush=True),
        )
        result = ReportEngine(runtime).build(job)
        if not comparison.get("equal"):
            raise AssertionError("No se comparo el artefacto generado.")
        print(
            f"OK: job {args.job_id} coincide en datos, estructura y estilos con {witness.name}."
        )
        print(f"Resultado QA: {result.output_uri}")


if __name__ == "__main__":
    main()
